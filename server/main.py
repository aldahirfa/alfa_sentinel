import os

from fastapi import FastAPI, HTTPException, Header, Request, Depends, Query
from fastapi.responses import RedirectResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from fastapi.middleware.cors import CORSMiddleware
from starlette.middleware.sessions import SessionMiddleware
from pydantic import BaseModel
from dotenv import load_dotenv

from database import get_connection
from security import verify_password, hash_password

import secrets
import hashlib
import csv
import io
from datetime import datetime, timedelta
from urllib.parse import urlencode

from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
)
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

load_dotenv()

app = FastAPI()

# SessionMiddleware es lo que nos da la "pulsera de concierto": firma
# una cookie con itsdangerous para que el navegador la pueda guardar,
# y el servidor pueda confiar en ella sin volver a pedir contraseña
# en cada clic. SESSION_SECRET tiene que ser secreto y estable -- si
# cambia, todas las sesiones abiertas se invalidan de golpe.
SESSION_SECRET = os.getenv("SESSION_SECRET", "cambia-esto-en-produccion")

app.add_middleware(SessionMiddleware, secret_key=SESSION_SECRET)

# CORS solo para el nuevo frontend React (Vite corre en otro puerto,
# distinto origen aunque sea el mismo host). allow_credentials=True es
# necesario porque la sesión viaja como cookie -- sin esto, el navegador
# no la manda entre orígenes distintos. La lista es explícita (no "*")
# porque allow_credentials=True no funciona junto con un wildcard.
CORS_ORIGINS = os.getenv(
    "CORS_ORIGINS",
    "http://localhost:5173,http://127.0.0.1:5173"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=CORS_ORIGINS,
    allow_credentials=True,
    allow_methods=["GET", "POST", "PATCH", "DELETE"],
    allow_headers=["*"],
)

def time_ago(value):
    """'hace X minutos' -- lo usa el Panel de Control y varias APIs
    JSON (alertas recientes, actividad reciente, honeyfiles, endpoints)
    en vez de mostrar la fecha completa cada vez."""

    if value is None:
        return "—"

    seconds = int((datetime.now(value.tzinfo) - value).total_seconds())

    if seconds < 0:
        seconds = 0

    if seconds < 60:
        return f"hace {seconds}s"

    minutes = seconds // 60

    if minutes < 60:
        return f"hace {minutes} min"

    hours = minutes // 60

    if hours < 24:
        return f"hace {hours} h"

    days = hours // 24

    return f"hace {days} d"


# Umbral para considerar que un agente ONLINE está "al día": si su
# último heartbeat es más viejo que esto, lo marcamos como "requiere
# atención" en vez de "funcionando", aunque la BD todavía diga
# ONLINE. Es una aproximación -- hoy el agente manda heartbeat solo
# una vez al arrancar (no hay bucle automático todavía), así que este
# umbral es generoso a propósito.
#
# Este valor era una constante fija hasta el 2026-08-12; ahora vive en
# 'system_settings' (key 'agent_stale_seconds') y es editable de
# verdad desde /configuracion -- get_agent_stale_seconds() lo lee de
# la base en cada request. Esta constante queda solo como default de
# emergencia si la fila no existiera todavía (base vieja sin migrar).
AGENT_STALE_SECONDS_DEFAULT = 120


def get_system_setting(cursor, key, default=None, cast=str):
    """Lee un valor de 'system_settings'. Devuelve 'default' si la
    fila no existe o si el valor guardado no se puede convertir con
    'cast' -- nunca revienta la página por un dato de configuración
    mal cargado a mano."""

    cursor.execute("SELECT value FROM system_settings WHERE key = %s;", (key,))
    row = cursor.fetchone()

    if row is None:
        return default

    try:
        return cast(row[0])
    except (TypeError, ValueError):
        return default


def get_agent_stale_seconds(cursor):
    """Reemplaza a la vieja constante AGENT_STALE_SECONDS -- único
    parámetro de 'Configuración > Agentes' que el servidor realmente
    vuelve a leer después de guardarlo."""

    return get_system_setting(cursor, "agent_stale_seconds", default=AGENT_STALE_SECONDS_DEFAULT, cast=int)


def log_audit(cursor, user_id, action, entity_type=None, entity_id=None, description=None):
    """Bitácora de auditoría real (2026-08-12) -- antes 'audit_logs'
    existía en el schema pero ningún endpoint escribía ahí. Se llama
    justo antes del commit() de cada mutación relevante, con el mismo
    cursor y la misma transacción que el cambio que registra: si el
    cambio se revierte por un error después, el log también (no queda
    una entrada de auditoría de algo que en realidad no se guardó).
    'ip_address' se deja NULL -- no todos los endpoints que mutan
    datos tienen 'Request' en su firma hoy, y agregarlo a todos para
    esto solo hubiera agrandado el alcance de este cambio."""

    cursor.execute(
        """
        INSERT INTO audit_logs (user_id, action, entity_type, entity_id, description)
        VALUES (%s, %s, %s, %s, %s);
        """,
        (user_id, action, entity_type, entity_id, description)
    )


# Sirve server/static/* en /static/* -- ahí vive el logo (logo-icon.png)
# que usa el frontend React (LoginGate.tsx, Sidebar.tsx).
app.mount("/static", StaticFiles(directory="static"), name="static")


class AgentCreate(BaseModel):
    hostname: str
    os: str
    os_version: str | None = None
    ip_address: str | None = None
    agent_version: str | None = None

class EnrollmentRequest(BaseModel):
    token: str
    hostname: str
    os: str
    os_version: str | None = None
    ip_address: str | None = None
    agent_version: str | None = None


class EventCreate(BaseModel):
    event_type: str
    description: str | None = None
    process_id: int | None = None
    process_name: str | None = None
    metadata: dict | None = None


class AlertCreate(BaseModel):
    """Contrato reescrito 2026-08-16 junto con el motor heurístico
    definitivo (ver PENDIENTES.md): el agente ya NO decide severidad
    ni risk_score -- solo reporta qué reglas detectó activas
    (matched_rules, nombres de 'heuristic_rules.name'). El servidor
    (POST /agent/alerts, report_alert) es quien calcula peso,
    correlación, score final y severidad. 'severity'/'risk_score'/
    'rule_name' (contrato viejo) se eliminan -- no quedaba ningún
    agente en uso con el contrato anterior que hubiera que soportar en
    paralelo."""
    title: str
    description: str | None = None
    matched_rules: list[str] = []


class LoginRequest(BaseModel):
    username: str
    password: str


class UserCreate(BaseModel):
    username: str
    password: str
    full_name: str
    email: str | None = None
    role: str = "admin"


class IncidentCreate(BaseModel):
    alert_id: int
    classification: str | None = None


class AlertStatusUpdate(BaseModel):
    status: str


class IncidentAlertLink(BaseModel):
    alert_id: int


class IncidentAssign(BaseModel):
    user_id: int | None = None


class IncidentStatusUpdate(BaseModel):
    status: str


class IncidentClassify(BaseModel):
    classification: str


class IncidentDescriptionUpdate(BaseModel):
    description: str


class RuleUpdate(BaseModel):
    weight: float | None = None
    is_active: bool | None = None
    threshold: float | None = None
    window_seconds: int | None = None


class AgentRuleUpdate(BaseModel):
    """Override de una regla para UN endpoint puntual (tabla
    'agent_rule', 2026-08-16 -- ver PENDIENTES.md). A diferencia de
    RuleUpdate (heuristic_rules, campos NOT NULL), acá un campo en
    'null' es una instrucción válida y distinta de "no mandar el
    campo": 'null' significa "volver a heredar el valor global para
    ESE campo puntual" (sección 4 de la especificación -- no hay que
    repetir todos los parámetros para personalizar uno solo). Por eso
    el endpoint no puede usar "if payload.x is not None" como
    RuleUpdate -- necesita distinguir "campo ausente" de "campo
    presente con valor null", y para eso usa payload.model_fields_set
    en vez de mirar los valores solos."""
    threshold: float | None = None
    window_seconds: int | None = None
    weight: float | None = None
    is_active: bool | None = None


class ReportGenerate(BaseModel):
    report_type: str
    period: str
    format: str
    endpoint_id: int | None = None


class SettingUpdate(BaseModel):
    value: str


class UserUpdate(BaseModel):
    full_name: str | None = None
    email: str | None = None
    is_active: bool | None = None
    role: str | None = None


class ProfileUpdate(BaseModel):
    full_name: str
    email: str | None = None


class PasswordChange(BaseModel):
    current_password: str
    new_password: str


def resolve_agent_id(cursor, x_agent_credential: str) -> int:
    """Traduce la credencial que manda el agente en el header a su
    agent_id. Centraliza lo que antes estaba repetido en cada endpoint
    que necesita saber "qué agente me está hablando"."""

    credential_hash = hashlib.sha256(
        x_agent_credential.encode()
    ).hexdigest()

    cursor.execute(
        """
        SELECT agent_id
        FROM agent_credentials
        WHERE credential_hash = %s
          AND status = 'ACTIVE';
        """,
        (credential_hash,)
    )

    result = cursor.fetchone()

    if result is None:
        raise HTTPException(
            status_code=401,
            detail="Credencial inválida"
        )

    return result[0]


def get_current_user(request: Request) -> dict:
    """El 'portero': lee la sesión (la pulsera) y devuelve quién es la
    persona logueada. Si no hay sesión válida, corta con 401 antes de
    que la ruta protegida llegue a ejecutarse."""

    user = request.session.get("user")

    if user is None:
        raise HTTPException(
            status_code=401,
            detail="No has iniciado sesión"
        )

    return user


def require_role(role_name: str):
    """Fábrica de dependencias: require_role('admin') exige sesión
    válida Y que el rol coincida. Se usa así en una ruta:
        Depends(require_role('admin'))
    """

    def dependency(user: dict = Depends(get_current_user)) -> dict:

        if role_name not in user.get("roles", []):
            raise HTTPException(
                status_code=403,
                detail=f"Se requiere el rol '{role_name}'"
            )

        return user

    return dependency


@app.post("/login")
def login(credentials: LoginRequest, request: Request):

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT id, password_hash, full_name, is_active
                FROM users
                WHERE username = %s;
                """,
                (credentials.username,)
            )

            row = cursor.fetchone()

            # Mismo mensaje de error tanto si el usuario no existe
            # como si la contraseña está mal -- así no le confirmamos
            # a quien intenta entrar si el usuario es válido o no.
            if row is None:
                raise HTTPException(
                    status_code=401,
                    detail="Usuario o contraseña incorrectos"
                )

            user_id, password_hash, full_name, is_active = row

            if not is_active:
                raise HTTPException(
                    status_code=401,
                    detail="Usuario deshabilitado"
                )

            if not verify_password(credentials.password, password_hash):
                raise HTTPException(
                    status_code=401,
                    detail="Usuario o contraseña incorrectos"
                )

            cursor.execute(
                """
                SELECT roles.name
                FROM roles
                JOIN user_roles ON user_roles.role_id = roles.id
                WHERE user_roles.user_id = %s;
                """,
                (user_id,)
            )

            roles = [role_row[0] for role_row in cursor.fetchall()]

            cursor.execute(
                """
                UPDATE users
                SET last_login_at = CURRENT_TIMESTAMP
                WHERE id = %s;
                """,
                (user_id,)
            )

            connection.commit()

        request.session["user"] = {
            "id": user_id,
            "username": credentials.username,
            "full_name": full_name,
            "roles": roles
        }

        return {
            "message": "Sesión iniciada",
            "username": credentials.username,
            "roles": roles
        }

    finally:
        connection.close()


@app.post("/logout")
def logout(request: Request):

    request.session.clear()

    return {
        "message": "Sesión cerrada"
    }


@app.get("/me")
def me(user: dict = Depends(get_current_user)):
    """Para probar rápido si la sesión está activa y qué rol tiene."""

    return user


@app.put("/me")
def update_profile(
    profile: ProfileUpdate,
    request: Request,
    user: dict = Depends(get_current_user)
):
    """Cada persona edita su propio nombre/correo -- no hay 'user_id'
    en el body, siempre se actualiza a quien está logueado (sale de
    la sesión), para que nadie pueda editar el perfil de otro solo
    cambiando un número."""

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            if profile.email:

                cursor.execute(
                    "SELECT id FROM users WHERE email = %s AND id != %s;",
                    (profile.email, user["id"])
                )

                if cursor.fetchone():
                    raise HTTPException(
                        status_code=409,
                        detail="Ese correo ya está en uso por otra cuenta"
                    )

            cursor.execute(
                """
                UPDATE users
                SET full_name = %s, email = %s
                WHERE id = %s;
                """,
                (profile.full_name, profile.email, user["id"])
            )

            connection.commit()

    finally:
        connection.close()

    # La sesión guarda una copia de full_name para no consultar la BD
    # en cada página (por eso el nombre aparece en la barra superior
    # sin una query extra) -- hay que refrescarla aquí o el cambio no
    # se vería hasta la próxima vez que la persona inicie sesión.
    request.session["user"] = {**user, "full_name": profile.full_name}

    return {
        "message": "Perfil actualizado",
        "full_name": profile.full_name,
        "email": profile.email
    }


@app.post("/me/password")
def change_password(
    change: PasswordChange,
    user: dict = Depends(get_current_user)
):

    if len(change.new_password) < 8:
        raise HTTPException(
            status_code=400,
            detail="La nueva contraseña debe tener al menos 8 caracteres"
        )

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                "SELECT password_hash FROM users WHERE id = %s;",
                (user["id"],)
            )

            row = cursor.fetchone()

            if row is None or not verify_password(change.current_password, row[0]):
                raise HTTPException(
                    status_code=401,
                    detail="La contraseña actual no es correcta"
                )

            new_hash = hash_password(change.new_password)

            cursor.execute(
                "UPDATE users SET password_hash = %s WHERE id = %s;",
                (new_hash, user["id"])
            )

            connection.commit()

    finally:
        connection.close()

    return {"message": "Contraseña actualizada"}


@app.get("/api/perfil")
def api_perfil(user: dict = Depends(get_current_user)):
    """Datos de la pantalla Perfil en React (menú de usuario > "Mi
    perfil"). 'roles' sale de la sesión (ya calculado en el login, no
    hace falta otra consulta); 'is_active' se agrega acá porque es una
    columna real de 'users' -- no un valor inventado."""

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                """
                SELECT username, full_name, email, is_active, created_at, last_login_at
                FROM users
                WHERE id = %s;
                """,
                (user["id"],)
            )

            row = cursor.fetchone()

    finally:
        connection.close()

    if row is None:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")

    return {
        "username": row[0],
        "full_name": row[1],
        "email": row[2],
        "is_active": row[3],
        "created_at": row[4].strftime("%d/%m/%Y") if row[4] else None,
        "last_login_at": row[5].strftime("%d/%m/%Y %H:%M:%S") if row[5] else None,
        "roles": user.get("roles", []),
    }


@app.get("/api/roles")
def api_roles(user: dict = Depends(get_current_user)):
    """Catálogo real de roles -- fuente para el selector de 'Rol' en
    crear/editar usuario (frontend) y para validar en el backend que un
    role_id/nombre de rol enviado realmente existe (ver POST /users,
    PATCH /users/{id}). Si mañana se agrega una fila nueva en 'roles',
    aparece acá sin tocar código (2026-08-16, ver PENDIENTES.md,
    auditoría de catálogos duplicados)."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT id, name, description FROM roles ORDER BY name;")
            rows = cursor.fetchall()
    finally:
        connection.close()

    return {
        "roles": [
            {"id": r[0], "name": r[1], "description": r[2]}
            for r in rows
        ]
    }


@app.post("/users")
def create_user(
    new_user: UserCreate,
    current_user: dict = Depends(require_role("admin"))
):
    """Reemplaza a bootstrap_admin.py para el uso diario: ahora un
    admin ya logueado puede dar de alta a otras personas desde el
    propio sistema, en vez de correr un script a mano cada vez."""

    if len(new_user.password) < 8:
        raise HTTPException(
            status_code=400,
            detail="La contraseña debe tener al menos 8 caracteres"
        )

    password_hash = hash_password(new_user.password)

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                "SELECT id FROM users WHERE username = %s;",
                (new_user.username,)
            )

            if cursor.fetchone():
                raise HTTPException(
                    status_code=409,
                    detail=f"Ya existe un usuario '{new_user.username}'"
                )

            # El rol pedido tiene que existir en 'roles' -- no se crea
            # uno nuevo silenciosamente a partir de un string que mandó
            # el frontend (antes lo hacía: cualquier typo en el campo de
            # texto generaba una fila basura en 'roles'. Ver
            # PENDIENTES.md, auditoría de catálogos duplicados,
            # 2026-08-16). El frontend ahora elige de un <select>
            # poblado con GET /api/roles, pero el backend no confía en
            # eso -- valida igual.
            cursor.execute(
                "SELECT id FROM roles WHERE name = %s;",
                (new_user.role,)
            )

            role_row = cursor.fetchone()

            if role_row is None:
                raise HTTPException(
                    status_code=422,
                    detail=f"El rol '{new_user.role}' no existe. Los roles disponibles se obtienen de GET /api/roles."
                )

            role_id = role_row[0]

            cursor.execute(
                """
                INSERT INTO users (username, password_hash, full_name, email)
                VALUES (%s, %s, %s, %s)
                RETURNING id;
                """,
                (
                    new_user.username,
                    password_hash,
                    new_user.full_name,
                    new_user.email
                )
            )

            user_id = cursor.fetchone()[0]

            cursor.execute(
                """
                INSERT INTO user_roles (user_id, role_id)
                VALUES (%s, %s);
                """,
                (user_id, role_id)
            )

            log_audit(
                cursor, current_user["id"], "CREATE_USER", "users", user_id,
                f"{new_user.username} ({new_user.role})"
            )

            connection.commit()

        return {
            "message": "Usuario creado correctamente",
            "user_id": user_id,
            "username": new_user.username,
            "role": new_user.role
        }

    finally:
        connection.close()


@app.patch("/users/{user_id}")
def update_user(
    user_id: int,
    payload: UserUpdate,
    current_user: dict = Depends(require_role("admin"))
):
    """Editar/desactivar un usuario (2026-08-12) -- hasta ahora
    'POST /users' solo permitía crear, no había forma de corregir un
    dato o desactivar una cuenta sin ir directo a la base. Mismo nivel
    de acceso que crear usuarios (solo admin), consistente con que
    'GET /usuarios' hoy no lo exige (gap ya documentado en
    PENDIENTES.md) -- este endpoint sí lo exige porque escribe."""

    if payload.full_name is None and payload.email is None and payload.is_active is None and payload.role is None:
        raise HTTPException(status_code=422, detail="Nada para actualizar")

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute("SELECT username FROM users WHERE id = %s;", (user_id,))
            existing = cursor.fetchone()

            if existing is None:
                raise HTTPException(status_code=404, detail="Usuario no encontrado")

            fields = []
            values = []
            change_parts = []

            if payload.full_name is not None:
                fields.append("full_name = %s")
                values.append(payload.full_name)
                change_parts.append(f"nombre -> {payload.full_name}")

            if payload.email is not None:
                fields.append("email = %s")
                values.append(payload.email)
                change_parts.append(f"email -> {payload.email}")

            if payload.is_active is not None:
                fields.append("is_active = %s")
                values.append(payload.is_active)
                change_parts.append(f"activo -> {payload.is_active}")

            if fields:
                fields.append("updated_at = CURRENT_TIMESTAMP")
                values.append(user_id)

                cursor.execute(
                    f"UPDATE users SET {', '.join(fields)} WHERE id = %s;",
                    values
                )

            if payload.role is not None:

                # Igual que POST /users: el rol tiene que existir ya en
                # 'roles', no se crea uno nuevo a partir de texto libre.
                cursor.execute("SELECT id FROM roles WHERE name = %s;", (payload.role,))
                role_row = cursor.fetchone()

                if role_row is None:
                    raise HTTPException(
                        status_code=422,
                        detail=f"El rol '{payload.role}' no existe. Los roles disponibles se obtienen de GET /api/roles."
                    )

                role_id = role_row[0]

                # Reemplaza los roles existentes en vez de sumarlos --
                # hoy la UI solo maneja "un rol por usuario" (mismo
                # criterio que POST /users), no roles múltiples.
                cursor.execute("DELETE FROM user_roles WHERE user_id = %s;", (user_id,))
                cursor.execute(
                    "INSERT INTO user_roles (user_id, role_id) VALUES (%s, %s);",
                    (user_id, role_id)
                )
                change_parts.append(f"rol -> {payload.role}")

            log_audit(
                cursor, current_user["id"], "UPDATE_USER", "users", user_id,
                f"{existing[0]}: {', '.join(change_parts)}" if change_parts else existing[0]
            )

            connection.commit()

        return {"message": "Usuario actualizado", "user_id": user_id}

    finally:
        connection.close()


@app.get("/")
def root():
    return {
        "message": "Servidor funcionando"
    }


def require_session_user(request: Request):
    """Versión 'de página' de get_current_user: en vez de devolver 401
    en JSON, devuelve None para que la ruta que llama decida mandar al
    navegador a /login. Evita repetir las mismas dos líneas en cada
    página protegida."""

    return request.session.get("user")


def _endpoint_cte(stale_seconds):
    """Antes era el módulo-level 'ENDPOINT_CTE', formateado una sola
    vez al importar el módulo con la constante fija. Ahora es función
    porque 'stale_seconds' se lee de 'system_settings' en cada
    request (ver get_agent_stale_seconds) -- si quedara como string ya
    formateado al arrancar el proceso, cambiar el valor desde
    /configuracion no tendría efecto hasta reiniciar el servidor."""

    return """
    WITH endpoint_risk AS (
        SELECT alerts.agent_id,
               MAX(severity_levels.min_score) AS worst_min_score
        FROM alerts
        JOIN severity_levels ON severity_levels.id = alerts.severity_id
        WHERE alerts.status = 'NEW'
        GROUP BY alerts.agent_id
    ),
    endpoint_data AS (
        SELECT agents.id, endpoints.hostname, endpoints.os AS operating_system, endpoints.os_version,
               endpoints.ip_address, agents.agent_version,
               agents.status, agents.last_seen_at, agents.enrolled_at,
               CASE
                   WHEN agents.status != 'ONLINE' THEN 'offline'
                   WHEN agents.last_seen_at >= CURRENT_TIMESTAMP - INTERVAL '{stale_seconds} seconds' THEN 'ok'
                   ELSE 'attention'
               END AS status_bucket,
               -- 'Peor severidad' de este endpoint: se busca en
               -- severity_levels la fila cuyo min_score coincide con
               -- el máximo alcanzado (sin alertas -> se usa la fila
               -- de menor min_score, el nivel más bajo del catálogo).
               -- Sin CASE ni diccionario hardcodeado -- el nombre y el
               -- orden salen enteramente de la tabla real.
               COALESCE(worst_sev.name, lowest_sev.name) AS risk_bucket
        FROM agents
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        LEFT JOIN endpoint_risk ON endpoint_risk.agent_id = agents.id
        LEFT JOIN severity_levels AS worst_sev ON worst_sev.min_score = endpoint_risk.worst_min_score
        CROSS JOIN (SELECT name FROM severity_levels ORDER BY min_score ASC LIMIT 1) AS lowest_sev
    )
""".format(stale_seconds=stale_seconds)

ENDPOINTS_PAGE_SIZE = 25


def _effective_agent_rules_cte():
    """Motor heurístico -- configuración por endpoint (2026-08-16, ver
    PENDIENTES.md, "Implementación final del motor heurístico y
    configuración por endpoint"). 'heuristic_rules' es la configuración
    GLOBAL de una regla; 'agent_rule' es un override OPCIONAL para un
    (agent_id, rule_id) puntual -- no se crea ninguna tabla nueva, esta
    es exactamente la que ya existía para este propósito.

    Resolución de la configuración EFECTIVA (la que el agente debe
    aplicar y la que el servidor debe usar para calcular el score):
    - threshold/window_seconds/weight: COALESCE(agent_rule.x,
      heuristic_rules.x) -- si el override existe pero ese campo
      puntual quedó en NULL, se hereda el valor global campo por
      campo, sin obligar a repetir todos los parámetros (sección 4 de
      la especificación).
    - is_active: si existe una fila en 'agent_rule' para ese
      (agent_id, rule_id), manda agent_rule.is_active (aunque sea
      TRUE, aunque el override no toque threshold/window/weight) --
      la sola presencia de la fila es la señal de "este endpoint tiene
      un override para esta regla". Si no existe fila, manda
      heuristic_rules.is_active (sección 17).

    Se parametriza con %(agent_id)s -- quien arme la consulta final
    debe pasar ese parámetro (y los que agregue el WHERE/SELECT que
    siga a este CTE)."""

    return """
    WITH effective_rules AS (
        SELECT
            heuristic_rules.id,
            heuristic_rules.name,
            heuristic_rules.description,
            heuristic_rules.event_type_id,
            heuristic_rules.metric_type_id,
            COALESCE(agent_rule.weight, heuristic_rules.weight) AS effective_weight,
            COALESCE(agent_rule.threshold, heuristic_rules.threshold) AS effective_threshold,
            COALESCE(agent_rule.window_seconds, heuristic_rules.window_seconds) AS effective_window_seconds,
            CASE
                WHEN agent_rule.id IS NOT NULL THEN agent_rule.is_active
                ELSE heuristic_rules.is_active
            END AS effective_is_active,
            heuristic_rules.weight AS global_weight,
            heuristic_rules.threshold AS global_threshold,
            heuristic_rules.window_seconds AS global_window_seconds,
            heuristic_rules.is_active AS global_is_active,
            agent_rule.id AS override_id,
            agent_rule.weight AS override_weight,
            agent_rule.threshold AS override_threshold,
            agent_rule.window_seconds AS override_window_seconds,
            agent_rule.is_active AS override_is_active
        FROM heuristic_rules
        LEFT JOIN agent_rule
            ON agent_rule.rule_id = heuristic_rules.id
           AND agent_rule.agent_id = %(agent_id)s
    )
    """


# ============================================================
# AISLAMIENTO -- el estado pertenece al ENDPOINT, no al incidente
# (2026-08-18, ver PENDIENTES.md, "Revisión y corrección integral de
# ALFA-Sentinel", problema H).
#
# BUG REAL ENCONTRADO (no solo cosmético): antes de esto, varias
# consultas ('COMBINED_CTE' para la tabla de Incidentes y Alertas,
# get_incidente_drawer(), GET /alerts/open, GET /api/respuesta, y el
# guard de isolate_incident_manually() que impide pedir un aislamiento
# duplicado) buscaban el estado de aislamiento filtrando por
# 'host_isolations.incident_id = <este incidente puntual>'. Un mismo
# endpoint puede tener varios incidentes (PC-01 -> INC-001, INC-002,
# INC-003); aislar desde INC-001 inserta una fila con
# incident_id = INC-001.id -- así que consultar por incident_id desde
# INC-002 o INC-003 nunca la encontraba, y esas pantallas seguían
# ofreciendo "Aislar" sobre un endpoint que YA estaba aislado. Peor
# todavía: el guard de isolate_incident_manually() tenía el mismo
# defecto, así que un segundo clic en "Aislar" desde un incidente
# DISTINTO del mismo endpoint no era bloqueado -- se llegaba a insertar
# una segunda orden 'REQUESTED' real para un endpoint que el agente ya
# estaba aislando o ya había aislado.
#
# El aislamiento es una propiedad del AGENTE/ENDPOINT, no del incidente
# que lo disparó (host_isolations.agent_id existe y es NOT NULL,
# 'incident_id' solo registra CUÁL incidente lo originó, para
# trazabilidad -- ver database/schema.sql). Estas tres funciones son la
# única fuente de esa lógica, reutilizada en todos los lugares que antes
# tenían su propia consulta (a veces correcta, a veces no) -- equivalente
# al 'is_endpoint_isolated(agent_id)' pedido explícitamente. Se
# implementan como fragmentos de SQL (mismo patrón ya establecido por
# _effective_agent_rules_cte() arriba) en vez de una función Python que
# ejecute su propia query, porque se usan como subconsulta DENTRO de
# consultas más grandes (listas paginadas, CTEs) donde ejecutar una
# consulta Python aparte por fila sería mucho más lento -- 'agent_id_expr'
# es el nombre de columna/alias/placeholder ('agents.id',
# 'incidents.agent_id', '%s') ya disponible en el FROM de quien la usa;
# se interpola tal cual porque es SQL armado en Python, no un valor de
# usuario.
def _agent_isolation_status_sql(agent_id_expr):
    """Estado de la orden de aislamiento MÁS RECIENTE de este agente,
    sin importar su estado actual (incluye 'RELEASED', necesario para
    que un aislamiento liberado se siga viendo como tal en vez de
    desaparecer -- ver CASE-E de
    tests/heuristic/test_tiempo_real_orden_consistencia.py, el mismo
    criterio que ya usaba el drawer de endpoint)."""
    return f"""(
        SELECT host_isolations.status FROM host_isolations
        WHERE host_isolations.agent_id = {agent_id_expr}
        ORDER BY host_isolations.requested_at DESC LIMIT 1
    )"""


def _agent_isolation_id_sql(agent_id_expr):
    """Id de esa misma fila más reciente -- lo que necesita el botón
    'Liberar' para llamar a POST /host-isolations/{id}/release."""
    return f"""(
        SELECT host_isolations.id FROM host_isolations
        WHERE host_isolations.agent_id = {agent_id_expr}
        ORDER BY host_isolations.requested_at DESC LIMIT 1
    )"""


def _agent_is_isolated_sql(agent_id_expr):
    """Booleano: ¿este agente está aislado AHORA MISMO? A diferencia de
    las dos funciones de arriba (que traen la fila más reciente exista o
    no un aislamiento activo, para mostrar 'Liberado'), esta es la
    condición real para decidir si corresponde ofrecer 'Aislar' (no) o
    bloquear una segunda orden (sí) -- 'RELEASE_REQUESTED' cuenta como
    aislado todavía (el agente no confirmó la liberación), 'RELEASED'/
    'ISOLATION_FAILED'/'RELEASE_FAILED' no."""
    return f"""EXISTS (
        SELECT 1 FROM host_isolations
        WHERE host_isolations.agent_id = {agent_id_expr}
          AND host_isolations.status IN ('REQUESTED', 'EXECUTED', 'RELEASE_REQUESTED')
          AND host_isolations.released_at IS NULL
    )"""


# EVENT_TYPE_LABELS_ES y ALERT_RULE_LABELS_ES se eliminaron (2026-08-16,
# ver PENDIENTES.md, auditoría de catálogos duplicados): duplicaban
# event_types.description y heuristic_rules.name respectivamente. Los
# nombres de regla se muestran tal cual vienen de heuristic_rules.name,
# sin traducir.

# RISK_LABELS_ES y ALERT_SEVERITY_LABELS_ES se eliminaron (2026-08-16,
# corrección arquitectónica: "si un dato existe en un catálogo de
# PostgreSQL, ese dato es la fuente de verdad"). 'severity_levels.name'
# ahora ES el nombre definitivo en español (BAJO/MEDIO/ALTO/CRÍTICO,
# ver migración en database/migration_2026-08-16_severity_levels_espanol.sql) -- ya no
# hace falta traducir nada, el valor que devuelve cualquier consulta a
# 'severity_levels'/'alerts.severity_id' es directamente lo que se le
# muestra al usuario. En la práctica BAJO casi nunca aparece en
# 'alerts': report_alert exige matched_rules no vacío, así que la
# alerta siempre nace con al menos un peso > 0.

# La nueva 'alerts.status' (alfa_sentinel) es un VARCHAR sin CHECK
# constraint -- estos 5 valores ya no vienen impuestos por la base,
# los define este diccionario. El mapeo a "estado gestionable" (Nueva/
# En investigación/Confirmada/Cerrada/Falso positivo) sigue siendo una
# decisión de producto, no algo que venga dado.
ALERT_STATUS_LABELS_ES = {
    "NEW": "Nueva",
    "ACKNOWLEDGED": "En investigación",
    "ESCALATED": "Confirmada",
    "CLOSED": "Cerrada",
    "FALSE_POSITIVE": "Falso positivo",
}

# "Corrección definitiva en la lógica y presentación de ALERTAS"
# (2026-08-18, ver PENDIENTES.md): el título VISIBLE de una alerta (o
# de un incidente, que es el mismo concepto agrupado -- ver más abajo)
# ya NO es el nombre de la primera regla que llegó ni el de la de
# mayor peso -- es un título general que representa el NIVEL DE RIESGO
# final (severity_levels.name, que ya existe y ya se calcula en todos
# lados), consistente con que la alerta representa el EPISODIO
# COMPLETO, no una señal individual. Los nombres de reglas
# individuales (Consumo CPU Elevado, Acceso Honeyfile, etc.) siguen
# existiendo tal cual -- pasan a mostrarse SOLO dentro del detalle,
# como "señales/reglas que contribuyeron" (alert_rule/heuristic_rules,
# sin tocar esas tablas). Este diccionario es la única fuente de esos
# 4 nombres -- si `severity_levels` alguna vez agrega un nivel nuevo,
# el fallback a BAJO evita que una alerta se quede sin título en vez
# de romper.
ALERT_GENERAL_TITLE_ES = {
    "BAJO": "ACTIVIDAD ANÓMALA",
    "MEDIO": "ACTIVIDAD SOSPECHOSA",
    "ALTO": "POSIBLE ATAQUE DE RANSOMWARE",
    "CRÍTICO": "ATAQUE DE RANSOMWARE PROBABLE",
}


def alert_general_title(severity_name):
    """Título general por severidad -- ver ALERT_GENERAL_TITLE_ES.
    Se calcula EN CADA LECTURA a partir de la severidad ACTUAL (nunca
    se guarda un título fijo en 'alerts.title'/'incidents.title' para
    mostrarlo tal cual), así que cuando un episodio recibe evidencia
    nueva y su severidad sube, el título general "se recalcula" solo
    -- no hace falta ningún UPDATE adicional ni ningún estado a
    mantener sincronizado (sección 11 de la especificación: "cuando
    una nueva evidencia se incorpora al episodio... recalcular título
    general"). 'alerts.title'/'incidents.title' (las columnas
    guardadas) NO se tocan -- siguen existiendo con el texto original
    que mandó el agente para el primer evento del episodio, útil como
    evidencia/búsqueda interna, pero ya no se usan para el título que
    ve el analista."""
    return ALERT_GENERAL_TITLE_ES.get(severity_name, ALERT_GENERAL_TITLE_ES["BAJO"])

# Igual que ALERT_STATUS_LABELS_ES: 'incidents.status' ya no tiene
# CHECK constraint en la base nueva, así que estos 4 valores los fija
# este diccionario. "CONTAINED" significa que ya se tomaron las
# acciones necesarias y el incidente está bajo control -- no que se
# confirmó que era ransomware. Esa determinación es aparte (ver
# INCIDENT_CLASSIFICATION_LABELS_ES).
INCIDENT_STATUS_LABELS_ES = {
    "OPEN": "Abierto",
    "IN_PROGRESS": "En investigación",
    "CONTAINED": "Contenido",
    "CLOSED": "Cerrado",
}

# 'host_isolations.status' -- agregado 2026-08-16 junto con el motor
# heurístico, corregido 2026-08-17 (ver PENDIENTES.md, "Corrección
# definitiva del motor heurístico..."): 'REQUESTED' es lo que
# report_alert() escribe cuando se cumple la condición de aislamiento
# (sección 30) -- una orden real, no una nota informativa. El agente de
# ese endpoint (agent/isolation_sync.py) la recoge y la ejecuta
# (agent/isolation_executor.py), y confirma vía
# POST /agent/isolation-status/report: 'EXECUTED' si pudo aislar de
# verdad, 'ISOLATION_FAILED' si lo intentó y no pudo (sin privilegios,
# comando no disponible, error real del SO -- nunca se finge éxito).
# 'RECOMMENDED' queda como valor LEGADO -- ninguna ruta nueva lo
# escribe, se conserva en el diccionario solo para traducir filas
# viejas de una base creada antes de esta corrección.
#
# Extendido 2026-08-17 (ver PENDIENTES.md, "Aislamiento de host --
# modo development, laboratorio y producción") con la operación
# inversa (sección 18 de esa especificación: "UNISOLATE"), reutilizando
# la misma columna 'status' sin CHECK constraint (ver comentario en
# database/schema.sql) -- mismo criterio que se usó para agregar
# REQUESTED/EXECUTED/ISOLATION_FAILED en la corrección anterior, sin
# tocar la estructura de la tabla:
# 'RELEASE_REQUESTED' (orden de liberar, el agente todavía no
# confirmó) -> 'RELEASED' (confirmado) o, si falla, vuelve a
# 'EXECUTED' (sigue aislado de verdad, que es el estado real -- no
# existe un 'RELEASE_FAILED' persistido porque nada cambió realmente
# respecto de antes de intentar liberar; el motivo del fallo queda en
# 'result' igual que cualquier otro intento).
ISOLATION_STATUS_LABELS_ES = {
    "RECOMMENDED": "Recomendado (legado, no ejecutado)",
    "REQUESTED": "Solicitado (esperando confirmación del agente)",
    "EXECUTED": "Ejecutado",
    "ISOLATION_FAILED": "Falló la ejecución",
    "RELEASE_REQUESTED": "Liberación solicitada (esperando confirmación del agente)",
    "RELEASED": "Liberado",
}

ISOLATION_TYPE_LABELS_ES = {
    "NETWORK": "Aislamiento de red",
}

# Clasificación del resultado de la investigación -- separada del
# estado a propósito, para no mezclar "en qué etapa del ciclo de vida
# está" con "qué se determinó que era". Igual que status/rule: sin
# CHECK constraint en la base nueva, los valores los fija este
# diccionario.
INCIDENT_CLASSIFICATION_LABELS_ES = {
    "CONFIRMED": "Confirmado",
    "POSSIBLE_THREAT": "Posible amenaza",
    "FALSE_POSITIVE": "Falso positivo",
    "LEGITIMATE_ACTIVITY": "Actividad legítima",
    "UNDETERMINED": "No determinado",
}

# Página /reportes (2026-08-12). Solo 3 tipos -- cubren lo directivo
# (SECURITY), lo operativo (ENDPOINTS) e investigativo (INCIDENTS) sin
# inflar la interfaz. No hay generación automática/"por sistema"
# todavía: 'generated_by' siempre sale de la sesión de quien lo pide
# (ver PENDIENTES.md).
REPORT_TYPE_LABELS_ES = {
    "SECURITY": "Informe de Seguridad General",
    "ENDPOINTS": "Informe de Actividad de Endpoints",
    "INCIDENTS": "Informe de Incidentes",
}

REPORT_PERIOD_OPTIONS = [
    ("7d", "Últimos 7 días"),
    ("30d", "Últimos 30 días"),
    ("90d", "Últimos 90 días"),
    ("all", "Todo el histórico"),
]

REPORTS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "generated_reports")

REPORT_FORMAT_MEDIA_TYPES = {
    "PDF": "application/pdf",
    "XLSX": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}


def _resolve_report_period(period: str):
    """Traduce el preset elegido en el formulario a un rango de fechas
    real. 'all' arranca en la fecha de creación de la base más vieja
    posible que tiene sentido acá (no hay datos de antes del propio
    proyecto), así que se usa una fecha fija bien anterior en vez de
    NULL, para no tener que ramificar el resto del código en "con
    filtro de fecha" / "sin filtro de fecha"."""

    now = datetime.now()

    if period == "7d":
        return now - timedelta(days=7), now, "Últimos 7 días"
    elif period == "30d":
        return now - timedelta(days=30), now, "Últimos 30 días"
    elif period == "90d":
        return now - timedelta(days=90), now, "Últimos 90 días"
    elif period == "all":
        return datetime(2000, 1, 1), now, "Todo el histórico"
    else:
        raise HTTPException(status_code=422, detail=f"Período desconocido: '{period}'")

# A diferencia de Eventos/Detecciones (15m/1h/24h, pensado para
# actividad reciente), acá se usan ventanas más largas porque un
# incidente puede seguir abierto días. Son intervalos relativos a
# AHORA, no días de calendario -- "Hoy" sería engañoso si en realidad
# filtra "últimas 24 horas" contadas desde este instante.
INCIDENTES_SINCE_OPTIONS = {
    "24h": ("Últimas 24 horas", "24 hours"),
    "7d": ("Últimos 7 días", "7 days"),
    "30d": ("Últimos 30 días", "30 days"),
}


@app.get("/api/endpoints/{agent_id}/drawer")
def get_endpoint_drawer_data(agent_id: int, request: Request):
    """API para obtener la información completa del Host Drawer (panel lateral)."""
    user = require_session_user(request)
    if user is None:
        return JSONResponse({"error": "No autorizado"}, status_code=401)

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            stale_seconds = get_agent_stale_seconds(cursor)

            cursor.execute(
                """
                SELECT agents.id, endpoints.hostname, endpoints.os, endpoints.os_version,
                       endpoints.ip_address, agents.agent_version, agents.status,
                       agents.last_seen_at, agents.enrolled_at
                FROM agents
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                WHERE agents.id = %s;
                """,
                (agent_id,)
            )
            row = cursor.fetchone()
            if not row:
                return JSONResponse({"error": "Endpoint no encontrado"}, status_code=404)

            # Mismo cálculo que _endpoint_cte()/api_endpoints() (lista de
            # Endpoints en React) -- Healthy/Warning/Offline según el
            # último heartbeat contra el umbral configurado. No es un
            # dato nuevo, es la misma fórmula real aplicada acá para
            # que el drawer y la lista digan lo mismo.
            if row[6] != "ONLINE":
                agent_health = "OFFLINE"
            elif row[7] and (datetime.now(row[7].tzinfo) - row[7]).total_seconds() <= stale_seconds:
                agent_health = "HEALTHY"
            else:
                agent_health = "WARNING"

            # Alertas activas e incidentes asociados a este endpoint --
            # mismas tablas/columnas que ya usa el resto del sistema
            # (alerts.status='NEW', incidents.status!='CLOSED'), sin
            # estructuras nuevas.
            cursor.execute(
                "SELECT COUNT(*) FROM alerts WHERE agent_id = %s AND status = 'NEW';",
                (agent_id,)
            )
            alerts_active = cursor.fetchone()[0]

            cursor.execute(
                "SELECT COUNT(*), COUNT(*) FILTER (WHERE status != 'CLOSED') FROM incidents WHERE agent_id = %s;",
                (agent_id,)
            )
            incidents_total, incidents_active = cursor.fetchone()

            # Nivel de riesgo -- se ordena por severity_levels.min_score
            # (no por un CASE con nombres hardcodeados: el orden real
            # de severidad ya está en la tabla). Si el endpoint no
            # tiene alertas abiertas, se cae al nivel más bajo del
            # catálogo, el que quede.
            cursor.execute(
                """
                SELECT severity_levels.name, severity_levels.min_score, severity_levels.max_score
                FROM alerts
                JOIN severity_levels ON severity_levels.id = alerts.severity_id
                WHERE alerts.agent_id = %s AND alerts.status = 'NEW'
                ORDER BY severity_levels.min_score DESC
                LIMIT 1;
                """,
                (agent_id,)
            )
            risk_row = cursor.fetchone()
            if risk_row is None:
                cursor.execute("SELECT name, min_score, max_score FROM severity_levels ORDER BY min_score ASC LIMIT 1;")
                risk_row = cursor.fetchone()
            risk_bucket, risk_min_score, risk_max_score = risk_row
            # Score representativo del nivel (punto medio del rango
            # real en severity_levels) -- reemplaza al diccionario fijo
            # que había antes (peor caso, no se conoce el score exacto
            # de la última alerta acá, solo el nivel).
            representative_risk_score = (float(risk_min_score) + float(risk_max_score)) / 2

            # Aislamiento activo -- real desde la corrección definitiva
            # del motor heurístico (2026-08-17, ver PENDIENTES.md):
            # 'REQUESTED' (orden en curso, el agente todavía no
            # confirmó) y 'EXECUTED' (confirmado) cuentan como aislado
            # de cara al usuario; 'ISOLATION_FAILED' NO cuenta. Este era
            # ya el criterio CORRECTO (agent_id, no incident_id) -- desde
            # 2026-08-18 (problema H, ver PENDIENTES.md) se expresa con
            # _agent_is_isolated_sql(), la misma fuente única que ahora
            # usan también COMBINED_CTE, get_incidente_drawer(),
            # /alerts/open, /api/respuesta y el guard de
            # isolate_incident_manually().
            cursor.execute(
                f"SELECT {_agent_is_isolated_sql('%s')};",
                (agent_id,)
            )
            is_isolated = bool(cursor.fetchone()[0])
            # Agregados 2026-08-17 (ver PENDIENTES.md, "Corrección de
            # tiempo real, ordenamiento y consistencia") -- 'is_isolated'
            # (booleano) ya alcanzaba para pintar el badge de estado,
            # pero el nuevo botón "Liberar" de este drawer necesita el
            # id real de la fila y el status exacto.
            #
            # OJO: esto es una consulta APARTE de la de 'is_isolated' de
            # arriba, no la misma fila -- la de arriba filtra
            # 'released_at IS NULL' porque solo le interesa si el
            # endpoint está aislado AHORA. Reusar ese mismo resultado acá
            # rompía la consistencia entre pantallas (sección 18): una
            # vez que el agente confirma 'RELEASED' (released_at deja de
            # ser NULL), esa consulta ya no encuentra ninguna fila y
            # 'isolation_status' quedaba en None -- mientras que
            # /api/incidentes, el drawer del incidente y /alerts/open
            # (que sí usan "la fila más reciente sin importar su estado
          # actual", igual que en COMBINED_CTE) seguían mostrando
            # 'RELEASED' para el mismo aislamiento. Detectado por
            # tests/heuristic/test_tiempo_real_orden_consistencia.py
            # (CASE-E). Acá se usa el mismo criterio que esos otros tres
            # lugares: la fila más reciente por 'requested_at', exista o
            # no un aislamiento activo en este momento.
            cursor.execute(
                f"SELECT {_agent_isolation_id_sql('%s')}, {_agent_isolation_status_sql('%s')};",
                (agent_id, agent_id)
            )
            iso_row_latest = cursor.fetchone()
            isolation_id = iso_row_latest[0] if iso_row_latest else None
            isolation_status = iso_row_latest[1] if iso_row_latest else None

            # Incidente activo más reciente de este endpoint -- lo que
            # necesita el botón "Aislar endpoint manualmente" (sección
            # 20 de "Aislamiento de host -- modo development,
            # laboratorio y producción", 2026-08-17, ver PENDIENTES.md)
            # para saber a qué incidente asociar la orden. 'host_isolations.
            # incident_id' es NOT NULL -- sin un incidente activo real,
            # no hay a qué asociar el aislamiento manual (no se inventa
            # un incidente solo para poder aislar).
            cursor.execute(
                "SELECT id FROM incidents WHERE agent_id = %s AND status != 'CLOSED' ORDER BY opened_at DESC LIMIT 1;",
                (agent_id,)
            )
            active_incident_row = cursor.fetchone()
            active_incident_id = active_incident_row[0] if active_incident_row else None

            # Honeyfiles en este host
            cursor.execute(
                "SELECT COUNT(*) FROM honeyfiles WHERE agent_id = %s;",
                (agent_id,)
            )
            honeyfiles_total = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT file_path, detected_at FROM events
                WHERE agent_id = %s AND (file_path ILIKE '%%honeyfile%%' OR file_path ILIKE '%%!0_%%')
                ORDER BY id DESC LIMIT 1;
                """,
                (agent_id,)
            )
            violated_evt = cursor.fetchone()
            violated_file = violated_evt[0] if violated_evt else None
            violated_at = violated_evt[1] if violated_evt else None

            # Último evento/alerta. 'alerts' ya no tiene 'file_path' como
            # columna propia (nunca la tuvo en realidad -- ver
            # PENDIENTES.md). Corregido 2026-08-18 (ver PENDIENTES.md,
            # "Corrección definitiva en la lógica y presentación de
            # ALERTAS"): antes esto traía la alerta y su regla en UNA
            # sola consulta con LEFT JOIN alert_rule/heuristic_rules
            # -- si esa alerta tenía más de una regla vinculada, el
            # 'ORDER BY alerts.id DESC LIMIT 1' se aplicaba sobre filas
            # ya multiplicadas por el JOIN y terminaba agarrando una
            # regla arbitraria (nunca garantizada, cualquiera de las
            # vinculadas). Ahora son dos consultas: la alerta más
            # reciente primero (sin el join, una sola fila real), y
            # después SUS reglas por separado, para poder aplicar el
            # mismo orden de relevancia que el resto del sistema
            # (sort_contributing_rules) y no depender de qué fila
            # devolvía Postgres primero.
            cursor.execute(
                """
                SELECT alerts.id, severity_levels.name, alerts.created_at
                FROM alerts
                JOIN severity_levels ON severity_levels.id = alerts.severity_id
                WHERE alerts.agent_id = %s
                ORDER BY alerts.id DESC LIMIT 1;
                """,
                (agent_id,)
            )
            alert_row = cursor.fetchone()
            latest_alert = None
            if alert_row:
                latest_alert_id, latest_severity, latest_created_at = alert_row

                cursor.execute(
                    """
                    SELECT heuristic_rules.name, alert_rule.weight_applied, alert_rule.matched_at
                    FROM alert_rule
                    JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                    WHERE alert_rule.alert_id = %s;
                    """,
                    (latest_alert_id,)
                )
                latest_alert_rules = sort_contributing_rules(cursor.fetchall())

                latest_alert = {
                    # Título general por severidad, no el nombre de la
                    # primera regla que llegó -- ver alert_general_title().
                    "title": alert_general_title(latest_severity),
                    "severity": latest_severity,
                    "created_at": latest_created_at.strftime("%d/%m/%Y %H:%M:%S") if latest_created_at else "",
                    "file_path": None,
                    # La regla más relevante (no arbitraria) entre las
                    # que contribuyeron -- ver sort_contributing_rules().
                    "rule_name": latest_alert_rules[0][0] if latest_alert_rules else ""
                }

            return {
                "id": row[0],
                "agent_code": f"AGT-{row[0]:06d}",
                "hostname": row[1],
                "operating_system": row[2],
                "os_version": row[3] or "",
                "architecture": None,
                "ip_address": str(row[4]) if row[4] else "127.0.0.1",
                "mac_address": None,
                "agent_version": row[5] or "v1.0.0 (watchdog)",
                "status": row[6],
                "agent_health": agent_health,
                "last_seen_at": row[7].strftime("%d/%m/%Y %H:%M:%S") if row[7] else "Nunca",
                "last_seen_ago": time_ago(row[7]),
                "enrolled_at": row[8].strftime("%d/%m/%Y") if row[8] else "",
                "risk_bucket": risk_bucket,
                "risk_score": representative_risk_score,
                "is_isolated": is_isolated,
                "isolation_id": isolation_id,
                "isolation_status": isolation_status,
                "active_incident_id": active_incident_id,
                "alerts_active": alerts_active,
                "incidents_total": incidents_total,
                "incidents_active": incidents_active,
                "honeyfiles_total": honeyfiles_total,
                "honeyfiles_violated_file": violated_file,
                "honeyfiles_violated_ago": time_ago(violated_at) if violated_at else None,
                "latest_alert": latest_alert
            }
    finally:
        connection.close()


@app.get("/api/honeyfiles")
def api_honeyfiles(
    agent_id: int | None = Query(None),
    status: str = Query(""),
    os_filter: str = Query("", alias="os"),
    search: str = Query(""),
    user: dict = Depends(get_current_user)
):
    """Datos de la pantalla Honeyfiles en React -- consulta, KPIs y
    Wizard de Despliegue (available_agents). No pagina (el inventario
    real hoy es chico).

    Nota: 'status' filtra sobre el estado ya calculado (ACTIVE ->
    TRIGGERED cuando hay activaciones), no directo contra la columna
    'honeyfiles.status' -- la columna en sí solo guarda ACTIVE/INACTIVE,
    'TRIGGERED' es un estado derivado en Python. Filtrar antes de
    calcularlo directo en SQL nunca hubiera encontrado nada con
    status=TRIGGERED; acá se filtra después, sobre el mismo valor que
    ve el analista."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            where_clauses = []
            params = {}

            if agent_id:
                where_clauses.append("honeyfiles.agent_id = %(agent_id)s")
                params["agent_id"] = agent_id
            if os_filter:
                where_clauses.append("endpoints.os ILIKE %(os)s")
                params["os"] = f"%{os_filter}%"
            if search:
                where_clauses.append("(honeyfiles.file_name ILIKE %(search)s OR honeyfiles.file_path ILIKE %(search)s OR endpoints.hostname ILIKE %(search)s)")
                params["search"] = f"%{search}%"

            where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

            cursor.execute(
                f"""
                SELECT honeyfiles.id, honeyfiles.file_name, honeyfiles.file_path,
                       honeyfiles.file_type, honeyfiles.status, honeyfiles.created_at,
                       honeyfiles.last_checked_at, agents.id AS agent_id, endpoints.hostname,
                       endpoints.ip_address, endpoints.os, endpoints.os_version,
                       agents.agent_version, agents.status AS agent_status,
                       agents.last_seen_at
                FROM honeyfiles
                JOIN agents ON agents.id = honeyfiles.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                {where_sql}
                ORDER BY honeyfiles.id DESC;
                """,
                params
            )
            rows = cursor.fetchall()

            cursor.execute("SELECT honeyfile_id, COUNT(*) FROM honeyfile_activations GROUP BY honeyfile_id;")
            activations_dict = dict(cursor.fetchall())

            cursor.execute("SELECT COUNT(*) FROM honeyfiles;")
            total_honeyfiles = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM honeyfiles WHERE status = 'ACTIVE';")
            active_honeyfiles_raw = cursor.fetchone()[0]
            cursor.execute(
                "SELECT COUNT(DISTINCT honeyfile_id) FROM honeyfile_activations "
                "JOIN honeyfiles ON honeyfiles.id = honeyfile_activations.honeyfile_id "
                "WHERE honeyfiles.status = 'ACTIVE';"
            )
            triggered_honeyfiles = cursor.fetchone()[0]
            active_honeyfiles = active_honeyfiles_raw - triggered_honeyfiles

            cursor.execute("SELECT COUNT(*) FROM agent_honeyfile_templates WHERE status = 'PENDING';")
            pending_deployments = cursor.fetchone()[0]
            cursor.execute("SELECT COUNT(*) FROM agent_honeyfile_templates WHERE status = 'FAILED';")
            failed_deployments = cursor.fetchone()[0]

            cursor.execute("SELECT DISTINCT os FROM endpoints ORDER BY os;")
            distinct_os = [r[0] for r in cursor.fetchall()]

            cursor.execute(
                """
                SELECT agents.id, endpoints.hostname, endpoints.os, endpoints.os_version,
                       endpoints.ip_address, agents.status, agents.last_seen_at
                FROM agents
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                ORDER BY agents.status DESC, endpoints.hostname ASC;
                """
            )
            agent_rows = cursor.fetchall()
            available_agents = [
                {
                    "id": r[0],
                    "hostname": r[1],
                    "operating_system": r[2],
                    "os_version": r[3] or "",
                    "ip_address": str(r[4]) if r[4] else "127.0.0.1",
                    "status": r[5],
                    "is_live": (r[5] == "ONLINE" and r[6] is not None and (datetime.now(r[6].tzinfo) - r[6]).total_seconds() < 30) if r[6] else False
                }
                for r in agent_rows
            ]
    finally:
        connection.close()

    honeyfiles_list = []
    for r in rows:
        hf_id = r[0]
        act_cnt = activations_dict.get(hf_id, 0)
        status_val = r[4]
        if act_cnt > 0 and status_val == "ACTIVE":
            status_val = "TRIGGERED"

        last_seen = r[14]
        is_agent_live = (r[13] == "ONLINE" and last_seen is not None and (datetime.now(last_seen.tzinfo) - last_seen).total_seconds() < 30) if last_seen else False

        honeyfiles_list.append({
            "id": hf_id,
            "file_name": r[1],
            "file_path": r[2],
            "file_type": (r[3] or "FILE").upper(),
            "status": status_val,
            "created_at": r[5].strftime("%d/%m/%Y %H:%M:%S") if r[5] else None,
            "last_checked_at": r[6].strftime("%d/%m/%Y %H:%M:%S") if r[6] else None,
            "agent_id": r[7],
            "hostname": r[8],
            "ip_address": str(r[9]) if r[9] else "127.0.0.1",
            "operating_system": r[10],
            "os_version": r[11] or "",
            "agent_version": r[12] or "v1.0.0",
            "agent_status": r[13],
            "is_agent_live": is_agent_live,
            "activations_count": act_cnt
        })

    if status:
        honeyfiles_list = [hf for hf in honeyfiles_list if hf["status"] == status]

    return {
        "summary": {
            "total": total_honeyfiles,
            "active": active_honeyfiles,
            "triggered": triggered_honeyfiles,
            "pending_deployments": pending_deployments,
            "failed_deployments": failed_deployments,
        },
        "distinct_os": distinct_os,
        "available_agents": available_agents,
        "filtered_total": len(honeyfiles_list),
        "honeyfiles": honeyfiles_list,
    }


@app.get("/api/honeyfiles/{honeyfile_id}/detail")
def get_honeyfile_detail_api(honeyfile_id: int, request: Request):
    """API para obtener la información completa del Host Drawer de un Honeyfile."""
    user = require_session_user(request)
    if user is None:
        return JSONResponse({"error": "No autorizado"}, status_code=401)

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            # Join con 'endpoints' para hostname/ip/os (ya no viven en
            # 'agents'). 'file_hash' se lee
            # directo de 'honeyfiles' -- ya no se fabrica un hash de la
            # ruta acá, porque ahora el agente calcula el sha256 real del
            # contenido que escribió y lo manda al reportar el archivo.
            cursor.execute(
                """
                SELECT honeyfiles.id, honeyfiles.file_name, honeyfiles.file_path,
                       honeyfiles.file_type, honeyfiles.status, honeyfiles.created_at,
                       honeyfiles.last_checked_at, honeyfiles.file_hash, agents.id AS agent_id,
                       endpoints.hostname, endpoints.ip_address, endpoints.os, endpoints.os_version,
                       agents.agent_version, agents.status AS agent_status, agents.last_seen_at
                FROM honeyfiles
                JOIN agents ON agents.id = honeyfiles.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                WHERE honeyfiles.id = %s;
                """,
                (honeyfile_id,)
            )
            r = cursor.fetchone()
            if not r:
                return JSONResponse({"error": "Honeyfile no encontrado"}, status_code=404)

            # Historial de activaciones desde honeyfile_activations
            cursor.execute(
                """
                SELECT detected_at, operation, process_name, process_id
                FROM honeyfile_activations
                WHERE honeyfile_id = %s
                ORDER BY id DESC LIMIT 10;
                """,
                (honeyfile_id,)
            )
            act_rows = cursor.fetchall()
            activations = [
                {
                    "detected_at": ar[0].strftime("%d/%m %H:%M") if ar[0] else "",
                    "operation": ar[1],
                    "process_name": ar[2] or "proceso desconocido",
                    "process_id": ar[3] or 0
                }
                for ar in act_rows
            ]

            is_online = (r[14] == "ONLINE" and r[15] is not None and (datetime.now(r[15].tzinfo) - r[15]).total_seconds() < 30) if r[15] else False

            status_val = r[4]
            if len(activations) > 0 and status_val == "ACTIVE":
                status_val = "TRIGGERED"

            return {
                "id": r[0],
                "file_name": r[1],
                "file_path": r[2],
                "file_type": f"Archivo {(r[3] or 'FILE').upper()}",
                "sha256_hash": r[7] or "No disponible",
                "status": status_val,
                "created_at": r[5].strftime("%d/%m/%Y %H:%M") if r[5] else "",
                "last_checked_at": r[6].strftime("%d/%m/%Y %H:%M:%S") if r[6] else "Hace momentos",
                "agent_id": r[8],
                "agent_code": f"AGT-{r[8]:06d}",
                "hostname": r[9],
                "ip_address": str(r[10]) if r[10] else "127.0.0.1",
                "operating_system": r[11],
                "os_version": r[12] or "",
                "agent_version": r[13] or "v1.0.0",
                "is_online": is_online,
                "activations": activations,
                "activations_count": len(activations)
            }
    finally:
        connection.close()


@app.post("/api/honeyfiles/{honeyfile_id}/toggle-status")
def toggle_honeyfile_status_api(honeyfile_id: int, request: Request):
    """API para alternar el estado (ACTIVE / INACTIVE) o desactivar un Honeyfile."""
    user = require_session_user(request)
    if user is None:
        return JSONResponse({"error": "No autorizado"}, status_code=401)

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT status FROM honeyfiles WHERE id = %s;", (honeyfile_id,))
            row = cursor.fetchone()
            if not row:
                return JSONResponse({"error": "Honeyfile no encontrado"}, status_code=404)

            new_status = "INACTIVE" if row[0] in ("ACTIVE", "TRIGGERED") else "ACTIVE"
            cursor.execute(
                "UPDATE honeyfiles SET status = %s, last_checked_at = CURRENT_TIMESTAMP WHERE id = %s;",
                (new_status, honeyfile_id)
            )
            connection.commit()
            return {"id": honeyfile_id, "status": new_status, "message": f"Estado actualizado a {new_status}"}
    finally:
        connection.close()


@app.post("/api/honeyfiles/deploy")
def deploy_honeyfile_api(request: Request, body: dict = None):
    """Crea una plantilla de honeyfile ('honeyfile_templates') y la
    asigna a los agentes elegidos ('agent_honeyfile_templates', en
    PENDING).

    Antes esto insertaba directo en 'honeyfiles' como si el archivo ya
    existiera en el endpoint con solo tocar un botón acá -- ningún
    agente recibía la orden ni la ejecutaba, así que la fila era
    ficticia (el mismo tipo de problema que el toggle de aislamiento
    de red que se encontró y se sacó de Endpoints). Ahora esta ruta
    solo dice "qué debería existir"; la fila real en 'honeyfiles' se
    crea recién cuando el agente la escribe de verdad y lo confirma
    en POST /agent/honeyfile-policy/report."""
    user = require_session_user(request)
    if user is None:
        return JSONResponse({"error": "No autorizado"}, status_code=401)

    if not body:
        return JSONResponse({"error": "Datos inválidos"}, status_code=400)

    display_name = (body.get("file_name") or "").strip()
    file_type = (body.get("file_type") or "txt").strip().lower()
    target_path = (body.get("target_path") or "").strip()
    platform = (body.get("platform") or "all").strip().lower()
    auto_deploy = bool(body.get("auto_deploy"))
    content = (body.get("content") or "").strip()
    agent_ids = body.get("agent_ids", []) or []

    if not display_name:
        return JSONResponse({"error": "Falta el nombre del archivo"}, status_code=400)
    if not target_path:
        return JSONResponse({"error": "Falta la ruta de destino en el cliente"}, status_code=400)
    if platform not in ("windows", "linux", "all"):
        return JSONResponse({"error": "Plataforma objetivo inválida"}, status_code=400)
    if not auto_deploy and not agent_ids:
        return JSONResponse({"error": "Selecciona al menos un endpoint destino, o marca despliegue automático"}, status_code=400)

    if not content:
        # Contenido genérico por defecto -- se muestra editable en el
        # wizard antes de desplegar, no se oculta ni se inventa nada
        # distinto de lo que el analista ve.
        content = "Documento confidencial. No modificar ni distribuir sin autorización."

    ext = f".{file_type}"
    full_file_name = display_name if display_name.lower().endswith(ext) else f"{display_name}{ext}"

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                INSERT INTO honeyfile_templates (
                    name, file_name, file_type, file_path,
                    operating_system, content, auto_deploy, created_by
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id;
                """,
                (
                    display_name, full_file_name, file_type.upper(), target_path,
                    platform.upper(), content, auto_deploy, user["id"]
                )
            )
            template_id = cursor.fetchone()[0]

            assigned_count = 0
            for aid in agent_ids:
                cursor.execute(
                    """
                    INSERT INTO agent_honeyfile_templates (agent_id, template_id, status)
                    VALUES (%s, %s, 'PENDING')
                    ON CONFLICT (agent_id, template_id) DO NOTHING
                    RETURNING id;
                    """,
                    (aid, template_id)
                )
                if cursor.fetchone() is not None:
                    assigned_count += 1

            connection.commit()
    finally:
        connection.close()

    if auto_deploy and not agent_ids:
        message = "Plantilla creada. Se creará sola en cada endpoint compatible la próxima vez que su agente se ejecute."
    elif auto_deploy:
        message = f"Plantilla creada y asignada a {assigned_count} endpoint(s) ahora mismo; también se aplicará sola en cualquier otro endpoint compatible que aparezca después."
    else:
        message = f"Plantilla creada y asignada a {assigned_count} endpoint(s). Se creará cuando el agente correspondiente vuelva a ejecutarse."

    return {
        "success": True,
        "template_id": template_id,
        "assigned_count": assigned_count,
        "message": message
    }


@app.get("/agent/honeyfile-policy")
def get_honeyfile_policy(x_agent_credential: str = Header(...)):
    """El agente llama esto al arrancar Y periódicamente mientras sigue
    corriendo (2026-08-17, ver PENDIENTES.md, "Honeyfiles: despliegue
    automático, rutas, integridad, reconciliación y ejecución en
    tiempo real" -- agent/honeyfile_sync.py) para saber qué honeyfiles
    debería tener en disco.

    Devuelve dos listas:
    - 'pending': asignaciones sin crear todavía (manuales desde el
      Wizard, resueltas ahora mismo desde una plantilla con
      auto_deploy=TRUE, o reintentos de un intento previo que falló).
      El agente tiene que escribirlas y reportarlas.
    - 'existing': asignaciones YA creadas en una ejecución anterior,
      enriquecidas con el contenido/tipo de la plantilla original y el
      hash que el servidor tiene registrado -- no es solo "para saber
      qué vigilar", es lo que el agente necesita para RECONCILIAR en
      cada sincronización (sección 22 de la especificación): si el
      archivo ya no está en disco, puede recrearlo con el mismo
      contenido; si el hash actual no coincide con 'expected_hash',
      puede detectar la alteración sin tener que haber estado
      corriendo en el momento exacto en que ocurrió."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            agent_id = resolve_agent_id(cursor, x_agent_credential)

            cursor.execute(
                """
                SELECT endpoints.os
                FROM agents
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                WHERE agents.id = %s;
                """,
                (agent_id,)
            )
            os_row = cursor.fetchone()
            agent_os = (os_row[0] or "").upper() if os_row else ""

            # Resolver auto_deploy: cualquier plantilla activa marcada
            # auto_deploy=TRUE cuyo SO objetivo coincida (o sea 'ALL')
            # con el de este endpoint, que todavía no tenga una fila de
            # asignación para este agente, se asigna recién ahora --
            # perezoso a propósito: así una plantilla creada después de
            # que el agente ya existía igual lo alcanza en su próxima
            # ejecución, sin tener que sembrar filas por adelantado
            # para agentes que ni siquiera existían todavía.
            cursor.execute(
                """
                SELECT honeyfile_templates.id
                FROM honeyfile_templates
                WHERE honeyfile_templates.is_active = TRUE
                  AND honeyfile_templates.auto_deploy = TRUE
                  AND (honeyfile_templates.operating_system = 'ALL'
                       OR honeyfile_templates.operating_system = %s)
                  AND NOT EXISTS (
                      SELECT 1 FROM agent_honeyfile_templates
                      WHERE agent_honeyfile_templates.template_id = honeyfile_templates.id
                        AND agent_honeyfile_templates.agent_id = %s
                  );
                """,
                (agent_os, agent_id)
            )
            new_auto_template_ids = [r[0] for r in cursor.fetchall()]

            for template_id in new_auto_template_ids:
                cursor.execute(
                    """
                    INSERT INTO agent_honeyfile_templates (agent_id, template_id, status)
                    VALUES (%s, %s, 'PENDING')
                    ON CONFLICT (agent_id, template_id) DO NOTHING;
                    """,
                    (agent_id, template_id)
                )

            if new_auto_template_ids:
                connection.commit()

            # Pendientes de crear (recién asignadas arriba, asignaciones
            # manuales anteriores, o intentos previos que fallaron y se
            # reintentan).
            cursor.execute(
                """
                SELECT agent_honeyfile_templates.id, honeyfile_templates.id,
                       honeyfile_templates.file_name, honeyfile_templates.file_type,
                       honeyfile_templates.file_path, honeyfile_templates.content
                FROM agent_honeyfile_templates
                JOIN honeyfile_templates ON honeyfile_templates.id = agent_honeyfile_templates.template_id
                WHERE agent_honeyfile_templates.agent_id = %s
                  AND agent_honeyfile_templates.status IN ('PENDING', 'FAILED')
                  AND honeyfile_templates.is_active = TRUE;
                """,
                (agent_id,)
            )
            pending = [
                {
                    "assignment_id": r[0],
                    "template_id": r[1],
                    "file_name": r[2],
                    "file_type": r[3],
                    "file_path": r[4],
                    "content": r[5]
                }
                for r in cursor.fetchall()
            ]

            # Ya creados en ejecuciones/ciclos de sincronización
            # anteriores -- enriquecido con contenido y hash esperado
            # para que el agente pueda reconciliar (recrear si falta,
            # detectar si el hash cambió) sin tener que volver a pedir
            # nada aparte. LEFT JOIN a 'honeyfiles' (no INNER): una
            # asignación puede estar en CREATED sin que todavía exista
            # su fila en 'honeyfiles' -- caso borde real (createdo se
            # perdió la respuesta del report), no se descarta la
            # asignación por eso.
            cursor.execute(
                """
                SELECT agent_honeyfile_templates.id, honeyfile_templates.id,
                       honeyfile_templates.file_name, honeyfile_templates.file_type,
                       honeyfile_templates.file_path, honeyfile_templates.content,
                       honeyfiles.id, honeyfiles.file_hash
                FROM agent_honeyfile_templates
                JOIN honeyfile_templates ON honeyfile_templates.id = agent_honeyfile_templates.template_id
                LEFT JOIN honeyfiles ON honeyfiles.agent_id = agent_honeyfile_templates.agent_id
                                     AND honeyfiles.template_id = agent_honeyfile_templates.template_id
                WHERE agent_honeyfile_templates.agent_id = %s
                  AND agent_honeyfile_templates.status = 'CREATED'
                  AND honeyfile_templates.is_active = TRUE;
                """,
                (agent_id,)
            )
            existing = [
                {
                    "assignment_id": r[0],
                    "template_id": r[1],
                    "file_name": r[2],
                    "file_type": r[3],
                    "file_path": r[4],
                    "content": r[5],
                    "honeyfile_id": r[6],
                    "expected_hash": r[7],
                }
                for r in cursor.fetchall()
            ]

            return {"pending": pending, "existing": existing}
    finally:
        connection.close()


@app.get("/agent/rule-policy")
def get_rule_policy(x_agent_credential: str = Header(...)):
    """Agregado 2026-08-12, extendido 2026-08-16 (configuración por
    endpoint, ver PENDIENTES.md): el agente pide esto en cada ejecución
    (sigue siendo un script de una sola pasada sin bucle) para
    enterarse de los valores EFECTIVOS de peso/umbral/ventana de cada
    regla activa PARA ESTE AGENTE puntual -- ya no solo los globales de
    'heuristic_rules', sino el resultado de aplicar el override de
    'agent_rule' si este endpoint tiene uno (ver
    _effective_agent_rules_cte()). Reglas cuyo is_active EFECTIVO sea
    FALSE no se incluyen -- el agente ya no las evalúa en absoluto, en
    vez de evaluarlas con un umbral inalcanzable."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:

            agent_id = resolve_agent_id(cursor, x_agent_credential)

            cursor.execute(
                _effective_agent_rules_cte() + """
                SELECT name, effective_weight, effective_threshold, effective_window_seconds
                FROM effective_rules
                WHERE effective_is_active = TRUE;
                """,
                {"agent_id": agent_id}
            )

            rules = [
                {
                    "name": r[0],
                    "weight": float(r[1]),
                    "threshold": float(r[2]),
                    "window_seconds": r[3]
                }
                for r in cursor.fetchall()
            ]

        return {"rules": rules}
    finally:
        connection.close()


@app.get("/agent/isolation-status")
def get_isolation_status(x_agent_credential: str = Header(...)):
    """El agente llama esto periódicamente (agent/isolation_sync.py,
    2026-08-17, ver PENDIENTES.md, "Corrección definitiva del motor
    heurístico...") para saber si el servidor ordenó aislar (o liberar)
    ESTE endpoint. Mismo patrón de polling que GET
    /agent/honeyfile-policy y GET /agent/rule-policy -- sin
    WebSockets/SSE/brokers, el agente sigue siendo quien pregunta,
    nunca el servidor quien empuja.

    Devuelve como mucho UNA orden pendiente por agente -- ya sea de
    aislar ('REQUESTED') o de liberar ('RELEASE_REQUESTED', sección 18
    de "Aislamiento de host -- modo development, laboratorio y
    producción", 2026-08-17, ver PENDIENTES.md); nunca ambas a la vez,
    porque una orden de liberar solo se crea sobre una fila que ya está
    'EXECUTED', y report_alert()/el endpoint de aislamiento manual ya
    evitan crear una segunda fila 'REQUESTED' mientras una siga
    REQUESTED o EXECUTED. El campo 'action' le dice al agente cuál de
    las dos ejecutar sin tener que inferirlo del estado."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            agent_id = resolve_agent_id(cursor, x_agent_credential)

            cursor.execute(
                """
                SELECT id, isolation_type, reason,
                       CASE WHEN status = 'RELEASE_REQUESTED' THEN 'RELEASE' ELSE 'ISOLATE' END AS action
                FROM host_isolations
                WHERE agent_id = %s AND status IN ('REQUESTED', 'RELEASE_REQUESTED')
                ORDER BY requested_at ASC
                LIMIT 1;
                """,
                (agent_id,)
            )
            row = cursor.fetchone()

        if row is None:
            return {"pending": None}

        return {
            "pending": {
                "isolation_id": row[0],
                "isolation_type": row[1],
                "reason": row[2],
                "action": row[3],
            }
        }
    finally:
        connection.close()


class IsolationStatusReport(BaseModel):
    isolation_id: int
    status: str  # 'EXECUTED'/'ISOLATION_FAILED' (aislar) o 'RELEASED'/'RELEASE_FAILED' (liberar) -- cualquier otro valor se rechaza (ver abajo)
    result: str | None = None


@app.post("/agent/isolation-status/report")
def report_isolation_status(
    report: IsolationStatusReport,
    x_agent_credential: str = Header(...)
):
    """El agente confirma acá el resultado REAL de haber intentado
    ejecutar una orden de aislamiento o de liberación
    (agent/isolation_executor.py) -- recién en este momento la fila de
    'host_isolations' deja de ser una orden pendiente y refleja lo que
    de verdad pasó en el endpoint (sección 28 de la especificación
    original: "agente ejecuta -> agente confirma -> servidor actualiza
    estado"; sección 18 de "Aislamiento de host -- modo development,
    laboratorio y producción", 2026-08-17, ver PENDIENTES.md, extiende
    esto mismo a "UNISOLATE").

    Cuatro resultados posibles, deliberadamente (no se inventa un
    estado intermedio como 'ISOLATING'/'RELEASING' -- la ejecución
    real, sea un comando de firewall real o la simulación de
    desarrollo, es una operación de una sola pasada desde la
    perspectiva del agente, sin un paso intermedio observable que
    valga la pena persistir):
    - 'EXECUTED': se aisló de verdad (o, en desarrollo, se completó el
      flujo simulado sin tocar el firewall real). Se registra
      'executed_at'. Solo válido si la orden pendiente era 'REQUESTED'.
    - 'ISOLATION_FAILED': se intentó aislar y falló (sin privilegios,
      comando no disponible, error real del SO, o la verificación
      posterior no confirmó las reglas -- ver agent/isolation_executor.py)
      -- 'result' trae el motivo. NO se reintenta solo -- la orden
      queda en este estado hasta que un analista decida (mismo criterio
      que 'FAILED' en honeyfiles: no ocultar un fallo real fingiendo
      éxito). Solo válido si la orden pendiente era 'REQUESTED'.
    - 'RELEASED': se liberó de verdad (o, en desarrollo, se completó el
      flujo simulado). Se registra 'released_at'. Solo válido si la
      orden pendiente era 'RELEASE_REQUESTED'.
    - 'RELEASE_FAILED': se intentó liberar y falló -- el endpoint SIGUE
      aislado de verdad (nada cambió), así que la fila vuelve a
      'EXECUTED' en vez de quedar en un estado nuevo ('result' deja
      registrado el motivo del fallo igual). Solo válido si la orden
      pendiente era 'RELEASE_REQUESTED'."""

    if report.status not in ("EXECUTED", "ISOLATION_FAILED", "RELEASED", "RELEASE_FAILED"):
        raise HTTPException(status_code=422, detail="status debe ser 'EXECUTED', 'ISOLATION_FAILED', 'RELEASED' o 'RELEASE_FAILED'")

    is_release_report = report.status in ("RELEASED", "RELEASE_FAILED")
    expected_pending_status = "RELEASE_REQUESTED" if is_release_report else "REQUESTED"

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            agent_id = resolve_agent_id(cursor, x_agent_credential)

            # La orden tiene que ser de ESTE agente y seguir en el
            # estado pendiente que corresponde a este tipo de reporte
            # -- no se confía en que isolation_id venga "limpio", y no
            # tiene sentido reportar el resultado de una orden que otro
            # reporte ya resolvió (evita una carrera de doble reporte),
            # ni confirmar una liberación sobre una orden que en
            # realidad era de aislar (o viceversa).
            cursor.execute(
                "SELECT id FROM host_isolations WHERE id = %s AND agent_id = %s AND status = %s;",
                (report.isolation_id, agent_id, expected_pending_status)
            )
            if cursor.fetchone() is None:
                raise HTTPException(
                    status_code=404,
                    detail=f"No hay una orden {expected_pending_status} con ese id para este agente"
                )

            if report.status == "EXECUTED":
                cursor.execute(
                    "UPDATE host_isolations SET status = 'EXECUTED', executed_at = CURRENT_TIMESTAMP, result = %s WHERE id = %s;",
                    (report.result, report.isolation_id)
                )
            elif report.status == "ISOLATION_FAILED":
                cursor.execute(
                    "UPDATE host_isolations SET status = 'ISOLATION_FAILED', result = %s WHERE id = %s;",
                    (report.result, report.isolation_id)
                )
            elif report.status == "RELEASED":
                cursor.execute(
                    "UPDATE host_isolations SET status = 'RELEASED', released_at = CURRENT_TIMESTAMP, result = %s WHERE id = %s;",
                    (report.result, report.isolation_id)
                )
            else:  # RELEASE_FAILED -- sigue aislado de verdad, vuelve a EXECUTED (ver docstring)
                cursor.execute(
                    "UPDATE host_isolations SET status = 'EXECUTED', result = %s WHERE id = %s;",
                    (report.result, report.isolation_id)
                )

            connection.commit()

        return {"message": "Resultado de aislamiento registrado", "status": report.status}
    finally:
        connection.close()


@app.post("/incidents/{incident_id}/isolate")
def isolate_incident_manually(incident_id: int, user: dict = Depends(get_current_user)):
    """Disparo MANUAL de aislamiento desde la consola (botón "Aislar",
    2026-08-17, ver PENDIENTES.md, "Aislamiento de host -- modo
    development, laboratorio y producción", sección 20: "debe utilizar
    exactamente el mismo mecanismo de backend/agente que el aislamiento
    automático... solo cambia el origen de la orden").

    Deliberadamente NO es una segunda implementación: inserta la MISMA
    fila 'REQUESTED' en 'host_isolations' que report_alert() inserta
    para el camino automático (sección 30 de la especificación
    original), con el mismo guard de no-duplicar y el mismo agente de
    endpoint (agent/isolation_sync.py + agent/isolation_executor.py)
    recogiéndola y ejecutándola de verdad. La única diferencia real es
    'requested_by' -- acá SÍ se completa (columna que ya existía en el
    schema, sin usar hasta ahora porque el camino automático no tiene
    un usuario detrás)."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:

            cursor.execute("SELECT agent_id, status FROM incidents WHERE id = %s;", (incident_id,))
            row = cursor.fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Incidente no encontrado")
            agent_id, incident_status = row

            if incident_status == "CLOSED":
                raise HTTPException(status_code=409, detail="No se puede aislar un endpoint por un incidente ya cerrado")

            # Mismo guard que usa report_alert() para el camino
            # automático (sección 17 de la especificación original: "no
            # duplicar evidencia/órdenes ya registradas"; sección 27 de
            # la de host: "pulsar Aislar dos veces... no debe romper el
            # sistema") -- si ya hay una orden en curso o cumplida para
            # este ENDPOINT, no se crea una segunda.
            #
            # BUG REAL corregido 2026-08-18 (problema H/J, ver
            # PENDIENTES.md, "Revisión y corrección integral de
            # ALFA-Sentinel"): este guard filtraba antes por
            # 'incident_id = %s' (el incidente puntual desde el que se
            # clickeó "Aislar"), no por el agente. Un mismo endpoint con
            # varios incidentes (PC-01 -> INC-001, INC-002, INC-003)
            # podía aislarse desde INC-001 y, como el guard consultaba
            # 'incident_id = INC-002.id' (que nunca tiene esa fila --
            # está en INC-001.id), un clic en "Aislar" desde INC-002
            # pasaba este chequeo sin problema e insertaba una SEGUNDA
            # orden 'REQUESTED' real para un endpoint que el agente ya
            # estaba aislando o ya había aislado -- exactamente la prueba
            # obligatoria Q/14 del usuario ("no se puede ejecutar
            # aislamiento dos veces sobre el mismo endpoint"). Ahora usa
            # _agent_is_isolated_sql(), la misma fuente única que ya usan
            # todas las pantallas para decidir si mostrar "Aislar" o
            # "Endpoint aislado".
            cursor.execute(
                f"SELECT {_agent_isolation_status_sql('%s')} WHERE {_agent_is_isolated_sql('%s')};",
                (agent_id, agent_id)
            )
            existing = cursor.fetchone()
            if existing is not None:
                raise HTTPException(
                    status_code=409,
                    detail=f"Este endpoint ya tiene una orden de aislamiento en curso o cumplida (estado actual: {ISOLATION_STATUS_LABELS_ES.get(existing[0], existing[0])})."
                )

            reason = f"Aislamiento manual solicitado por {user['full_name']} desde la consola."

            cursor.execute(
                """
                INSERT INTO host_isolations (agent_id, incident_id, isolation_type, status, reason, requested_by)
                VALUES (%s, %s, 'NETWORK', 'REQUESTED', %s, %s)
                RETURNING id;
                """,
                (agent_id, incident_id, reason, user["id"])
            )
            isolation_id = cursor.fetchone()[0]

            log_audit(cursor, user["id"], "MANUAL_ISOLATE_REQUEST", "host_isolations", isolation_id, reason)

            connection.commit()

        return {"message": "Orden de aislamiento enviada -- el agente del endpoint la ejecutará en breve", "isolation_id": isolation_id, "status": "REQUESTED"}
    finally:
        connection.close()


@app.post("/host-isolations/{isolation_id}/release")
def release_host_isolation(isolation_id: int, user: dict = Depends(get_current_user)):
    """Operación inversa -- UNISOLATE (sección 18 de "Aislamiento de
    host -- modo development, laboratorio y producción", 2026-08-17,
    ver PENDIENTES.md): "CONSOLA -> SERVIDOR -> AGENTE -> UNISOLATE ->
    restaurar comunicación -> confirmación". Mismo mecanismo de
    polling/ejecución/confirmación que aislar (GET
    /agent/isolation-status, agent/isolation_sync.py,
    agent/isolation_executor.py::execute_release()), nunca recupera
    archivos -- solo restaura el estado de red."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:

            cursor.execute("SELECT status FROM host_isolations WHERE id = %s;", (isolation_id,))
            row = cursor.fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Orden de aislamiento no encontrada")

            if row[0] != "EXECUTED":
                raise HTTPException(
                    status_code=409,
                    detail=f"Solo se puede liberar un aislamiento con estado 'Ejecutado' (estado actual: {ISOLATION_STATUS_LABELS_ES.get(row[0], row[0])})."
                )

            cursor.execute(
                "UPDATE host_isolations SET status = 'RELEASE_REQUESTED' WHERE id = %s;",
                (isolation_id,)
            )

            log_audit(cursor, user["id"], "MANUAL_RELEASE_REQUEST", "host_isolations", isolation_id, f"Liberación solicitada por {user['full_name']}.")

            connection.commit()

        return {"message": "Orden de liberación enviada -- el agente del endpoint la ejecutará en breve", "isolation_id": isolation_id, "status": "RELEASE_REQUESTED"}
    finally:
        connection.close()


class HoneyfileReportItem(BaseModel):
    assignment_id: int
    status: str
    file_path: str | None = None
    file_name: str | None = None
    file_type: str | None = None
    file_hash: str | None = None
    error: str | None = None


class HoneyfileReport(BaseModel):
    results: list[HoneyfileReportItem]


@app.post("/agent/honeyfile-policy/report")
def report_honeyfile_policy(
    report: HoneyfileReport,
    x_agent_credential: str = Header(...)
):
    """El agente confirma acá qué pudo crear/reconciliar de verdad y qué
    no. Recién en este momento aparece o se actualiza la fila real en
    'honeyfiles' -- antes de esto, un honeyfile 'PENDING' es solo una
    intención, no un archivo que exista en ningún disco.

    Tres estados posibles por ítem (2026-08-17, ver PENDIENTES.md,
    "Honeyfiles: despliegue automático, rutas, integridad,
    reconciliación y ejecución en tiempo real"):

    - 'CREATED': se escribió el archivo (primera vez, o se recreó
      porque había desaparecido -- reconciliación caso B). UPSERT
      contra (agent_id, template_id) en vez de INSERT liso: esto se
      llama tanto al crear por primera vez como en cada ciclo de
      sincronización que vuelve a verificar honeyfiles ya existentes
      (ver GET /agent/honeyfile-policy, lista 'existing'), así que un
      INSERT sin ON CONFLICT duplicaría la fila la segunda vez que el
      agente reconfirma un honeyfile que ya tenía.
    - 'MODIFIED': el hash real en disco ya no coincide con el
      'expected_hash' que el servidor le pasó (reconciliación caso C).
      Solo se actualiza el hash registrado -- nunca se restaura el
      contenido ni se toca el estado de la asignación, que sigue
      'CREATED' (el archivo existe, solo que alguien lo alteró; eso lo
      capta HR-03 si watchdog lo vio en vivo, no es responsabilidad de
      este endpoint recuperarlo).
    - Cualquier otro valor ('FAILED'): no se pudo crear/recrear.
      'agent_honeyfile_templates' pasa a FAILED, así que el próximo
      GET /agent/honeyfile-policy lo vuelve a ofrecer como 'pending'
      (ver ese endpoint, WHERE status IN ('PENDING','FAILED')) -- sirve
      igual para un fallo de creación inicial que para un fallo al
      intentar recrear un honeyfile que había desaparecido."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            agent_id = resolve_agent_id(cursor, x_agent_credential)

            created_count = 0
            modified_count = 0
            failed_count = 0

            for item in report.results:
                # La asignación tiene que ser de este agente -- no se
                # confía en que assignment_id venga "limpio". Se trae
                # también template_id: es la clave real de UPSERT
                # contra 'honeyfiles' (UNIQUE(agent_id, template_id),
                # ver database/schema.sql y la migración
                # 2026-08-17_honeyfiles_template_id.sql), nunca se
                # confía en un template_id que mandara el agente.
                cursor.execute(
                    """
                    SELECT id, template_id FROM agent_honeyfile_templates
                    WHERE id = %s AND agent_id = %s;
                    """,
                    (item.assignment_id, agent_id)
                )
                row = cursor.fetchone()
                if row is None:
                    continue
                template_id = row[1]

                if item.status == "CREATED":
                    cursor.execute(
                        """
                        INSERT INTO honeyfiles (
                            agent_id, template_id, file_path, file_name,
                            file_type, file_hash, status, last_checked_at
                        )
                        VALUES (%s, %s, %s, %s, %s, %s, 'ACTIVE', CURRENT_TIMESTAMP)
                        ON CONFLICT (agent_id, template_id) DO UPDATE SET
                            file_path = EXCLUDED.file_path,
                            file_name = EXCLUDED.file_name,
                            file_type = EXCLUDED.file_type,
                            file_hash = EXCLUDED.file_hash,
                            status = 'ACTIVE',
                            last_checked_at = CURRENT_TIMESTAMP;
                        """,
                        (
                            agent_id, template_id, item.file_path, item.file_name,
                            item.file_type, item.file_hash
                        )
                    )
                    cursor.execute(
                        """
                        UPDATE agent_honeyfile_templates
                        SET status = 'CREATED', updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s;
                        """,
                        (item.assignment_id,)
                    )
                    created_count += 1

                elif item.status == "MODIFIED":
                    cursor.execute(
                        """
                        UPDATE honeyfiles
                        SET file_hash = %s, last_checked_at = CURRENT_TIMESTAMP
                        WHERE agent_id = %s AND template_id = %s;
                        """,
                        (item.file_hash, agent_id, template_id)
                    )
                    modified_count += 1

                else:
                    cursor.execute(
                        """
                        UPDATE agent_honeyfile_templates
                        SET status = 'FAILED', updated_at = CURRENT_TIMESTAMP
                        WHERE id = %s;
                        """,
                        (item.assignment_id,)
                    )
                    failed_count += 1

            connection.commit()

            return {
                "message": "Reporte de honeyfiles procesado",
                "created_count": created_count,
                "modified_count": modified_count,
                "failed_count": failed_count
            }
    finally:
        connection.close()


@app.get("/api/users")
def api_users(user: dict = Depends(get_current_user)):
    """Datos de la subsección Usuarios y Roles de Administración en
    React. Cualquier sesión válida puede leer esta lista (gap ya
    documentado: solo crear/editar exige rol admin, ver POST /users y
    PATCH /users/{id})."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT users.id, users.username, users.full_name, users.email,
                       STRING_AGG(roles.name, ', ' ORDER BY roles.name) AS roles,
                       users.is_active, users.last_login_at, users.created_at
                FROM users
                LEFT JOIN user_roles ON user_roles.user_id = users.id
                LEFT JOIN roles ON roles.id = user_roles.role_id
                GROUP BY users.id
                ORDER BY users.id;
                """
            )
            rows = cursor.fetchall()
    finally:
        connection.close()

    return {
        "is_admin": "admin" in user.get("roles", []),
        "users": [
            {
                "id": r[0],
                "username": r[1],
                "full_name": r[2],
                "email": r[3],
                "roles": r[4],
                "is_active": r[5],
                "last_login_at": r[6].strftime("%d/%m/%Y %H:%M:%S") if r[6] else None,
                "created_at": r[7].strftime("%d/%m/%Y %H:%M:%S") if r[7] else None,
            }
            for r in rows
        ],
    }


@app.get("/alerts/open")
def alerts_open(user: dict = Depends(get_current_user)):
    """JSON liviano para la campanita de notificaciones -- se consulta
    solo, sin pasar por cada ruta de página. Trae las alertas todavía
    sin revisar (status = 'NEW'), sin importar la severidad: hasta las
    de severidad BAJO quedan fuera porque esas ni siquiera generan alerta
    (el agente solo llama a send_alert cuando is_suspicious() es
    True).

    Extendido 2026-08-17 (ver PENDIENTES.md, "Alertas flotantes
    globales de alta prioridad"): además de lo que ya usaba la
    campanita, ahora trae 'risk_score', 'incident_id' y
    'isolation_status' -- lo que necesita la capa de notificaciones
    flotantes globales para decidir prioridad visual (CRÍTICO vs ALTO),
    armar el enlace "Ver incidente" cuando ya existe uno, y mostrar el
    estado REAL de una acción de aislamiento (nunca 'RECOMMENDED' como
    si fuera el resultado final, sección 23 de esa especificación). Se
    reutiliza este mismo endpoint en vez de crear uno nuevo a propósito
    -- misma fuente de datos, una sola consulta, sin duplicar lo que la
    campanita ya pedía (sección 19: "no crear una petición
    independiente" por función nueva)."""

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            # 'isolation_status' corregido 2026-08-18 (problema H, ver
            # PENDIENTES.md): antes filtraba por
            # 'host_isolations.incident_id = alerts.incident_id' -- si el
            # endpoint fue aislado desde OTRO incidente (o la alerta
            # todavía no tiene incident_id porque no escaló), esta
            # columna quedaba NULL aunque el endpoint SÍ estuviera
            # aislado, y la notificación flotante/campanita ofrecían
            # "Aislar" sobre un endpoint ya aislado. Ver
            # _agent_isolation_status_sql() -- misma fuente que
            # COMBINED_CTE, get_incidente_drawer() y GET /api/respuesta.
            cursor.execute(
                f"""
                SELECT alerts.id, severity_levels.name, alerts.title,
                       endpoints.hostname, alerts.created_at,
                       alerts.risk_score, alerts.incident_id,
                       {_agent_isolation_status_sql("agents.id")} AS isolation_status
                FROM alerts
                JOIN agents ON agents.id = alerts.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                JOIN severity_levels ON severity_levels.id = alerts.severity_id
                WHERE alerts.status = 'NEW'
                ORDER BY alerts.created_at DESC
                LIMIT 10;
                """
            )

            rows = cursor.fetchall()

            cursor.execute(
                "SELECT COUNT(*) FROM alerts WHERE status = 'NEW';"
            )

            total = cursor.fetchone()[0]

    finally:
        connection.close()

    return {
        "count": total,
        "alerts": [
            {
                "id": row[0],
                "severity": row[1],
                # Título general por severidad, no el nombre de la
                # primera regla que llegó -- ver alert_general_title()
                # (2026-08-18, ver PENDIENTES.md, "Corrección definitiva
                # en la lógica y presentación de ALERTAS"). Esto alimenta
                # tanto la campana (NotificationsBell.tsx) como la
                # notificación flotante global (FloatingAlertCard.tsx) --
                # una sola fuente para ambas.
                "title": alert_general_title(row[1]),
                "hostname": row[3],
                "created_at": row[4].strftime("%d/%m %H:%M"),
                "risk_score": float(row[5]),
                "incident_id": row[6],
                "isolation_status": row[7],
            }
            for row in rows
        ]
    }


# ============================================================
# API JSON para el Panel de Control en React (2026-08-14, hoy la
# única interfaz -- el dashboard Jinja2 que calculaba estas mismas
# cifras para HTML se eliminó junto con el resto del frontend viejo,
# ver PENDIENTES.md). Mismas consultas y mismo criterio de
# "abierto"/"en línea" que se usaban ahí; si mañana cambia una regla
# de negocio acá (por ejemplo, qué cuenta como "endpoint en riesgo"),
# es la única fuente de verdad -- ya no hay un segundo camino de
# código en paralelo que replicar.
# ============================================================

# Escala de color fija -- esto SÍ es presentación legítima (el nombre
# y el rango de cada nivel salen de severity_levels, pero qué color
# hex le corresponde a cada uno es una decisión de diseño que no vive
# en ninguna tabla). Las claves son los 'name' reales de
# severity_levels tras la migración a español (ver
# database/migration_2026-08-16_severity_levels_espanol.sql).
RISK_COLOR_HEX = {
    "BAJO": "#16a34a",
    "MEDIO": "#ca8a04",
    "ALTO": "#ea580c",
    "CRÍTICO": "#dc2626",
}


@app.get("/api/dashboard/overview")
def api_dashboard_overview(user: dict = Depends(get_current_user)):
    """Snapshot completo para el Panel de Control en React. Pensado
    para pedirse en el primer render y volver a pedirse entero cada
    tantos segundos (polling, igual que /dashboard/live) -- no hay
    websockets ni push real desde el agente."""

    try:
        connection = get_connection()
    except Exception:
        return {"db_ok": False}

    try:
        with connection.cursor() as cursor:

            # --- KPIs de endpoints ---
            cursor.execute("SELECT COUNT(*) FROM agents;")
            endpoints_total = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM agents WHERE status = 'ONLINE';")
            endpoints_online = cursor.fetchone()[0]

            # 'Aislados': real desde la corrección definitiva del motor
            # heurístico (2026-08-17, ver PENDIENTES.md) -- cuenta
            # endpoints con una orden de aislamiento vigente (REQUESTED
            # o EXECUTED) todavía no liberada. Filtro de status
            # agregado en esa misma corrección: sin él, una orden que
            # el agente reportó como ISOLATION_FAILED (nunca se aisló
            # de verdad) también tiene released_at NULL y se hubiera
            # contado como "aislado" -- mismo criterio que el resto de
            # las consultas de is_isolated en este archivo.
            cursor.execute(
                "SELECT COUNT(DISTINCT agent_id) FROM host_isolations WHERE status IN ('REQUESTED', 'EXECUTED', 'RELEASE_REQUESTED') AND released_at IS NULL;"
            )
            endpoints_isolated = cursor.fetchone()[0]

            endpoints_offline = max(endpoints_total - endpoints_online - endpoints_isolated, 0)

            # --- Alertas activas + tendencia vs. periodo anterior (24h) ---
            cursor.execute("SELECT COUNT(*) FROM alerts WHERE status = 'NEW';")
            alerts_active = cursor.fetchone()[0]

            cursor.execute(
                "SELECT COUNT(*) FROM alerts WHERE created_at >= CURRENT_TIMESTAMP - INTERVAL '24 hours';"
            )
            alerts_last_24h = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT COUNT(*) FROM alerts
                WHERE created_at >= CURRENT_TIMESTAMP - INTERVAL '48 hours'
                  AND created_at < CURRENT_TIMESTAMP - INTERVAL '24 hours';
                """
            )
            alerts_prev_24h = cursor.fetchone()[0]

            alerts_trend_pct = (
                round(((alerts_last_24h - alerts_prev_24h) / alerts_prev_24h) * 100, 1)
                if alerts_prev_24h > 0 else None
            )

            # --- Incidentes activos ---
            cursor.execute("SELECT COUNT(*) FROM incidents WHERE status = 'OPEN';")
            incidents_active = cursor.fetchone()[0]

            # --- Honeyfiles activados hoy (vía alerts+alert_rule -- la
            # tabla honeyfile_activations existe pero nada la escribe
            # todavía, ver PENDIENTES.md) ---
            cursor.execute(
                """
                SELECT COUNT(*)
                FROM alerts
                JOIN alert_rule ON alert_rule.alert_id = alerts.id
                JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                WHERE heuristic_rules.name = 'Acceso Honeyfile'
                  AND alerts.created_at >= date_trunc('day', CURRENT_TIMESTAMP);
                """
            )
            honeyfiles_activated_today = cursor.fetchone()[0]

            # Catálogo real de severity_levels, ordenado por min_score
            # (Bajo -> Crítico) -- fuente única para todo lo que sigue
            # en esta función: nombres, orden y color. Nada de esto se
            # hardcodea en Python.
            cursor.execute("SELECT id, name, min_score FROM severity_levels ORDER BY min_score;")
            severity_catalog_rows = cursor.fetchall()
            severity_names_by_min_score = {row[2]: row[1] for row in severity_catalog_rows}
            lowest_severity_name = severity_catalog_rows[0][1]

            # --- Distribución de riesgo por endpoint (peor severidad
            # entre sus alertas abiertas; el nivel más bajo del
            # catálogo = sin ninguna alerta abierta) ---
            cursor.execute(
                """
                SELECT severity_levels.name, COUNT(DISTINCT alerts.agent_id)
                FROM alerts
                JOIN severity_levels ON severity_levels.id = alerts.severity_id
                WHERE alerts.status = 'NEW'
                GROUP BY severity_levels.name;
                """
            )
            severity_rows = dict(cursor.fetchall())

            cursor.execute("SELECT COUNT(DISTINCT agent_id) FROM alerts WHERE status = 'NEW';")
            agents_with_open_alerts = cursor.fetchone()[0]
            severity_rows[lowest_severity_name] = severity_rows.get(lowest_severity_name, 0) + max(
                endpoints_total - agents_with_open_alerts, 0
            )

            risk_distribution = [
                {"level": name, "count": severity_rows.get(name, 0), "color": RISK_COLOR_HEX.get(name, "#6b7280")}
                for _, name, _ in severity_catalog_rows
            ]

            # --- Endpoints que requieren atención -- MAX(min_score) en
            # vez de un CASE con nombres hardcodeados: el "peor" nivel
            # de cada endpoint se calcula con el orden real del
            # catálogo (severity_names_by_min_score de arriba). ---
            cursor.execute(
                """
                SELECT endpoints.hostname, endpoints.os, agents.status, agents.last_seen_at,
                       MAX(severity_levels.min_score) AS risk_min_score,
                       COUNT(*) AS alert_count
                FROM alerts
                JOIN agents ON agents.id = alerts.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                JOIN severity_levels ON severity_levels.id = alerts.severity_id
                WHERE alerts.status = 'NEW'
                GROUP BY endpoints.hostname, endpoints.os, agents.status, agents.last_seen_at
                ORDER BY risk_min_score DESC, alert_count DESC
                LIMIT 5;
                """
            )
            endpoints_at_risk = [
                {
                    "hostname": r[0],
                    "os": r[1],
                    "status": r[2],
                    "last_seen_ago": time_ago(r[3]),
                    "severity": severity_names_by_min_score[r[4]],
                    "alerts_count": r[5],
                }
                for r in cursor.fetchall()
            ]

            # --- Alertas recientes. 'process' va explícitamente null:
            # 'alerts' no tiene columna de proceso -- el agente no
            # correla eventos con el proceso responsable todavía (ver
            # PENDIENTES.md, "Atribución de proceso"). No se inventa. ---
            cursor.execute(
                """
                SELECT alerts.id, severity_levels.name, alerts.title,
                       endpoints.hostname, alerts.created_at, alerts.status
                FROM alerts
                JOIN agents ON agents.id = alerts.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                JOIN severity_levels ON severity_levels.id = alerts.severity_id
                ORDER BY alerts.created_at DESC
                LIMIT 6;
                """
            )
            recent_alerts = [
                {
                    "id": r[0],
                    "severity": r[1],
                    # Título general por severidad -- ver alert_general_title()
                    # (2026-08-18, ver PENDIENTES.md).
                    "title": alert_general_title(r[1]),
                    "hostname": r[3],
                    "process": None,
                    "time": r[4].strftime("%H:%M"),
                    "status": ALERT_STATUS_LABELS_ES.get(r[5], r[5]),
                }
                for r in cursor.fetchall()
            ]

            # --- Honeyfiles: resumen + últimas activaciones. El
            # nombre de archivo puntual no está disponible todavía: la
            # alerta de "Acceso Honeyfile" no incluye qué archivo fue
            # (el agente no lo manda en el payload hoy), así que
            # 'file_name' va null en vez de inventado -- ver
            # PENDIENTES.md.
            cursor.execute("SELECT COUNT(*) FROM honeyfiles WHERE status = 'ACTIVE';")
            honeyfiles_active_total = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT endpoints.hostname, alerts.created_at
                FROM alerts
                JOIN alert_rule ON alert_rule.alert_id = alerts.id
                JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                JOIN agents ON agents.id = alerts.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                WHERE heuristic_rules.name = 'Acceso Honeyfile'
                ORDER BY alerts.created_at DESC
                LIMIT 5;
                """
            )
            honeyfile_recent = [
                {"hostname": r[0], "time": r[1].strftime("%H:%M"), "file_name": None}
                for r in cursor.fetchall()
            ]

            # --- Estado de endpoints (online/offline/aislados) + salud
            # de agentes, mismo criterio de conectividad que el resto
            # de la consola (agents.status, no heartbeat exacto). ---
            agent_health_pct = (
                round((endpoints_online / endpoints_total) * 100, 1)
                if endpoints_total > 0 else 100.0
            )

            # --- Principales tipos de detección (vectores) ---
            cursor.execute(
                """
                SELECT heuristic_rules.name, COUNT(*) AS n
                FROM alerts
                LEFT JOIN alert_rule ON alert_rule.alert_id = alerts.id
                LEFT JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                WHERE heuristic_rules.name IS NOT NULL
                GROUP BY heuristic_rules.name
                ORDER BY n DESC;
                """
            )
            top_detections = [
                {"rule_name": r[0], "rule_label": r[0], "count": r[1]}
                for r in cursor.fetchall()
            ]

            # --- Actividad reciente (feed unificado: alertas + eventos
            # crudos). Solo tipos reales -- no se agregan entradas como
            # "endpoint aislado" porque nada en el sistema genera eso
            # todavía. ---
            cursor.execute(
                """
                (
                    SELECT 'alert' AS kind, severity_levels.name AS sev,
                           alerts.title AS label, endpoints.hostname AS hostname,
                           alerts.created_at AS ts
                    FROM alerts
                    JOIN agents ON agents.id = alerts.agent_id
                    JOIN endpoints ON endpoints.id = agents.endpoint_id
                    JOIN severity_levels ON severity_levels.id = alerts.severity_id
                    ORDER BY alerts.created_at DESC
                    LIMIT 10
                )
                UNION ALL
                (
                    SELECT 'honeyfile_created' AS kind, NULL AS sev,
                           honeyfiles.file_name AS label, endpoints.hostname AS hostname,
                           honeyfiles.created_at AS ts
                    FROM honeyfiles
                    JOIN agents ON agents.id = honeyfiles.agent_id
                    JOIN endpoints ON endpoints.id = agents.endpoint_id
                    ORDER BY honeyfiles.created_at DESC
                    LIMIT 5
                )
                UNION ALL
                (
                    SELECT 'endpoint_registered' AS kind, NULL AS sev,
                           endpoints.hostname AS label, endpoints.hostname AS hostname,
                           agents.enrolled_at AS ts
                    FROM agents
                    JOIN endpoints ON endpoints.id = agents.endpoint_id
                    ORDER BY agents.enrolled_at DESC
                    LIMIT 5
                )
                ORDER BY ts DESC
                LIMIT 10;
                """
            )
            RECENT_ACTIVITY_LABELS = {
                "alert": "Nueva alerta",
                "honeyfile_created": "Honeyfile creado",
                "endpoint_registered": "Endpoint registrado",
            }
            recent_activity = [
                {
                    "kind": r[0],
                    "severity": r[1],
                    "type_label": RECENT_ACTIVITY_LABELS.get(r[0], r[0]),
                    # Para 'alert', título general por severidad -- ver
                    # alert_general_title() (2026-08-18, ver
                    # PENDIENTES.md) -- en vez de 'alerts.title' crudo
                    # (el texto del primer evento del episodio). Los
                    # demás tipos de esta actividad ('honeyfile_created'/
                    # 'endpoint_registered') no son alertas, su 'label'
                    # sigue siendo el que ya traía la consulta.
                    "label": alert_general_title(r[1]) if r[0] == "alert" else r[2],
                    "hostname": r[3],
                    "time": r[4].strftime("%H:%M"),
                    "ago": time_ago(r[4]),
                }
                for r in cursor.fetchall()
            ]

            # --- Estado del sistema. 'api_ok'/'db_ok' son ciertos por
            # haber llegado hasta acá sin excepción. 'detection_engine_ok'
            # es un proxy real (hay al menos una regla activa), no una
            # métrica de salud del proceso del agente en sí -- eso no
            # se puede ver desde el servidor. ---
            cursor.execute("SELECT COUNT(*) FROM heuristic_rules WHERE is_active = TRUE;")
            active_rules_count = cursor.fetchone()[0]

            cursor.execute("SELECT MAX(last_seen_at) FROM agents;")
            last_sync = cursor.fetchone()[0]

    finally:
        connection.close()

    return {
        "db_ok": True,
        "generated_at": datetime.now().strftime("%H:%M:%S"),
        "summary": {
            "endpoints_total": endpoints_total,
            "endpoints_online": endpoints_online,
            "endpoints_offline": endpoints_offline,
            "endpoints_isolated": endpoints_isolated,
            "alerts_active": alerts_active,
            "alerts_trend_pct": alerts_trend_pct,
            "incidents_active": incidents_active,
            "honeyfiles_activated_today": honeyfiles_activated_today,
        },
        "risk_distribution": risk_distribution,
        "endpoints_at_risk": endpoints_at_risk,
        "recent_alerts": recent_alerts,
        "honeyfile_activity": {
            "active_total": honeyfiles_active_total,
            "activated_today": honeyfiles_activated_today,
            "recent": honeyfile_recent,
        },
        "endpoint_status": {
            "online": endpoints_online,
            "offline": endpoints_offline,
            "isolated": endpoints_isolated,
            "agent_health_pct": agent_health_pct,
        },
        "top_detections": top_detections,
        "recent_activity": recent_activity,
        "system_status": {
            "api_ok": True,
            "db_ok": True,
            "agents_comm_ok": endpoints_online > 0,
            "detection_engine_ok": active_rules_count > 0,
            "agents_connected": endpoints_online,
            "agents_total": endpoints_total,
            "last_sync_ago": time_ago(last_sync),
        },
    }


@app.get("/api/dashboard/activity-series")
def api_dashboard_activity_series(
    period: str = Query("24h", pattern="^(24h|7d|30d)$"),
    user: dict = Depends(get_current_user)
):
    """Serie de tiempo para el gráfico grande de 'Actividad de
    seguridad'. Cuatro series, todas de tablas reales: alertas
    (alerts.created_at), actividad de archivos en general
    (events.detected_at -- volumen crudo, no solo lo sospechoso),
    incidentes abiertos (incidents.opened_at) y activaciones de
    honeyfile (mismo camino vía alert_rule que el resto del
    dashboard). 24h se agrupa por hora, 7d/30d por día.

    Nota: el parámetro se llama 'period', no 'range' -- 'range' tapa
    la función builtin de Python del mismo nombre, que se usa más
    abajo para armar los buckets (bug real, encontrado al probar esto
    contra Postgres de verdad)."""

    if period == "24h":
        interval, trunc_unit, bucket_count, fmt = "24 hours", "hour", 24, "%H:00"
    elif period == "7d":
        interval, trunc_unit, bucket_count, fmt = "7 days", "day", 7, "%d/%m"
    else:
        interval, trunc_unit, bucket_count, fmt = "30 days", "day", 30, "%d/%m"

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                f"""
                SELECT date_trunc('{trunc_unit}', created_at) AS bucket, COUNT(*)
                FROM alerts
                WHERE created_at >= CURRENT_TIMESTAMP - INTERVAL '{interval}'
                GROUP BY bucket;
                """
            )
            alerts_by_bucket = dict(cursor.fetchall())

            cursor.execute(
                f"""
                SELECT date_trunc('{trunc_unit}', detected_at) AS bucket, COUNT(*)
                FROM events
                WHERE detected_at >= CURRENT_TIMESTAMP - INTERVAL '{interval}'
                GROUP BY bucket;
                """
            )
            activity_by_bucket = dict(cursor.fetchall())

            cursor.execute(
                f"""
                SELECT date_trunc('{trunc_unit}', opened_at) AS bucket, COUNT(*)
                FROM incidents
                WHERE opened_at >= CURRENT_TIMESTAMP - INTERVAL '{interval}'
                GROUP BY bucket;
                """
            )
            incidents_by_bucket = dict(cursor.fetchall())

            cursor.execute(
                f"""
                SELECT date_trunc('{trunc_unit}', alerts.created_at) AS bucket, COUNT(*)
                FROM alerts
                JOIN alert_rule ON alert_rule.alert_id = alerts.id
                JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                WHERE heuristic_rules.name = 'Acceso Honeyfile'
                  AND alerts.created_at >= CURRENT_TIMESTAMP - INTERVAL '{interval}'
                GROUP BY bucket;
                """
            )
            honeyfiles_by_bucket = dict(cursor.fetchall())

    finally:
        connection.close()

    now = datetime.now()
    if trunc_unit == "hour":
        now_bucket = now.replace(minute=0, second=0, microsecond=0)
        step = timedelta(hours=1)
    else:
        now_bucket = now.replace(hour=0, minute=0, second=0, microsecond=0)
        step = timedelta(days=1)

    def _lookup(bucket_dict, ts):
        for key, value in bucket_dict.items():
            key_naive = key.replace(tzinfo=None) if key.tzinfo else key
            if key_naive == ts:
                return value
        return 0

    points = []
    for i in range(bucket_count - 1, -1, -1):
        ts = now_bucket - (step * i)
        points.append({
            "bucket": ts.strftime(fmt),
            "alerts": _lookup(alerts_by_bucket, ts),
            "activity": _lookup(activity_by_bucket, ts),
            "incidents": _lookup(incidents_by_bucket, ts),
            "honeyfiles": _lookup(honeyfiles_by_bucket, ts),
        })

    return {"range": period, "points": points}


@app.get("/api/endpoints")
def api_endpoints(
    search: str = "",
    status: str = Query("", pattern="^(ONLINE|OFFLINE|ISOLATED|)$"),
    risk: str = Query("", pattern="^(BAJO|MEDIO|ALTO|CRÍTICO|)$"),
    os_family: str = "",
    page: int = Query(1, ge=1),
    page_size: int = Query(10, ge=1, le=100),
    user: dict = Depends(get_current_user)
):
    """Lista de endpoints para la pantalla Endpoints en React. Usa el
    mismo _endpoint_cte() que el resto de las vistas de riesgo por
    endpoint (mismo criterio en todo el sistema, un solo camino de
    código).

    Diferencia real a propósito: acá "Estado" tiene un solo valor de
    cara al usuario (ONLINE/OFFLINE/ISOLATED, con ISOLATED con
    prioridad sobre el estado de conexión crudo), en vez de las tres
    categorías ok/attention/offline de /endpoints -- así se pidió esta
    pantalla. El estado de conexión más fino (Healthy/Warning/Offline)
    se conserva aparte como "agent_health", derivado de status_bucket
    (ok->HEALTHY, attention->WARNING, offline->OFFLINE). Riesgo sigue
    siendo un eje aparte (Normal/Sospechoso/Alto/Crítico) -- nunca se
    mezcla con conectividad, igual que en el resto del sistema.
    """
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            stale_seconds = get_agent_stale_seconds(cursor)

            cte = _endpoint_cte(stale_seconds) + f"""
            , endpoint_full AS (
                SELECT
                    endpoint_data.*,
                    {_agent_is_isolated_sql("endpoint_data.id")} AS is_isolated,
                    (
                        SELECT COUNT(*) FROM alerts a
                        WHERE a.agent_id = endpoint_data.id AND a.status = 'NEW'
                    ) AS alerts_count,
                    (
                        SELECT MAX(e.detected_at) FROM events e
                        WHERE e.agent_id = endpoint_data.id
                    ) AS last_activity
                FROM endpoint_data
            ),
            endpoint_view AS (
                SELECT
                    id, hostname, operating_system, os_version, ip_address,
                    status, last_seen_at, status_bucket, risk_bucket,
                    is_isolated, alerts_count, last_activity,
                    CASE
                        WHEN is_isolated THEN 'ISOLATED'
                        WHEN status = 'ONLINE' THEN 'ONLINE'
                        ELSE 'OFFLINE'
                    END AS conn_status
                FROM endpoint_full
            )
            """

            # Resumen -- sin filtrar, para las 5 tarjetas de arriba.
            cursor.execute(
                cte + """
                SELECT
                    COUNT(*) AS total_n,
                    COUNT(*) FILTER (WHERE conn_status = 'ONLINE') AS online_n,
                    COUNT(*) FILTER (WHERE conn_status = 'OFFLINE') AS offline_n,
                    COUNT(*) FILTER (WHERE conn_status = 'ISOLATED') AS isolated_n,
                    COUNT(*) FILTER (WHERE risk_bucket = 'CRÍTICO') AS critical_n
                FROM endpoint_view;
                """
            )
            total_n, online_n, offline_n, isolated_n, critical_n = cursor.fetchone()

            cursor.execute("SELECT DISTINCT os FROM endpoints ORDER BY os;")
            os_families = sorted({row[0].split(" ")[0] for row in cursor.fetchall() if row[0]})

            where_clauses = []
            params = {}
            if search:
                where_clauses.append("(hostname ILIKE %(search)s OR host(ip_address) ILIKE %(search)s)")
                params["search"] = f"%{search}%"
            if status:
                where_clauses.append("conn_status = %(status)s")
                params["status"] = status
            if risk:
                where_clauses.append("risk_bucket = %(risk)s")
                params["risk"] = risk
            if os_family:
                where_clauses.append("operating_system ILIKE %(os_family)s")
                params["os_family"] = f"{os_family}%"
            where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

            cursor.execute(cte + f"SELECT COUNT(*) FROM endpoint_view {where_sql};", params)
            filtered_total = cursor.fetchone()[0]

            total_pages = max(1, -(-filtered_total // page_size))
            current_page = min(page, total_pages)
            offset = (current_page - 1) * page_size

            page_params = dict(params)
            page_params["limit"] = page_size
            page_params["offset"] = offset

            cursor.execute(
                cte + f"""
                SELECT id, hostname, operating_system, os_version, ip_address,
                       conn_status, risk_bucket, status_bucket, last_seen_at,
                       alerts_count, last_activity
                FROM endpoint_view
                {where_sql}
                ORDER BY id DESC
                LIMIT %(limit)s OFFSET %(offset)s;
                """,
                page_params
            )
            rows = cursor.fetchall()
    finally:
        connection.close()

    agent_health_map = {"ok": "HEALTHY", "attention": "WARNING", "offline": "OFFLINE"}

    endpoints = [
        {
            "id": r[0],
            "hostname": r[1],
            "operating_system": r[2],
            "os_version": r[3],
            "ip_address": str(r[4]) if r[4] else "127.0.0.1",
            "conn_status": r[5],
            "risk": r[6],
            "agent_health": agent_health_map[r[7]],
            "last_seen_ago": time_ago(r[8]),
            "alerts_count": r[9],
            "last_activity_ago": time_ago(r[10]) if r[10] else None,
        }
        for r in rows
    ]

    return {
        "summary": {
            "total": total_n,
            "online": online_n,
            "offline": offline_n,
            "isolated": isolated_n,
            "critical": critical_n,
        },
        "os_families": os_families,
        "page": current_page,
        "page_size": page_size,
        "total_pages": total_pages,
        "filtered_total": filtered_total,
        "endpoints": endpoints,
    }


@app.post("/incidents")
def create_incident(
    incident: IncidentCreate,
    user: dict = Depends(get_current_user)
):
    """Escala una detección a incidente. 'incidents' ya no tiene
    'alert_id' ni existe la tabla puente 'incident_alerts' -- ahora
    'alerts.incident_id' es una FK directa nullable (un incidente
    agrupa varias alertas; cada alerta pertenece como máximo a un
    incidente). Si la detección ya estaba en un incidente, se
    devuelve ese en vez de crear uno nuevo."""

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                "SELECT incident_id FROM alerts WHERE id = %s;",
                (incident.alert_id,)
            )

            alert_check = cursor.fetchone()

            if alert_check is None:
                raise HTTPException(status_code=404, detail="Alerta no encontrada")

            if alert_check[0] is not None:
                return {
                    "message": "Ya existía un incidente para esta alerta",
                    "incident_id": alert_check[0]
                }

            cursor.execute(
                """
                SELECT alerts.agent_id, alerts.title, alerts.description
                FROM alerts WHERE alerts.id = %s;
                """,
                (incident.alert_id,)
            )

            agent_id, title, description = cursor.fetchone()

            cursor.execute(
                """
                INSERT INTO incidents (
                    agent_id, title, description, classification
                )
                VALUES (%s, %s, %s, %s)
                RETURNING id;
                """,
                (agent_id, title, description, incident.classification)
            )

            incident_id = cursor.fetchone()[0]

            cursor.execute(
                "UPDATE alerts SET incident_id = %s WHERE id = %s;",
                (incident_id, incident.alert_id)
            )

            connection.commit()

        return {
            "message": "Incidente creado",
            "incident_id": incident_id
        }

    finally:
        connection.close()


@app.patch("/incidents/{incident_id}/status")
def update_incident_status(
    incident_id: int,
    payload: IncidentStatusUpdate,
    user: dict = Depends(get_current_user)
):
    """Cambia el estado del incidente (Abierto -> En investigación ->
    Contenido -> Cerrado). La nueva 'incidents' no tiene 'updated_at'
    ni 'closed_by' -- solo se registra 'closed_at'."""

    if payload.status not in INCIDENT_STATUS_LABELS_ES:
        raise HTTPException(status_code=422, detail="Estado inválido")

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            if payload.status == "CLOSED":
                cursor.execute(
                    """
                    UPDATE incidents
                    SET status = %s, closed_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    RETURNING id;
                    """,
                    (payload.status, incident_id)
                )
            else:
                cursor.execute(
                    """
                    UPDATE incidents
                    SET status = %s, closed_at = NULL
                    WHERE id = %s
                    RETURNING id;
                    """,
                    (payload.status, incident_id)
                )

            updated = cursor.fetchone()

            if updated is None:
                raise HTTPException(status_code=404, detail="Incidente no encontrado")

            log_audit(
                cursor, user["id"], "UPDATE_INCIDENT_STATUS", "incidents", incident_id,
                f"Estado -> {payload.status}"
            )

            connection.commit()

        return {
            "message": "Estado actualizado",
            "incident_id": incident_id,
            "status": payload.status,
            "status_label": INCIDENT_STATUS_LABELS_ES[payload.status]
        }

    finally:
        connection.close()


@app.patch("/incidents/{incident_id}/assign")
def assign_incident(
    incident_id: int,
    payload: IncidentAssign,
    user: dict = Depends(get_current_user)
):
    """Reintroducido 2026-08-12 -- 'assigned_to'/'assigned_at' se
    habían sacado al adoptar alfa_sentinel tal cual (ver PENDIENTES.md),
    pero se pidió la función de Responsable de verdad. payload.user_id
    en None desasigna (vuelve a 'Sin asignar')."""

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            if payload.user_id is not None:

                cursor.execute("SELECT id FROM users WHERE id = %s;", (payload.user_id,))

                if cursor.fetchone() is None:
                    raise HTTPException(status_code=404, detail="Usuario no encontrado")

                cursor.execute(
                    """
                    UPDATE incidents
                    SET assigned_to = %s, assigned_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    RETURNING id;
                    """,
                    (payload.user_id, incident_id)
                )

            else:

                cursor.execute(
                    """
                    UPDATE incidents
                    SET assigned_to = NULL, assigned_at = NULL
                    WHERE id = %s
                    RETURNING id;
                    """,
                    (incident_id,)
                )

            updated = cursor.fetchone()

            if updated is None:
                raise HTTPException(status_code=404, detail="Incidente no encontrado")

            log_audit(
                cursor, user["id"],
                "ASSIGN_INCIDENT" if payload.user_id else "UNASSIGN_INCIDENT",
                "incidents", incident_id,
                f"Responsable -> user_id={payload.user_id}" if payload.user_id else "Responsable removido"
            )

            connection.commit()

        return {
            "message": "Responsable actualizado" if payload.user_id else "Incidente desasignado",
            "incident_id": incident_id,
            "user_id": payload.user_id
        }

    finally:
        connection.close()


@app.get("/api/rules")
def api_rules(user: dict = Depends(get_current_user)):
    """Pantalla Reglas Heurísticas (React) -- versión completa
    (2026-08-16, ver PENDIENTES.md) que expone TODO lo que la
    especificación de la pantalla pide: identificación, métrica
    (nombre/descripción/unidad, desde 'metric_types'), evento
    (nombre/descripción, desde 'event_types'), parámetros, actividad
    real (alertas en 30 días / última activación, calculadas desde
    'alert_rule') y auditoría (creación/última actualización). Nada de
    esto se inventa: sale de columnas reales o de sub-consultas sobre
    datos reales; cuando no hay dato (ej. una regla diferida sin
    ninguna alerta todavía), el campo queda en None y el cliente
    decide cómo mostrarlo ("Sin actividad registrada", etc.), no se
    rellena acá con un valor de relleno."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute(
                """
                SELECT
                    heuristic_rules.id,
                    heuristic_rules.name,
                    heuristic_rules.description,
                    heuristic_rules.weight,
                    heuristic_rules.threshold,
                    heuristic_rules.window_seconds,
                    heuristic_rules.is_active,
                    heuristic_rules.created_at,
                    heuristic_rules.updated_at,
                    event_types.name,
                    event_types.description,
                    metric_types.name,
                    metric_types.description,
                    metric_types.unit,
                    (
                        SELECT COUNT(*)
                        FROM alert_rule
                        JOIN alerts ON alerts.id = alert_rule.alert_id
                        WHERE alert_rule.rule_id = heuristic_rules.id
                          AND alerts.created_at >= NOW() - INTERVAL '30 days'
                    ) AS alerts_30d,
                    (
                        SELECT MAX(alerts.created_at)
                        FROM alert_rule
                        JOIN alerts ON alerts.id = alert_rule.alert_id
                        WHERE alert_rule.rule_id = heuristic_rules.id
                    ) AS last_triggered_at
                FROM heuristic_rules
                LEFT JOIN event_types ON event_types.id = heuristic_rules.event_type_id
                LEFT JOIN metric_types ON metric_types.id = heuristic_rules.metric_type_id
                ORDER BY heuristic_rules.weight DESC, heuristic_rules.name ASC;
                """
            )
            rows = cursor.fetchall()
    finally:
        connection.close()

    rules = [
        {
            "id": r[0],
            "name": r[1],
            # 'label' es heuristic_rules.name tal cual -- sin diccionario
            # de traducción paralelo (decisión 2026-08-16, ver
            # PENDIENTES.md: los nombres en heuristic_rules ya son los
            # definitivos, no se vuelven a traducir en código).
            "label": r[1],
            "description": r[2],
            "weight": float(r[3]),
            "threshold": float(r[4]),
            "window_seconds": r[5],
            "is_active": r[6],
            "created_at": r[7].strftime("%d/%m/%Y %H:%M:%S") if r[7] else None,
            "updated_at": r[8].strftime("%d/%m/%Y %H:%M:%S") if r[8] else None,
            "event_type_name": r[9],
            "event_type_label": r[10] if r[9] else "Cualquiera en la ventana",
            "event_type_description": r[10],
            "metric_type_name": r[11],
            "metric_type_description": r[12],
            "metric_unit": r[13],
            "alerts_30d": r[14],
            "last_triggered_at": r[15].strftime("%d/%m/%Y %H:%M:%S") if r[15] else None,
            # Calculados acá (no una columna nueva) para que el
            # cliente no tenga que mantener su propia copia de estos
            # dos sets -- una sola fuente de verdad, la misma que ya
            # usa PATCH /rules/{id} para bloquear ediciones inválidas.
            "is_deferred": r[1] in DEFERRED_RULE_NAMES,
            "is_honeyfile": r[1] == "Acceso Honeyfile",
            "has_fixed_scoring": r[1] in FIXED_SCORING_RULE_NAMES,
        }
        for r in rows
    ]

    return {
        "summary": {
            "total": len(rules),
            "active": sum(1 for r in rules if r["is_active"]),
            "inactive": sum(1 for r in rules if not r["is_active"]),
            "alerts_30d_total": sum(r["alerts_30d"] for r in rules),
        },
        "rules": rules,
    }


@app.patch("/rules/{rule_id}")
def update_rule(
    rule_id: int,
    payload: RuleUpdate,
    user: dict = Depends(get_current_user)
):
    """Página /configuracion. 'weight'/'is_active'/'threshold'/
    'window_seconds' son ahora los cuatro campos de 'heuristic_rules'
    que el sistema realmente lee después de guardarlos:

    - 'weight' se copia a 'alert_rule.weight_applied' en cada alerta
      nueva que matchea la regla (POST /agent/alerts) y participa del
      cálculo del riesgo agregado en reportes/analítica.
    - 'is_active' decide si una alerta nueva queda vinculada a la
      regla (is_active = FALSE en la consulta de POST /agent/alerts) y
      además ahora decide si el agente la evalúa en absoluto (ver
      GET /agent/rule-policy).
    - 'threshold'/'window_seconds' (2026-08-12): antes eran solo
      referencia visual porque el agente los tenía hardcodeados en
      FileActivityAnalyzer.__init__() (agent/heuristic_engine.py) y
      arrancaba sin pedirle nada a este ni a ningún otro endpoint.
      Ahora el agente pide GET /agent/rule-policy en cada ejecución
      (mismo patrón que ya existía para honeyfiles) y usa estos
      valores de verdad -- el cambio se aplica recién la próxima vez
      que el agente arranque, no en caliente sobre un proceso ya
      corriendo (sigue sin tener bucle en segundo plano, ver
      PENDIENTES.md)."""

    if all(v is None for v in (payload.weight, payload.is_active, payload.threshold, payload.window_seconds)):
        raise HTTPException(status_code=422, detail="Nada para actualizar")

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            # Se leen los valores ANTERIORES antes de tocar nada --
            # hace falta tanto para las validaciones (nombre de la
            # regla) como para la auditoría completa (sección 15 de la
            # especificación: "valores anteriores; valores nuevos").
            cursor.execute(
                "SELECT name, weight, threshold, window_seconds, is_active FROM heuristic_rules WHERE id = %s;",
                (rule_id,)
            )
            old_row = cursor.fetchone()

            if old_row is None:
                raise HTTPException(status_code=404, detail="Regla no encontrada")

            old_name, old_weight, old_threshold, old_window, old_is_active = old_row

            if payload.is_active is True and old_name in DEFERRED_RULE_NAMES:
                raise HTTPException(
                    status_code=422,
                    detail=(
                        "Esta regla está diferida: requiere datos que el agente no recopila "
                        "hoy (atribución de proceso a evento de archivo, o muestreo de CPU por "
                        "proceso). Activarla acá no haría que el agente empiece a evaluarla -- "
                        "ver la descripción de la regla."
                    )
                )

            if old_name in FIXED_SCORING_RULE_NAMES and (
                payload.weight is not None or payload.threshold is not None or payload.window_seconds is not None
            ):
                if old_name == "Acceso Honeyfile":
                    reason = (
                        "El peso de esta regla es fijo (100): cualquier interacción con un "
                        "honeyfile debe llevar el riesgo a CRÍTICO sin depender de otras reglas. "
                        "Solo se puede activar o desactivar, no ajustar su peso/umbral/ventana."
                    )
                else:
                    reason = (
                        "Esta regla no puntúa por umbral/ventana: es una bonificación de "
                        "correlación calculada por el servidor según cuántas reglas distintas "
                        "coincidieron (2 -> +5, 3 -> +10, 4+ -> +15). Solo se puede activar o "
                        "desactivar, no ajustar su peso/umbral/ventana."
                    )
                raise HTTPException(status_code=422, detail=reason)

            fields = []
            values = []

            if payload.weight is not None:
                if payload.weight < 0 or payload.weight > 100:
                    raise HTTPException(status_code=422, detail="El peso tiene que estar entre 0 y 100")
                fields.append("weight = %s")
                values.append(payload.weight)

            if payload.is_active is not None:
                fields.append("is_active = %s")
                values.append(payload.is_active)

            if payload.threshold is not None:
                if payload.threshold <= 0:
                    raise HTTPException(status_code=422, detail="El umbral tiene que ser mayor a 0")
                fields.append("threshold = %s")
                values.append(payload.threshold)

            if payload.window_seconds is not None:
                if payload.window_seconds <= 0:
                    raise HTTPException(status_code=422, detail="La ventana tiene que ser mayor a 0 segundos")
                fields.append("window_seconds = %s")
                values.append(payload.window_seconds)

            fields.append("updated_at = CURRENT_TIMESTAMP")
            values.append(rule_id)

            cursor.execute(
                f"""
                UPDATE heuristic_rules
                SET {', '.join(fields)}
                WHERE id = %s
                RETURNING id, name, weight, is_active, threshold, window_seconds, updated_at;
                """,
                values
            )

            updated = cursor.fetchone()

            change_parts = []
            if payload.weight is not None:
                change_parts.append(f"peso: {old_weight} -> {payload.weight}")
            if payload.is_active is not None:
                change_parts.append(f"activa: {old_is_active} -> {payload.is_active}")
            if payload.threshold is not None:
                change_parts.append(f"umbral: {old_threshold} -> {payload.threshold}")
            if payload.window_seconds is not None:
                change_parts.append(f"ventana: {old_window}s -> {payload.window_seconds}s")

            log_audit(
                cursor, user["id"], "UPDATE_RULE", "heuristic_rules", updated[0],
                f"{updated[1]}: {', '.join(change_parts)}"
            )

            connection.commit()

        return {
            "message": "Regla actualizada",
            "rule_id": updated[0],
            "name": updated[1],
            "weight": float(updated[2]),
            "is_active": updated[3],
            "threshold": float(updated[4]),
            "window_seconds": updated[5],
            "updated_at": updated[6].strftime("%d/%m/%Y %H:%M:%S") if updated[6] else None,
        }

    finally:
        connection.close()


@app.get("/api/agents/{agent_id}/rules")
def api_agent_rules(agent_id: int, user: dict = Depends(get_current_user)):
    """Configuración de reglas heurísticas para UN endpoint puntual
    (2026-08-16, ver PENDIENTES.md). Pantalla: Endpoints -> detalle de
    un endpoint -> "Configuración de reglas". Para cada una de las 12
    reglas devuelve el valor GLOBAL ('heuristic_rules'), el override
    puntual si existe ('agent_rule' para este agent_id) y el valor
    EFECTIVO resultante (el mismo que usan GET /agent/rule-policy y
    POST /agent/alerts) -- para que la interfaz pueda mostrar
    claramente "Valor global" vs "Valor personalizado" en vez de
    obligar al analista a calcular la herencia a mano."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:

            cursor.execute("SELECT agents.id, endpoints.hostname FROM agents JOIN endpoints ON endpoints.id = agents.endpoint_id WHERE agents.id = %s;", (agent_id,))
            agent_row = cursor.fetchone()
            if agent_row is None:
                raise HTTPException(status_code=404, detail="Agente no encontrado")

            cursor.execute(
                _effective_agent_rules_cte() + """
                SELECT
                    effective_rules.id,
                    effective_rules.name,
                    effective_rules.description,
                    event_types.name,
                    event_types.description,
                    metric_types.name,
                    effective_rules.global_weight,
                    effective_rules.global_threshold,
                    effective_rules.global_window_seconds,
                    effective_rules.global_is_active,
                    effective_rules.override_id,
                    effective_rules.override_weight,
                    effective_rules.override_threshold,
                    effective_rules.override_window_seconds,
                    effective_rules.override_is_active,
                    effective_rules.effective_weight,
                    effective_rules.effective_threshold,
                    effective_rules.effective_window_seconds,
                    effective_rules.effective_is_active
                FROM effective_rules
                LEFT JOIN event_types ON event_types.id = effective_rules.event_type_id
                LEFT JOIN metric_types ON metric_types.id = effective_rules.metric_type_id
                ORDER BY effective_rules.global_weight DESC, effective_rules.name ASC;
                """,
                {"agent_id": agent_id}
            )
            rows = cursor.fetchall()
    finally:
        connection.close()

    rules = [
        {
            "id": r[0],
            "name": r[1],
            "description": r[2],
            "event_type_name": r[3],
            "event_type_label": r[4] if r[3] else "Cualquiera en la ventana",
            "metric_type_name": r[5],
            "global": {
                "weight": float(r[6]),
                "threshold": float(r[7]),
                "window_seconds": r[8],
                "is_active": r[9],
            },
            "override": (
                {
                    "id": r[10],
                    "weight": float(r[11]) if r[11] is not None else None,
                    "threshold": float(r[12]) if r[12] is not None else None,
                    "window_seconds": r[13],
                    "is_active": r[14],
                }
                if r[10] is not None else None
            ),
            "effective": {
                "weight": float(r[15]),
                "threshold": float(r[16]),
                "window_seconds": r[17],
                "is_active": r[18],
            },
            "has_override": r[10] is not None,
            "is_deferred": r[1] in DEFERRED_RULE_NAMES,
            "is_honeyfile": r[1] == "Acceso Honeyfile",
            "has_fixed_scoring": r[1] in FIXED_SCORING_RULE_NAMES,
        }
        for r in rows
    ]

    return {
        "agent_id": agent_row[0],
        "hostname": agent_row[1],
        "rules": rules,
    }


@app.patch("/api/agents/{agent_id}/rules/{rule_id}")
def update_agent_rule(
    agent_id: int,
    rule_id: int,
    payload: AgentRuleUpdate,
    user: dict = Depends(get_current_user)
):
    """Crea o actualiza el override de UNA regla para UN endpoint
    (tabla 'agent_rule', sección 24/25 de la especificación -- no se
    crea ninguna tabla nueva). Semántica de PATCH parcial con NULL
    significativo (ver AgentRuleUpdate): un campo AUSENTE del body no
    se toca; un campo presente con valor 'null' se guarda como NULL en
    'agent_rule' (= "volver a heredar el valor global para ese campo
    puntual"); un campo presente con un valor concreto lo reemplaza.
    'is_active' es la excepción -- la columna no admite NULL (ver
    schema), así que mandarlo en null es un error de validación, no
    una instrucción de herencia."""

    fields_set = payload.model_fields_set

    if not fields_set:
        raise HTTPException(status_code=422, detail="Nada para actualizar")

    if "is_active" in fields_set and payload.is_active is None:
        raise HTTPException(
            status_code=422,
            detail="'is_active' no puede ser null -- mandá true o false explícito, o simplemente no lo incluyas en el body."
        )

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute("SELECT id FROM agents WHERE id = %s;", (agent_id,))
            if cursor.fetchone() is None:
                raise HTTPException(status_code=404, detail="Agente no encontrado")

            cursor.execute(
                "SELECT name, weight, threshold, window_seconds, is_active FROM heuristic_rules WHERE id = %s;",
                (rule_id,)
            )
            rule_row = cursor.fetchone()
            if rule_row is None:
                raise HTTPException(status_code=404, detail="Regla no encontrada")

            rule_name = rule_row[0]

            # Mismas dos restricciones que PATCH /rules/{id} (sección 19
            # de la especificación): HR-03/HR-12 no admiten
            # personalizar weight/threshold/window ni a nivel global ni
            # por endpoint -- solo pueden encenderse/apagarse. Y una
            # regla diferida (capacidad que el agente todavía no tiene)
            # no se puede activar desde ningún lado, para no sugerir
            # una funcionalidad que no existe.
            if rule_name in FIXED_SCORING_RULE_NAMES and fields_set & {"weight", "threshold", "window_seconds"}:
                if rule_name == "Acceso Honeyfile":
                    reason = (
                        "El peso de esta regla es fijo (100): cualquier interacción con un "
                        "honeyfile debe llevar el riesgo a CRÍTICO sin depender de otras reglas. "
                        "Para este endpoint solo se puede activar o desactivar, no ajustar su "
                        "peso/umbral/ventana."
                    )
                else:
                    reason = (
                        "Esta regla no puntúa por umbral/ventana: es una bonificación de "
                        "correlación calculada por el servidor. Para este endpoint solo se puede "
                        "activar o desactivar, no ajustar su peso/umbral/ventana."
                    )
                raise HTTPException(status_code=422, detail=reason)

            if "is_active" in fields_set and payload.is_active is True and rule_name in DEFERRED_RULE_NAMES:
                raise HTTPException(
                    status_code=422,
                    detail=(
                        "Esta regla está diferida: requiere datos que el agente no recopila hoy. "
                        "Activarla para este endpoint no haría que el agente empiece a evaluarla."
                    )
                )

            if "weight" in fields_set and payload.weight is not None and payload.weight < 0:
                raise HTTPException(status_code=422, detail="El peso tiene que ser mayor o igual a 0")

            if "threshold" in fields_set and payload.threshold is not None and payload.threshold < 0:
                raise HTTPException(status_code=422, detail="El umbral tiene que ser mayor o igual a 0")

            if "window_seconds" in fields_set and payload.window_seconds is not None and payload.window_seconds <= 0:
                raise HTTPException(status_code=422, detail="La ventana tiene que ser mayor a 0 segundos")

            cursor.execute(
                "SELECT id, threshold, window_seconds, weight, is_active FROM agent_rule WHERE agent_id = %s AND rule_id = %s;",
                (agent_id, rule_id)
            )
            existing = cursor.fetchone()

            if existing is None:
                final_threshold = payload.threshold if "threshold" in fields_set else None
                final_window = payload.window_seconds if "window_seconds" in fields_set else None
                final_weight = payload.weight if "weight" in fields_set else None
                final_is_active = payload.is_active if "is_active" in fields_set else True

                cursor.execute(
                    """
                    INSERT INTO agent_rule (agent_id, rule_id, threshold, window_seconds, weight, is_active)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    RETURNING id, threshold, window_seconds, weight, is_active;
                    """,
                    (agent_id, rule_id, final_threshold, final_window, final_weight, final_is_active)
                )
                action = "CREATE_AGENT_RULE_OVERRIDE"
            else:
                override_id, old_threshold, old_window, old_weight, old_is_active = existing
                final_threshold = payload.threshold if "threshold" in fields_set else old_threshold
                final_window = payload.window_seconds if "window_seconds" in fields_set else old_window
                final_weight = payload.weight if "weight" in fields_set else old_weight
                final_is_active = payload.is_active if "is_active" in fields_set else old_is_active

                cursor.execute(
                    """
                    UPDATE agent_rule
                    SET threshold = %s, window_seconds = %s, weight = %s, is_active = %s
                    WHERE id = %s
                    RETURNING id, threshold, window_seconds, weight, is_active;
                    """,
                    (final_threshold, final_window, final_weight, final_is_active, override_id)
                )
                action = "UPDATE_AGENT_RULE_OVERRIDE"

            updated = cursor.fetchone()

            log_audit(
                cursor, user["id"], action, "agent_rule", updated[0],
                f"{rule_name} (agent_id={agent_id}): threshold={updated[1]}, window_seconds={updated[2]}, "
                f"weight={updated[3]}, is_active={updated[4]}"
            )

            connection.commit()

        return {
            "message": "Configuración del endpoint actualizada",
            "override": {
                "id": updated[0],
                "threshold": float(updated[1]) if updated[1] is not None else None,
                "window_seconds": updated[2],
                "weight": float(updated[3]) if updated[3] is not None else None,
                "is_active": updated[4],
            },
        }

    finally:
        connection.close()


@app.delete("/api/agents/{agent_id}/rules/{rule_id}")
def delete_agent_rule(agent_id: int, rule_id: int, user: dict = Depends(get_current_user)):
    """Quita el override por completo -- el endpoint vuelve a heredar
    el valor global de la regla en los cuatro campos (no solo se
    resetean uno por uno, se elimina la fila entera de 'agent_rule')."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:

            cursor.execute(
                "SELECT heuristic_rules.name FROM heuristic_rules WHERE heuristic_rules.id = %s;",
                (rule_id,)
            )
            rule_row = cursor.fetchone()
            rule_name = rule_row[0] if rule_row else f"id={rule_id}"

            cursor.execute(
                "DELETE FROM agent_rule WHERE agent_id = %s AND rule_id = %s RETURNING id;",
                (agent_id, rule_id)
            )
            deleted = cursor.fetchone()

            if deleted is None:
                raise HTTPException(status_code=404, detail="Este endpoint no tiene una configuración personalizada para esa regla")

            log_audit(
                cursor, user["id"], "DELETE_AGENT_RULE_OVERRIDE", "agent_rule", deleted[0],
                f"{rule_name} (agent_id={agent_id}): override eliminado, vuelve a heredar el valor global"
            )

            connection.commit()

        return {"message": "Configuración personalizada eliminada -- este endpoint vuelve a usar el valor global"}
    finally:
        connection.close()


@app.patch("/settings/{key}")
def update_setting(
    key: str,
    payload: SettingUpdate,
    user: dict = Depends(get_current_user)
):
    """Página /configuracion > Agentes (2026-08-12). A propósito solo
    acepta claves que el servidor de verdad vuelve a leer -- hoy la
    única es 'agent_stale_seconds' (ver get_agent_stale_seconds()).
    No se expone un endpoint genérico "guardame cualquier key" sin
    lista blanca: eso permitiría crear settings que nadie consume,
    el mismo problema que ya se evitó con threshold/window_seconds
    en las reglas heurísticas."""

    KNOWN_SETTINGS = {
        "agent_stale_seconds": {
            "cast": int,
            "validate": lambda v: v > 0,
            "error": "Tiene que ser un número entero mayor a 0.",
        }
    }

    if key not in KNOWN_SETTINGS:
        raise HTTPException(
            status_code=404,
            detail=f"'{key}' no es un parámetro configurable -- no hay ningún mecanismo en el servidor que lo consuma."
        )

    spec = KNOWN_SETTINGS[key]

    try:
        cast_value = spec["cast"](payload.value)
    except (TypeError, ValueError):
        raise HTTPException(status_code=422, detail=spec["error"])

    if not spec["validate"](cast_value):
        raise HTTPException(status_code=422, detail=spec["error"])

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                """
                UPDATE system_settings
                SET value = %s, updated_at = CURRENT_TIMESTAMP, updated_by = %s
                WHERE key = %s
                RETURNING key, value;
                """,
                (str(cast_value), user["id"], key)
            )

            updated = cursor.fetchone()

            if updated is None:
                raise HTTPException(status_code=404, detail="Parámetro no encontrado")

            log_audit(
                cursor, user["id"], "UPDATE_SETTING", "system_settings", None,
                f"{key} -> {cast_value}"
            )

            connection.commit()

        return {"message": "Parámetro actualizado", "key": updated[0], "value": updated[1]}

    finally:
        connection.close()


@app.patch("/incidents/{incident_id}/classification")
def update_incident_classification(
    incident_id: int,
    payload: IncidentClassify,
    user: dict = Depends(get_current_user)
):
    """Clasificación del resultado -- deliberadamente separada del
    estado (ver comentario de INCIDENT_CLASSIFICATION_LABELS_ES)."""

    if payload.classification not in INCIDENT_CLASSIFICATION_LABELS_ES:
        raise HTTPException(status_code=422, detail="Clasificación inválida")

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                """
                UPDATE incidents
                SET classification = %s
                WHERE id = %s
                RETURNING id;
                """,
                (payload.classification, incident_id)
            )

            updated = cursor.fetchone()

            if updated is None:
                raise HTTPException(status_code=404, detail="Incidente no encontrado")

            connection.commit()

        return {
            "message": "Clasificación actualizada",
            "incident_id": incident_id,
            "classification": payload.classification,
            "classification_label": INCIDENT_CLASSIFICATION_LABELS_ES[payload.classification]
        }

    finally:
        connection.close()


# Notas de analista sobre detecciones/incidentes (alert_notes,
# incident_notes) y "Responsable" de incidente (assigned_to/
# assigned_at) se sacaron: esas tablas/columnas no existen en la
# nueva estructura (alfa_sentinel) -- decisión explícita de adoptarla
# tal cual. Ver PENDIENTES.md.

# 'severity' y 'detection_count' no son columnas de 'incidents' --
# se derivan de las alertas vinculadas vía 'alerts.incident_id'
# (antes: tabla puente 'incident_alerts', ahora FK directa).
# 'severity' es la más alta entre esas detecciones (mismo criterio de
# "peor caso" que ya usa ENDPOINT_CTE para el riesgo de un endpoint).
# Se arma como CTE para poder filtrar/contar por estos valores
# derivados en cada consulta que lo necesite.
INCIDENT_CTE = """
    WITH incident_data AS (
        SELECT incidents.id, incidents.title, incidents.description,
               incidents.status, incidents.classification,
               incidents.opened_at, incidents.closed_at,
               incidents.agent_id,
               endpoints.hostname,
               (
                   SELECT COUNT(*) FROM alerts
                   WHERE alerts.incident_id = incidents.id
               ) AS detection_count,
               (
                   SELECT severity_levels.name FROM alerts
                   JOIN severity_levels ON severity_levels.id = alerts.severity_id
                   WHERE alerts.incident_id = incidents.id
                   ORDER BY severity_levels.min_score DESC
                   LIMIT 1
               ) AS severity,
               (
                   SELECT COALESCE(MAX(alerts.risk_score), 0) FROM alerts
                   WHERE alerts.incident_id = incidents.id
               ) AS risk_score,
               incidents.assigned_to,
               assigned_user.full_name AS assigned_to_name
        FROM incidents
        JOIN agents ON agents.id = incidents.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        LEFT JOIN users AS assigned_user ON assigned_user.id = incidents.assigned_to
    )
"""

INCIDENTES_PAGE_SIZE = 25

# 'Incidentes y Alertas' unifica dos cosas de naturaleza distinta en
# una sola matriz, tal como lo pidió el mockup: incidentes ya agrupados
# (varias alertas relacionadas bajo un mismo caso) y alertas sueltas
# que todavía no se escalaron a incidente (alerts.incident_id IS NULL).
# Antes eran dos pantallas separadas (Detecciones / Incidentes) -- esa
# separación se mantiene a nivel de rutas (/detecciones/{id} sigue
# existiendo para el detalle de una alerta puntual) pero la LISTA
# ahora es una sola. 'status_bucket' traduce los dos vocabularios de
# estado reales (incidents.status e alerts.status, que no coinciden)
# a un set compartido solo para poder filtrar/colorear parejo -- el
# estado real de cada fila (status_label) se sigue mostrando tal cual
# está guardado, no se renombra.
COMBINED_CTE = f"""
    WITH combined AS (
        SELECT
            'incident' AS kind, incidents.id AS id, incidents.opened_at AS ts,
            incidents.status AS raw_status,
            CASE incidents.status
                WHEN 'OPEN' THEN 'nuevo' WHEN 'IN_PROGRESS' THEN 'investigando'
                WHEN 'CONTAINED' THEN 'contenido' WHEN 'CLOSED' THEN 'cerrado'
                ELSE 'nuevo'
            END AS status_bucket,
            endpoints.hostname, endpoints.ip_address, agents.id AS agent_id,
            (
                SELECT severity_levels.name FROM alerts
                JOIN severity_levels ON severity_levels.id = alerts.severity_id
                WHERE alerts.incident_id = incidents.id
                ORDER BY severity_levels.min_score DESC LIMIT 1
            ) AS severity,
            (
                SELECT COALESCE(MAX(alerts.risk_score), 0) FROM alerts
                WHERE alerts.incident_id = incidents.id
            ) AS risk_score,
            (
                SELECT STRING_AGG(DISTINCT heuristic_rules.name, ' + ') FROM alerts
                LEFT JOIN alert_rule ON alert_rule.alert_id = alerts.id
                LEFT JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                WHERE alerts.incident_id = incidents.id
            ) AS rule_names,
            (SELECT COUNT(*) FROM alerts WHERE alerts.incident_id = incidents.id) AS detection_count,
            incidents.assigned_to, assigned_user.full_name AS assigned_to_name,
            -- Corregido 2026-08-18 (problema H, ver PENDIENTES.md): antes
            -- filtraba por 'host_isolations.incident_id = incidents.id'
            -- -- si el aislamiento se ordenó desde OTRO incidente del
            -- mismo endpoint, esta fila no lo encontraba y seguía
            -- ofreciendo "Aislar" sobre un endpoint ya aislado. El
            -- aislamiento es del AGENTE, no del incidente -- ver
            -- _agent_isolation_status_sql().
            {_agent_isolation_status_sql("agents.id")} AS isolation_status
        FROM incidents
        JOIN agents ON agents.id = incidents.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        LEFT JOIN users AS assigned_user ON assigned_user.id = incidents.assigned_to

        UNION ALL

        SELECT
            'alert' AS kind, alerts.id AS id, alerts.created_at AS ts,
            alerts.status AS raw_status,
            CASE alerts.status
                WHEN 'NEW' THEN 'nuevo' WHEN 'ACKNOWLEDGED' THEN 'investigando'
                WHEN 'ESCALATED' THEN 'confirmado' WHEN 'CLOSED' THEN 'cerrado'
                WHEN 'FALSE_POSITIVE' THEN 'falso_positivo'
                ELSE 'nuevo'
            END AS status_bucket,
            endpoints.hostname, endpoints.ip_address, agents.id AS agent_id,
            severity_levels.name AS severity,
            alerts.risk_score AS risk_score,
            heuristic_rules.name AS rule_names,
            0 AS detection_count,
            NULL::BIGINT AS assigned_to, NULL::TEXT AS assigned_to_name,
            -- Corregido 2026-08-18 (problema H, ver PENDIENTES.md): esta
            -- rama es SIEMPRE 'alerts.incident_id IS NULL' (ver WHERE más
            -- abajo) -- la alerta en sí nunca tiene una orden de
            -- aislamiento propia (el aislamiento se asocia a un
            -- incidente), PERO el ENDPOINT de esa alerta puede estar
            -- aislado igual por un incidente distinto y anterior sobre el
            -- mismo agente -- antes esto era NULL fijo (siempre "no
            -- aislado"), lo que dejaba ofrecer "Aislar" desde una alerta
            -- suelta de un endpoint que ya estaba aislado. Ver
            -- _agent_isolation_status_sql().
            {_agent_isolation_status_sql("agents.id")} AS isolation_status
        FROM alerts
        JOIN agents ON agents.id = alerts.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        JOIN severity_levels ON severity_levels.id = alerts.severity_id
        LEFT JOIN alert_rule ON alert_rule.alert_id = alerts.id
        LEFT JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
        WHERE alerts.incident_id IS NULL
    )
"""

STATUS_BUCKET_LABELS_ES = {
    "nuevo": "Nuevo",
    "investigando": "En investigación",
    "confirmado": "Confirmado",
    "contenido": "Contenido",
    "cerrado": "Cerrado",
    "falso_positivo": "Falso positivo",
}


@app.get("/api/incidentes")
def api_incidentes(
    agent_id: int | None = Query(None),
    status_bucket: str = Query("", alias="status"),
    severity: str = Query(""),
    rule: str = Query(""),
    since: str = Query(""),
    search: str = Query(""),
    # Vista operativa vs. historial -- ver el mismo parámetro en
    # GET /api/alerts (problema G, 2026-08-18, ver PENDIENTES.md). Acá
    # el único bucket "final" real de un incidente es 'cerrado'
    # (incidents.status = 'CLOSED', ver INCIDENT_STATUS_BUCKETS abajo).
    view: str = Query("activas", pattern="^(activas|todos)$"),
    page: int = Query(1, ge=1),
    user: dict = Depends(get_current_user)
):
    """API para la pantalla Incidentes en React -- misma consulta
    (COMBINED_CTE), mismos filtros, mismos KPIs que ya usa la matriz
    combinada de incidentes/alertas (ver COMBINED_CTE más arriba).

    A propósito solo devuelve incidentes agrupados, no alertas sueltas
    -- pedido explícito (2026-08-15) para que la lista sea más rápida
    de leer. Las alertas sueltas ya tienen su propia pantalla dedicada
    (Alertas)."""

    # Solo los 4 buckets que de verdad puede tener un incidente
    # (OPEN/IN_PROGRESS/CONTAINED/CLOSED) -- 'confirmado' y
    # 'falso_positivo' son exclusivos de alerts.status y ahora que esta
    # lista no muestra alertas sueltas, ofrecerlos como filtro siempre
    # devolvería vacío.
    INCIDENT_STATUS_BUCKETS = {"nuevo", "investigando", "contenido", "cerrado"}

    status_bucket = status_bucket if status_bucket in INCIDENT_STATUS_BUCKETS else ""
    since = since if since in INCIDENTES_SINCE_OPTIONS else ""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:

            # Nombres reales de heuristic_rules -- no se mantiene un
            # diccionario paralelo (mismo criterio que en el resto del
            # archivo, ver auditoría de catálogos duplicados).
            cursor.execute("SELECT name FROM heuristic_rules ORDER BY name;")
            rule_names_catalog = [row[0] for row in cursor.fetchall()]
            rule = rule if rule in rule_names_catalog else ""

            # Igual que arriba, pero con severity_levels -- ya no hay
            # ALERT_SEVERITY_LABELS_ES paralelo, se valida contra la
            # tabla real, ordenada por severidad real (min_score).
            cursor.execute("SELECT name FROM severity_levels ORDER BY min_score;")
            severity_catalog = [row[0] for row in cursor.fetchall()]
            severity = severity if severity in severity_catalog else ""

            where_clauses = ["kind = 'incident'"]
            params = {}

            # Igual criterio que /api/alerts: si el analista eligió un
            # bucket de estado puntual (incluido 'cerrado'), ese filtro
            # manda; si no eligió ninguno y la vista es la de por
            # defecto ('activas'), se excluye 'cerrado' sin necesidad de
            # inventar un estado nuevo.
            if not status_bucket and view == "activas":
                where_clauses.append("status_bucket != 'cerrado'")

            if agent_id:
                where_clauses.append("agent_id = %(agent_id)s")
                params["agent_id"] = agent_id
            if status_bucket:
                where_clauses.append("status_bucket = %(status_bucket)s")
                params["status_bucket"] = status_bucket
            if severity:
                where_clauses.append("severity = %(severity)s")
                params["severity"] = severity
            if rule:
                where_clauses.append("rule_names ILIKE %(rule)s")
                params["rule"] = f"%{rule}%"
            if since:
                where_clauses.append("ts >= CURRENT_TIMESTAMP - INTERVAL %(since_interval)s")
                params["since_interval"] = INCIDENTES_SINCE_OPTIONS[since][1]
            if search:
                where_clauses.append(
                    "(hostname ILIKE %(search)s OR CAST(ip_address AS TEXT) ILIKE %(search)s "
                    "OR rule_names ILIKE %(search)s OR CAST(id AS TEXT) ILIKE %(search)s)"
                )
                params["search"] = f"%{search}%"

            where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

            cursor.execute(
                COMBINED_CTE + "SELECT COUNT(*) FROM combined WHERE kind = 'incident' AND severity = 'CRÍTICO' AND raw_status != 'CLOSED';"
            )
            critical_incidents = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM alerts WHERE status IN ('NEW', 'ACKNOWLEDGED');")
            active_alerts = cursor.fetchone()[0]

            # Real desde la corrección definitiva del motor heurístico
            # (2026-08-17, ver PENDIENTES.md) -- mismo filtro de status
            # que el resto de las consultas de aislamiento en este
            # archivo (ISOLATION_FAILED no cuenta como aislado).
            cursor.execute("SELECT COUNT(*) FROM host_isolations WHERE status IN ('REQUESTED', 'EXECUTED', 'RELEASE_REQUESTED') AND released_at IS NULL;")
            isolated_hosts = cursor.fetchone()[0]

            cursor.execute(
                "SELECT AVG(EXTRACT(EPOCH FROM (resolved_at - created_at))) FROM alerts WHERE resolved_at IS NOT NULL;"
            )
            mttr_row = cursor.fetchone()
            mttr_seconds = mttr_row[0] if mttr_row else None
            mttr_minutes = round(mttr_seconds / 60, 1) if mttr_seconds is not None else None

            cursor.execute("SELECT id, full_name FROM users ORDER BY full_name;")
            assignable_users = [{"id": r[0], "full_name": r[1]} for r in cursor.fetchall()]

            count_params = dict(params)
            cursor.execute(COMBINED_CTE + f"SELECT COUNT(*) FROM combined {where_sql};", count_params)
            filtered_total = cursor.fetchone()[0]

            total_pages = max(1, -(-filtered_total // INCIDENTES_PAGE_SIZE))
            current_page = min(page, total_pages)
            offset = (current_page - 1) * INCIDENTES_PAGE_SIZE

            page_params = dict(params)
            page_params["limit"] = INCIDENTES_PAGE_SIZE
            page_params["offset"] = offset

            cursor.execute(
                COMBINED_CTE + f"""
                SELECT kind, id, ts, raw_status, status_bucket, hostname, ip_address,
                       agent_id, severity, risk_score, rule_names, detection_count,
                       assigned_to, assigned_to_name, isolation_status
                FROM combined
                {where_sql}
                ORDER BY ts DESC
                LIMIT %(limit)s OFFSET %(offset)s;
                """,
                page_params
            )
            rows = cursor.fetchall()
    finally:
        connection.close()

    items = []
    for row in rows:
        (kind, item_id, ts, raw_status, bucket, hostname, ip_address, item_agent_id,
         severity_val, risk_score, rule_names, detection_count, assigned_to, assigned_to_name,
         isolation_status) = row

        items.append({
            "kind": kind,
            "id": item_id,
            "code": f"INC-{item_id:05d}" if kind == "incident" else f"ALT-{item_id:05d}",
            "created_at": ts.strftime("%d/%m/%Y %H:%M:%S"),
            "raw_status": raw_status,
            "status_bucket": bucket,
            "status_label": (INCIDENT_STATUS_LABELS_ES.get(raw_status, raw_status) if kind == "incident"
                              else ALERT_STATUS_LABELS_ES.get(raw_status, raw_status)),
            "hostname": hostname,
            "ip_address": str(ip_address) if ip_address else "127.0.0.1",
            "agent_id": item_agent_id,
            "severity": severity_val,
            "risk_score": float(risk_score) if risk_score is not None else None,
            "rule_label": " + ".join(
                n for n in (rule_names or "").split(" + ") if n
            ) or "—",
            "detection_count": detection_count,
            "assigned_to": assigned_to,
            "assigned_to_name": assigned_to_name,
            # Agregado 2026-08-17 (ver PENDIENTES.md, "Corrección de
            # tiempo real, ordenamiento y consistencia") -- lo que
            # necesita el botón "Aislar" de esta misma tabla (antes
            # deshabilitado permanentemente, sección 12 de esa
            # especificación) para saber si ya hay una orden en
            # curso/cumplida y no ofrecer aislar de nuevo.
            "isolation_status": isolation_status,
        })

    return {
        "summary": {
            "critical_incidents": critical_incidents,
            "active_alerts": active_alerts,
            "isolated_hosts": isolated_hosts,
            "mttr_minutes": mttr_minutes,
        },
        "filters": {
            "status_options": [
                {"value": k, "label": v} for k, v in STATUS_BUCKET_LABELS_ES.items()
                if k in INCIDENT_STATUS_BUCKETS
            ],
            "severity_options": [{"value": name, "label": name} for name in severity_catalog],
            "rule_options": [{"value": n, "label": n} for n in rule_names_catalog],
            "since_options": [{"value": k, "label": v[0]} for k, v in INCIDENTES_SINCE_OPTIONS.items()],
            "assignable_users": assignable_users,
        },
        "page": current_page,
        "page_size": INCIDENTES_PAGE_SIZE,
        "total_pages": total_pages,
        "filtered_total": filtered_total,
        "items": items,
    }


@app.get("/api/incidentes/{kind}/{item_id}/drawer")
def get_incidente_drawer(kind: str, item_id: int, request: Request):
    """Expediente para el panel lateral -- sirve tanto un incidente
    agrupado como una alerta suelta ('kind' viene de la fila que
    armó COMBINED_CTE). La 'cadena de evidencia' es real: eventos y
    activaciones de honeyfile del mismo agente en una ventana
    alrededor del momento en que se disparó (5 min antes, 1 min
    después) -- no se inventa un paso de 'ejecución de proceso' que
    el agente no reportó."""

    user = require_session_user(request)
    if user is None:
        return JSONResponse({"error": "No autorizado"}, status_code=401)

    if kind not in ("incident", "alert"):
        return JSONResponse({"error": "Tipo inválido"}, status_code=400)

    connection = get_connection()
    try:
        with connection.cursor() as cursor:

            if kind == "incident":

                cursor.execute(INCIDENT_CTE + "SELECT * FROM incident_data WHERE id = %s;", (item_id,))
                row = cursor.fetchone()
                if row is None:
                    return JSONResponse({"error": "Incidente no encontrado"}, status_code=404)

                (inc_id, title_val, description_val, status, classification, opened_at, closed_at,
                 agent_id, hostname, detection_count, severity, risk_score,
                 assigned_to, assigned_to_name) = row

                cursor.execute(
                    """
                    SELECT alerts.id, alerts.created_at, severity_levels.name, alerts.risk_score
                    FROM alerts
                    JOIN severity_levels ON severity_levels.id = alerts.severity_id
                    WHERE alerts.incident_id = %s
                    ORDER BY alerts.created_at ASC;
                    """,
                    (item_id,)
                )
                linked = cursor.fetchall()
                anchor_ts = linked[0][1] if linked else opened_at
                code = f"INC-{inc_id:05d}"
                status_label = INCIDENT_STATUS_LABELS_ES.get(status, status)

                # No aplican a un incidente agrupado -- son campos propios
                # de una alerta suelta (kind == 'alert').
                incident_id_val = None
                resolved_at_val = None

                # "Alerta de origen": la primera alerta (por fecha) que
                # quedó vinculada a este incidente -- ya sea porque
                # create_incident() la escaló manualmente, o porque el
                # motor automático la generó. Si más adelante se
                # vincularon más alertas al mismo incidente
                # (POST /incidents/{id}/alerts), esta sigue siendo la
                # que dio origen al caso, no la última.
                origin_alert = None
                if linked:
                    origin_row = linked[0]
                    origin_alert = {
                        "id": origin_row[0],
                        "code": f"ALT-{origin_row[0]:05d}",
                        "severity": origin_row[2],
                        "risk_score": float(origin_row[3]) if origin_row[3] is not None else None,
                    }

                # Reglas asociadas -- corregido 2026-08-18 (ver
                # PENDIENTES.md, "Corrección definitiva en la lógica y
                # presentación de ALERTAS"): antes esto quedaba
                # hardcodeado en una lista vacía ("no aplica a un
                # incidente agrupado"), así que el drawer de un
                # INCIDENTE nunca mostraba qué reglas lo componían --
                # solo el de una alerta suelta las mostraba. Un
                # incidente agrupa una o más alertas (alerts.incident_id
                # = item_id); sus reglas son la unión de las reglas de
                # TODAS esas alertas. No se fusionan/suman ocurrencias
                # de una misma regla entre distintas alertas del mismo
                # incidente -- cada coincidencia real se lista tal cual
                # (sección 3: "no ocultar reglas secundarias").
                cursor.execute(
                    """
                    SELECT heuristic_rules.name, alert_rule.weight_applied, alert_rule.matched_at
                    FROM alert_rule
                    JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                    JOIN alerts ON alerts.id = alert_rule.alert_id
                    WHERE alerts.incident_id = %s;
                    """,
                    (item_id,)
                )
                contributing_rows = cursor.fetchall()
                is_honeyfile = any(r[0] == "Acceso Honeyfile" for r in contributing_rows)
                matched_rules = [
                    {
                        "rule_name": r[0],
                        "weight_applied": float(r[1]),
                        "matched_at": r[2].strftime("%d/%m/%Y %H:%M:%S"),
                    }
                    for r in sort_contributing_rules(contributing_rows)
                ]

            else:

                cursor.execute(
                    """
                    SELECT alerts.id, alerts.title, alerts.description, alerts.status, alerts.created_at,
                           alerts.risk_score, severity_levels.name,
                           alerts.agent_id, alerts.incident_id, alerts.resolved_at
                    FROM alerts
                    JOIN severity_levels ON severity_levels.id = alerts.severity_id
                    WHERE alerts.id = %s;
                    """,
                    (item_id,)
                )
                row = cursor.fetchone()
                if row is None:
                    return JSONResponse({"error": "Alerta no encontrada"}, status_code=404)

                (alert_id, title_val, description_val, status, anchor_ts, risk_score, severity,
                 agent_id, incident_id_val, resolved_at_val) = row

                code = f"ALT-{alert_id:05d}"
                status_label = ALERT_STATUS_LABELS_ES.get(status, status)
                classification = None
                assigned_to = None
                assigned_to_name = None
                detection_count = 1
                # No aplica a una alerta suelta -- 'origin_alert' solo
                # tiene sentido para un incidente agrupado, ver arriba.
                origin_alert = None

                # Reglas asociadas (alert_rule) -- puede haber más de una
                # regla contribuyendo a la misma alerta, por eso es una
                # consulta aparte. Orden de relevancia real (sección 3
                # de "Corrección definitiva en la lógica y presentación
                # de ALERTAS", 2026-08-18, ver PENDIENTES.md), no
                # 'ORDER BY matched_at' simple -- ver sort_contributing_rules().
                cursor.execute(
                    """
                    SELECT heuristic_rules.name, alert_rule.weight_applied, alert_rule.matched_at
                    FROM alert_rule
                    JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                    WHERE alert_rule.alert_id = %s;
                    """,
                    (alert_id,)
                )
                contributing_rows = cursor.fetchall()
                is_honeyfile = any(r[0] == "Acceso Honeyfile" for r in contributing_rows)
                matched_rules = [
                    {
                        "rule_name": r[0],
                        "weight_applied": float(r[1]),
                        "matched_at": r[2].strftime("%d/%m/%Y %H:%M:%S"),
                    }
                    for r in sort_contributing_rules(contributing_rows)
                ]

            # Título general por severidad -- ver alert_general_title()
            # (2026-08-18, ver PENDIENTES.md, "Corrección definitiva en
            # la lógica y presentación de ALERTAS"). Aplica igual para
            # un incidente agrupado que para una alerta suelta: ambos
            # representan un EPISODIO (de una o varias señales), nunca
            # el nombre de una sola regla. 'title_val' (columna guardada
            # en 'incidents.title'/'alerts.title') sigue existiendo tal
            # cual en la base, solo se deja de usar para lo que se
            # muestra acá.
            title_val = alert_general_title(severity)

            cursor.execute(
                """
                SELECT endpoints.hostname, endpoints.ip_address, endpoints.os,
                       agents.status, agents.last_seen_at
                FROM agents
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                WHERE agents.id = %s;
                """,
                (agent_id,)
            )
            hostname, ip_address, operating_system, agent_status, last_seen_at = cursor.fetchone()

            stale_seconds = get_agent_stale_seconds(cursor)

            is_online = (
                agent_status == "ONLINE" and last_seen_at is not None
                and (datetime.now(last_seen_at.tzinfo) - last_seen_at).total_seconds() < stale_seconds
            ) if last_seen_at else False

            window_start = anchor_ts - timedelta(minutes=5)
            window_end = anchor_ts + timedelta(minutes=1)

            cursor.execute(
                """
                SELECT events.detected_at, event_types.name, events.file_path
                FROM events
                JOIN event_types ON event_types.id = events.event_type_id
                WHERE events.agent_id = %s
                  AND events.detected_at BETWEEN %s AND %s
                ORDER BY events.detected_at DESC LIMIT 25;
                """,
                (agent_id, window_start, window_end)
            )
            timeline = [
                {
                    "at_raw": r[0],
                    "at": r[0].strftime("%d/%m %H:%M:%S"),
                    "kind": "event",
                    "label": r[1],
                    "detail": r[2] or ""
                }
                for r in cursor.fetchall()
            ]

            cursor.execute(
                """
                SELECT honeyfile_activations.detected_at, honeyfile_activations.operation,
                       honeyfiles.file_name, honeyfile_activations.process_name,
                       honeyfile_activations.process_id
                FROM honeyfile_activations
                JOIN honeyfiles ON honeyfiles.id = honeyfile_activations.honeyfile_id
                WHERE honeyfile_activations.agent_id = %s
                  AND honeyfile_activations.detected_at BETWEEN %s AND %s
                ORDER BY honeyfile_activations.detected_at DESC LIMIT 10;
                """,
                (agent_id, window_start, window_end)
            )
            timeline += [
                {
                    "at_raw": r[0],
                    "at": r[0].strftime("%d/%m %H:%M:%S"),
                    "kind": "honeyfile",
                    "label": f"{r[1]} · {r[2]}",
                    "detail": f"proceso: {r[3] or 'No disponible'} (PID: {r[4] if r[4] is not None else 'No disponible'})"
                }
                for r in cursor.fetchall()
            ]

            timeline.sort(key=lambda i: i["at_raw"], reverse=True)
            for item in timeline:
                del item["at_raw"]

            # PROCESO INVOLUCRADO (sección 5 de "Corrección definitiva en
            # la lógica y presentación de ALERTAS", 2026-08-18, ver
            # PENDIENTES.md): 'alerts'/'incidents' NUNCA tuvieron columna
            # de proceso -- no se inventa un vínculo directo que la base
            # no tiene (ver PENDIENTES.md, "Alerta ↔ eventos que la
            # disparó", revertido a propósito en la reestructuración).
            # Se reutiliza la MISMA correlación aproximada por ventana de
            # tiempo que ya arma la 'cadena de evidencia' de arriba
            # (window_start/window_end, 5 min antes / 1 min después del
            # ancla del episodio) -- no una relación nueva, solo se le
            # pide un dato más: el proceso real, si el agente pudo
            # atribuirlo (agent/adapters/, best-effort, puede no estar
            # disponible). Prioridad: honeyfile_activations primero si
            # esta alerta/incidente incluye 'Acceso Honeyfile' (sección 7:
            # es la señal más fuerte y su atribución de proceso ya es la
            # más confiable del sistema), si no, el evento de archivo más
            # reciente en la ventana que sí tenga proceso atribuido.
            # 'ruta'/'usuario' NUNCA se completan -- ni 'events' ni
            # 'honeyfile_activations' tienen esas columnas (el agente las
            # calcula en memoria vía agent/adapters/ pero nunca las manda,
            # ver PENDIENTES.md, "Atribución de proceso en eventos de
            # archivo") -- se muestran honestamente como "No disponible"
            # en vez de fabricar un valor.
            process_name = None
            process_id = None

            if is_honeyfile:
                cursor.execute(
                    """
                    SELECT process_name, process_id FROM honeyfile_activations
                    WHERE agent_id = %s AND detected_at BETWEEN %s AND %s
                      AND (process_name IS NOT NULL OR process_id IS NOT NULL)
                    ORDER BY detected_at DESC LIMIT 1;
                    """,
                    (agent_id, window_start, window_end)
                )
                proc_row = cursor.fetchone()
                if proc_row:
                    process_name, process_id = proc_row

            if process_name is None and process_id is None:
                cursor.execute(
                    """
                    SELECT process_name, process_id FROM events
                    WHERE agent_id = %s AND detected_at BETWEEN %s AND %s
                      AND (process_name IS NOT NULL OR process_id IS NOT NULL)
                    ORDER BY detected_at DESC LIMIT 1;
                    """,
                    (agent_id, window_start, window_end)
                )
                proc_row = cursor.fetchone()
                if proc_row:
                    process_name, process_id = proc_row

            process = {
                "process_name": process_name,
                "process_id": process_id,
                # Ninguna tabla real tiene estas dos columnas hoy -- ver
                # comentario arriba. Siempre 'None' (el frontend lo
                # muestra como "No disponible"), a propósito, no un
                # descuido.
                "executable_path": None,
                "username": None,
            }

            # Incidente sobre el que aplicaría un aislamiento manual
            # desde este drawer (2026-08-17, ver PENDIENTES.md,
            # "Aislamiento de host -- modo development, laboratorio y
            # producción") -- para kind == 'incident' es el propio
            # item_id; para kind == 'alert' es incident_id_val (puede
            # ser None si la alerta todavía no escaló a incidente, en
            # cuyo caso no hay a qué asociar la orden). 'isolation_status'
            # es el estado más reciente de host_isolations para ese
            # incidente, o None si nunca se ordenó ninguno.
            isolatable_incident_id = item_id if kind == "incident" else incident_id_val
            # Corregido 2026-08-18 (problema H, ver PENDIENTES.md): antes
            # esto filtraba 'host_isolations.incident_id = %s' contra
            # 'isolatable_incident_id' -- si el aislamiento se había
            # ordenado desde OTRO incidente de este mismo endpoint (o
            # automáticamente desde una alerta que después escaló a un
            # incidente DISTINTO), este drawer no lo encontraba y ofrecía
            # "Aislar" de nuevo sobre un endpoint ya aislado. El
            # aislamiento pertenece al AGENTE (agent_id, ya resuelto más
            # arriba en esta función), no al incidente que lo originó --
            # ver _agent_isolation_status_sql()/_agent_isolation_id_sql().
            cursor.execute(
                f"""
                SELECT {_agent_isolation_id_sql("%s")}, {_agent_isolation_status_sql("%s")};
                """,
                (agent_id, agent_id)
            )
            iso_row = cursor.fetchone()
            isolation_id, isolation_status = iso_row if iso_row else (None, None)

    finally:
        connection.close()

    return {
        "kind": kind,
        "id": item_id,
        "code": code,
        "title": title_val,
        "description": description_val,
        "status": status,
        "status_label": status_label,
        "severity": severity,
        "risk_score": risk_score,
        "classification": classification,
        "classification_label": INCIDENT_CLASSIFICATION_LABELS_ES.get(classification, "Sin clasificar") if kind == "incident" else None,
        "assigned_to": assigned_to,
        "assigned_to_name": assigned_to_name,
        "hostname": hostname,
        "ip_address": str(ip_address) if ip_address else "127.0.0.1",
        "operating_system": operating_system,
        "is_online": is_online,
        "agent_id": agent_id,
        "detection_count": detection_count,
        "is_honeyfile": is_honeyfile,
        "incident_id": incident_id_val,
        "resolved_at": resolved_at_val.strftime("%d/%m/%Y %H:%M:%S") if resolved_at_val else None,
        "rules": matched_rules,
        # Proceso involucrado (sección 5, 2026-08-18, ver PENDIENTES.md)
        # -- ver el bloque de arriba para de dónde sale cada campo y por
        # qué 'executable_path'/'username' quedan siempre en None.
        "process": process,
        "timeline": timeline,
        # 'anchor_ts' ya se calculaba antes (para la ventana de la
        # cadena de evidencia) pero nunca se devolvía -- se agrega acá
        # tal cual, sin inventar nada nuevo. Para una alerta suelta es
        # su fecha de creación real; para un incidente, la de la
        # primera alerta vinculada (o su apertura si no tiene ninguna).
        "created_at": anchor_ts.strftime("%d/%m/%Y %H:%M:%S") if anchor_ts else None,
        # Solo viene poblado para kind == 'incident': la alerta que dio
        # origen al caso (la primera por fecha entre las vinculadas),
        # para el bloque "Alerta de origen" del drawer.
        "origin_alert": origin_alert,
        "isolatable_incident_id": isolatable_incident_id,
        "isolation_status": isolation_status,
        # Agregado 2026-08-17 (ver PENDIENTES.md, "Corrección de tiempo
        # real, ordenamiento y consistencia") -- lo que necesita el
        # nuevo botón "Liberar" del drawer para llamar a
        # POST /host-isolations/{id}/release sin tener que adivinar el
        # id de la fila.
        "isolation_id": isolation_id,
    }


@app.get("/api/alerts")
def api_alerts(
    search: str = "",
    severity: str = Query("", pattern="^(MEDIO|ALTO|CRÍTICO|)$"),
    status: str = Query("", pattern="^(NEW|ACKNOWLEDGED|ESCALATED|CLOSED|FALSE_POSITIVE|)$"),
    since: str = Query("", pattern="^(24h|7d|30d|)$"),
    rule: str = "",
    # Vista operativa vs. historial (2026-08-18, ver PENDIENTES.md,
    # "Revisión y corrección integral de ALFA-Sentinel", problema G):
    # antes esta pantalla mostraba SIEMPRE el historial completo, alertas
    # cerradas/falsos positivos incluidos, sin ningún filtro por
    # defecto -- el analista tenía que armar el filtro de estado a mano
    # cada vez para ver solo lo que necesita atención. 'activas' (default)
    # excluye CLOSED/FALSE_POSITIVE -- los 2 estados finales reales de
    # 'alerts.status' (ver ALERT_STATUS_LABELS_ES) -- sin inventar ningún
    # estado nuevo. 'todos' quita esa exclusión. Si 'status' viene
    # explícito (el analista eligió un estado puntual del desplegable,
    # incluida una de las cerradas), ESE filtro manda siempre, sin
    # importar 'view' -- elegir un estado a propósito es más específico
    # que la vista general. No se borra ningún registro -- 'todos' sigue
    # trayendo el historial completo real.
    view: str = Query("activas", pattern="^(activas|todos)$"),
    page: int = Query(1, ge=1),
    page_size: int = Query(15, ge=1, le=100),
    user: dict = Depends(get_current_user)
):
    """Lista dedicada de alertas para la pantalla Alertas en React --
    a diferencia de /api/incidentes, que unifica incidentes agrupados
    y alertas sueltas en un solo listado (COMBINED_CTE), acá se listan
    únicamente filas de 'alerts' tal cual, sin agrupar. Mismas tablas,
    mismos nombres de estado/severidad que el resto del sistema
    (ALERT_STATUS_LABELS_ES; la severidad ya no tiene diccionario
    paralelo, ver severity_levels)."""
    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            where_clauses = []
            params = {}

            if not status and view == "activas":
                where_clauses.append("alerts.status NOT IN ('CLOSED', 'FALSE_POSITIVE')")

            if search:
                # Busca sobre 'alerts.title' -- la columna GUARDADA
                # (texto del primer evento del episodio, ej. "...--
                # python.exe"), que sigue existiendo tal cual aunque ya
                # no se use como título visible (ver alert_general_title
                # más abajo). Se deja así a propósito: sigue siendo un
                # texto real y buscable (nombres de proceso, etc.) --
                # cambiar qué se busca no fue pedido y no es necesario
                # para corregir el título.
                where_clauses.append("(endpoints.hostname ILIKE %(search)s OR alerts.title ILIKE %(search)s)")
                params["search"] = f"%{search}%"
            if severity:
                where_clauses.append("severity_levels.name = %(severity)s")
                params["severity"] = severity
            if status:
                where_clauses.append("alerts.status = %(status)s")
                params["status"] = status
            if since:
                interval = {"24h": "24 hours", "7d": "7 days", "30d": "30 days"}[since]
                where_clauses.append(f"alerts.created_at >= CURRENT_TIMESTAMP - INTERVAL '{interval}'")
            if rule:
                # EXISTS en vez de JOIN (2026-08-18, ver PENDIENTES.md,
                # "Corrección definitiva en la lógica y presentación de
                # ALERTAS") -- un JOIN contra alert_rule/heuristic_rules
                # multiplica filas cuando una alerta tiene más de una
                # regla vinculada (la causa real del bug de título/regla
                # arbitraria que reportó el usuario). EXISTS filtra sin
                # multiplicar nada.
                where_clauses.append(
                    """EXISTS (
                        SELECT 1 FROM alert_rule
                        JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                        WHERE alert_rule.alert_id = alerts.id AND heuristic_rules.name = %(rule)s
                    )"""
                )
                params["rule"] = rule

            where_sql = ("WHERE " + " AND ".join(where_clauses)) if where_clauses else ""

            # Sin LEFT JOIN alert_rule/heuristic_rules acá -- no hace
            # falta para nada de lo que se selecciona (severidad/status/
            # fecha ya viven en 'alerts', el filtro por regla usa EXISTS
            # arriba). Antes estaba solo para poder mostrar UN nombre de
            # regla por alerta más abajo, lo que obligaba a un DISTINCT
            # ON frágil (ver historial del bug en PENDIENTES.md,
            # "Corrección de tiempo real, ordenamiento y consistencia").
            # Ahora el título es general por severidad y la cantidad de
            # reglas se cuenta con una subconsulta escalar -- no hay
            # ningún JOIN que pueda multiplicar filas en esta consulta.
            base_from = """
                FROM alerts
                JOIN severity_levels ON severity_levels.id = alerts.severity_id
                JOIN agents ON agents.id = alerts.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
            """

            # Resumen -- sin filtrar, para las 5 tarjetas de arriba.
            cursor.execute(
                f"""
                SELECT
                    COUNT(*) AS total_n,
                    COUNT(*) FILTER (WHERE alerts.status NOT IN ('CLOSED', 'FALSE_POSITIVE')) AS active_n,
                    COUNT(*) FILTER (WHERE severity_levels.name = 'CRÍTICO' AND alerts.status NOT IN ('CLOSED', 'FALSE_POSITIVE')) AS critical_n,
                    COUNT(*) FILTER (WHERE alerts.status = 'ACKNOWLEDGED') AS investigating_n,
                    COUNT(*) FILTER (WHERE alerts.status IN ('CLOSED', 'FALSE_POSITIVE')) AS resolved_n
                {base_from};
                """
            )
            total_n, active_n, critical_n, investigating_n, resolved_n = cursor.fetchone()

            cursor.execute("SELECT DISTINCT name FROM heuristic_rules ORDER BY name;")
            rule_names = [r[0] for r in cursor.fetchall()]

            cursor.execute(
                f"SELECT COUNT(*) {base_from} {where_sql};",
                params
            )
            filtered_total = cursor.fetchone()[0]

            total_pages = max(1, -(-filtered_total // page_size))
            current_page = min(page, total_pages)
            offset = (current_page - 1) * page_size

            page_params = dict(params)
            page_params["limit"] = page_size
            page_params["offset"] = offset

            # Sin DISTINCT ON: cada alerta aparece en 'alerts' una sola
            # vez (a diferencia del join anterior contra alert_rule, que
            # la multiplicaba por cada regla vinculada), así que un
            # ORDER BY / LIMIT / OFFSET directo sobre 'alerts.created_at
            # DESC' ya es correcto en todas las páginas -- ver
            # PENDIENTES.md, "Corrección de tiempo real, ordenamiento y
            # consistencia" para el bug original de esta misma consulta.
            # 'rule_count' (2026-08-18, ver PENDIENTES.md, "Corrección
            # definitiva en la lógica y presentación de ALERTAS")
            # reemplaza al nombre de una sola regla -- la tabla ya no
            # muestra una regla individual como si fuera representativa
            # de toda la alerta, solo cuántas contribuyeron; el detalle
            # (drawer) sigue listando cada una.
            cursor.execute(
                f"""
                SELECT
                    alerts.id, severity_levels.name, endpoints.hostname,
                    alerts.risk_score, alerts.status, alerts.created_at, alerts.incident_id,
                    alerts.agent_id,
                    (
                        SELECT COUNT(*) FROM alert_rule
                        WHERE alert_rule.alert_id = alerts.id
                    ) AS rule_count
                {base_from}
                {where_sql}
                ORDER BY alerts.created_at DESC
                LIMIT %(limit)s OFFSET %(offset)s;
                """,
                page_params
            )
            rows = cursor.fetchall()
    finally:
        connection.close()

    alerts = [
        {
            "id": r[0],
            "severity": r[1],
            # Título general por severidad -- ver alert_general_title()
            # (2026-08-18, ver PENDIENTES.md, "Corrección definitiva en
            # la lógica y presentación de ALERTAS"). Ya NO es el nombre
            # de una regla individual ("Consumo de CPU elevado",
            # "Acceso Honeyfile", etc.) -- esas reglas siguen existiendo,
            # se ven todas juntas en el detalle de la alerta.
            "title": alert_general_title(r[1]),
            "hostname": r[2],
            "risk_score": float(r[3]),
            "status": r[4],
            "status_label": ALERT_STATUS_LABELS_ES.get(r[4], r[4]),
            "created_at": r[5].strftime("%d/%m/%Y %H:%M:%S"),
            "incident_id": r[6],
            "agent_id": r[7],
            "rule_count": r[8],
        }
        for r in rows
    ]

    return {
        "summary": {
            "total": total_n,
            "active": active_n,
            "critical": critical_n,
            "investigating": investigating_n,
            "resolved": resolved_n,
        },
        "rules": [{"value": n, "label": n} for n in rule_names],
        "page": current_page,
        "page_size": page_size,
        "total_pages": total_pages,
        "filtered_total": filtered_total,
        "alerts": alerts,
    }


def _alfa_kv_table(data):
    """Tabla de dos columnas (campo/valor) con el mismo look en todas
    las secciones de la ficha del reporte PDF."""

    table = Table(data, colWidths=[4 * cm, 12.5 * cm])
    table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("TEXTCOLOR", (0, 0), (0, -1), colors.HexColor("#6b7280")),
        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
    ]))
    return table


def _alfa_table(data, col_widths=None):
    """Tabla con encabezado (primera fila) para listas dentro del
    reporte PDF -- reglas disparadas, cadena de evidencia."""

    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("FONTSIZE", (0, 0), (-1, -1), 8.5),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#111827")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f9fafb")]),
    ]))
    return table


@app.get("/api/respuesta")
def api_respuesta(user: dict = Depends(get_current_user)):
    """Datos de la pantalla Acciones de Respuesta en React.

    Desde la corrección definitiva del motor heurístico (2026-08-17,
    ver PENDIENTES.md), el aislamiento AUTOMÁTICO es real de punta a
    punta: cuando report_alert() determina que corresponde (sección
    30), deja una orden 'REQUESTED' en 'host_isolations' que el agente
    del endpoint recoge y ejecuta (agent/isolation_sync.py +
    agent/isolation_executor.py), confirmando 'EXECUTED' o
    'ISOLATION_FAILED' según el resultado real.

    Extendido 2026-08-17 (ver PENDIENTES.md, "Aislamiento de host --
    modo development, laboratorio y producción"): el disparo MANUAL
    desde la consola (POST /incidents/{id}/isolate) ya existe y usa
    exactamente el mismo mecanismo -- esta pantalla ahora también deja
    disparar aislamiento manual sobre los incidentes críticos listados
    abajo, y liberar (POST /host-isolations/{id}/release) un
    aislamiento ya ejecutado desde el historial.

    Esta pantalla muestra:
    1. El estado real de 'host_isolations' (historial completo, con el
       resultado que reportó cada agente, incluido quién lo solicitó
       cuando fue manual).
    2. Los incidentes críticos abiertos ahora mismo, con un botón para
       aislar manualmente si todavía no tienen una orden en curso.
    """

    connection = get_connection()
    try:
        with connection.cursor() as cursor:

            cursor.execute(
                "SELECT COUNT(DISTINCT agent_id) FROM host_isolations WHERE status IN ('REQUESTED', 'EXECUTED', 'RELEASE_REQUESTED') AND released_at IS NULL;"
            )
            isolated_now = cursor.fetchone()[0]

            cursor.execute("SELECT COUNT(*) FROM host_isolations;")
            total_isolations = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT host_isolations.id, host_isolations.isolation_type, host_isolations.status,
                       host_isolations.reason, host_isolations.requested_at, host_isolations.executed_at,
                       host_isolations.released_at, host_isolations.result,
                       endpoints.hostname, users.full_name, host_isolations.incident_id
                FROM host_isolations
                JOIN agents ON agents.id = host_isolations.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                LEFT JOIN users ON users.id = host_isolations.requested_by
                ORDER BY host_isolations.requested_at DESC
                LIMIT 50;
                """
            )
            isolation_rows = cursor.fetchall()

            # 'isolation_status'/'isolation_id' corregidos 2026-08-18
            # (problema H, ver PENDIENTES.md): antes filtraban por
            # 'host_isolations.incident_id = incidents.id' -- la misma
            # tabla de "Incidentes críticos" de Respuesta mostraba
            # "Aislar" sobre un incidente cuyo endpoint YA estaba aislado
            # desde otro incidente distinto. Ver
            # _agent_isolation_status_sql()/_agent_isolation_id_sql().
            cursor.execute(
                f"""
                SELECT incidents.id, incidents.title, incidents.status, incidents.opened_at,
                       endpoints.hostname, incidents.assigned_to, assigned_user.full_name,
                       (
                           SELECT severity_levels.name FROM alerts
                           JOIN severity_levels ON severity_levels.id = alerts.severity_id
                           WHERE alerts.incident_id = incidents.id
                           ORDER BY severity_levels.min_score DESC
                           LIMIT 1
                       ) AS severity,
                       {_agent_isolation_status_sql("incidents.agent_id")} AS isolation_status,
                       {_agent_isolation_id_sql("incidents.agent_id")} AS isolation_id
                FROM incidents
                JOIN agents ON agents.id = incidents.agent_id
                JOIN endpoints ON endpoints.id = agents.endpoint_id
                LEFT JOIN users AS assigned_user ON assigned_user.id = incidents.assigned_to
                WHERE incidents.status != 'CLOSED'
                ORDER BY incidents.opened_at DESC;
                """
            )
            incident_rows = cursor.fetchall()
    finally:
        connection.close()

    critical_incidents = [
        {
            "id": r[0],
            "code": f"INC-{r[0]:05d}",
            # Título general por severidad -- ver alert_general_title()
            # (2026-08-18, ver PENDIENTES.md, "Corrección definitiva en
            # la lógica y presentación de ALERTAS"). Un incidente
            # agrupa una o más alertas -- el título ya no es el texto
            # del primer evento que lo originó.
            "title": alert_general_title(r[7]),
            "status": r[2],
            "status_label": INCIDENT_STATUS_LABELS_ES.get(r[2], r[2]),
            "opened_at": r[3].strftime("%d/%m/%Y %H:%M:%S") if r[3] else None,
            "hostname": r[4],
            "assigned_to": r[5],
            "assigned_to_name": r[6],
            "severity": r[7],
            "isolation_status": r[8],
            "isolation_id": r[9],
        }
        for r in incident_rows
        if r[7] in ("ALTO", "CRÍTICO")
    ]

    isolations = [
        {
            "id": r[0],
            "isolation_type": r[1],
            "isolation_type_label": ISOLATION_TYPE_LABELS_ES.get(r[1], r[1]),
            "status": r[2],
            "status_label": ISOLATION_STATUS_LABELS_ES.get(r[2], r[2]),
            "reason": r[3],
            "requested_at": r[4].strftime("%d/%m/%Y %H:%M:%S") if r[4] else None,
            "executed_at": r[5].strftime("%d/%m/%Y %H:%M:%S") if r[5] else None,
            "released_at": r[6].strftime("%d/%m/%Y %H:%M:%S") if r[6] else None,
            "result": r[7],
            "hostname": r[8],
            "requested_by_name": r[9],
            "incident_id": r[10],
        }
        for r in isolation_rows
    ]

    return {
        "summary": {
            "isolated_now": isolated_now,
            "total_isolations": total_isolations,
            "critical_incidents_open": len(critical_incidents),
        },
        "critical_incidents": critical_incidents,
        "isolations": isolations,
    }


def _alfa_styles():
    """Hoja de estilos compartida por los generadores de PDF de
    /reportes (mismo look que el reporte de un incidente puntual,
    /incidentes/{id}/reporte.pdf)."""

    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="AlfaTitle", parent=styles["Title"], fontSize=18, spaceAfter=4))
    styles.add(ParagraphStyle(name="AlfaSubtitle", parent=styles["Normal"], fontSize=9.5, textColor=colors.HexColor("#6b7280"), spaceAfter=16))
    styles.add(ParagraphStyle(name="AlfaSection", parent=styles["Heading2"], fontSize=13, spaceBefore=14, spaceAfter=6, textColor=colors.HexColor("#111827")))
    styles.add(ParagraphStyle(name="AlfaNote", parent=styles["Normal"], fontSize=8.5, textColor=colors.HexColor("#6b7280"), spaceAfter=10))
    return styles


def _xlsx_section(ws, title, headers, rows):
    """Sección repetible dentro de una hoja de cálculo de /reportes:
    título en negrita, encabezado opcional (para tablas) y filas de
    datos, seguidos de una fila en blanco de separación."""

    ws.append([title])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)

    if headers:
        ws.append(headers)
        header_row = ws.max_row
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="111827", end_color="111827", fill_type="solid")

    for r in rows:
        ws.append(r)

    ws.append([])


# --- Recolección de datos (una función por tipo de informe) ---
# Todas devuelven solo lo que sale de una consulta real -- ninguna
# arma "proceso ejecutado", "línea de comando" ni hashes que el agente
# no reporta hoy (ver PENDIENTES.md, "Atribución de proceso en
# eventos de archivo").

def _gather_security_report_data(cursor, start, end, endpoint_id):
    params = {"start": start, "end": end}
    ep_filter = ""
    hf_filter = ""
    if endpoint_id:
        ep_filter = "AND endpoints.id = %(endpoint_id)s"
        hf_filter = "AND agents.endpoint_id = %(endpoint_id)s"
        params["endpoint_id"] = endpoint_id

    cursor.execute(
        f"""
        SELECT severity_levels.name, COUNT(*)
        FROM alerts
        JOIN severity_levels ON severity_levels.id = alerts.severity_id
        JOIN agents ON agents.id = alerts.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        WHERE alerts.created_at BETWEEN %(start)s AND %(end)s
        {ep_filter}
        GROUP BY severity_levels.name;
        """,
        params
    )
    severity_counts = dict(cursor.fetchall())

    cursor.execute(
        f"""
        SELECT incidents.status, COUNT(*)
        FROM incidents
        JOIN agents ON agents.id = incidents.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        WHERE incidents.opened_at BETWEEN %(start)s AND %(end)s
        {ep_filter}
        GROUP BY incidents.status;
        """,
        params
    )
    incident_status_counts = dict(cursor.fetchall())

    cursor.execute(
        f"""
        SELECT incidents.classification, COUNT(*)
        FROM incidents
        JOIN agents ON agents.id = incidents.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        WHERE incidents.opened_at BETWEEN %(start)s AND %(end)s
          AND incidents.classification IS NOT NULL
        {ep_filter}
        GROUP BY incidents.classification;
        """,
        params
    )
    classification_counts = dict(cursor.fetchall())

    cursor.execute(
        f"""
        SELECT heuristic_rules.name, COUNT(*)
        FROM alert_rule
        JOIN alerts ON alerts.id = alert_rule.alert_id
        JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
        JOIN agents ON agents.id = alerts.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        WHERE alerts.created_at BETWEEN %(start)s AND %(end)s
        {ep_filter}
        GROUP BY heuristic_rules.name
        ORDER BY COUNT(*) DESC;
        """,
        params
    )
    rule_counts = cursor.fetchall()

    cursor.execute(
        f"""
        SELECT
            COUNT(*) FILTER (WHERE honeyfiles.status = 'ACTIVE'),
            COUNT(*) FILTER (WHERE honeyfiles.status = 'TRIGGERED'),
            COUNT(*)
        FROM honeyfiles
        JOIN agents ON agents.id = honeyfiles.agent_id
        WHERE 1=1 {hf_filter};
        """,
        params
    )
    hf_active, hf_triggered, hf_total = cursor.fetchone()

    cursor.execute(
        f"""
        SELECT AVG(EXTRACT(EPOCH FROM (alerts.resolved_at - alerts.created_at)))
        FROM alerts
        JOIN agents ON agents.id = alerts.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        WHERE alerts.resolved_at IS NOT NULL
          AND alerts.created_at BETWEEN %(start)s AND %(end)s
        {ep_filter};
        """,
        params
    )
    avg_resolution_seconds = cursor.fetchone()[0]

    return {
        "severity_counts": severity_counts,
        "total_alerts": sum(severity_counts.values()),
        "incident_status_counts": incident_status_counts,
        "total_incidents": sum(incident_status_counts.values()),
        "classification_counts": classification_counts,
        "rule_counts": rule_counts,
        "honeyfiles_active": hf_active or 0,
        "honeyfiles_triggered": hf_triggered or 0,
        "honeyfiles_total": hf_total or 0,
        "avg_resolution_seconds": float(avg_resolution_seconds) if avg_resolution_seconds is not None else None,
    }


def _gather_endpoints_report_data(cursor, start, end, endpoint_id):
    params = {"start": start, "end": end}
    ep_filter = ""
    if endpoint_id:
        ep_filter = "AND endpoints.id = %(endpoint_id)s"
        params["endpoint_id"] = endpoint_id

    stale_seconds = get_agent_stale_seconds(cursor)

    # 'is_online' se calcula en SQL (CURRENT_TIMESTAMP, con el mismo
    # criterio que _endpoint_cte() en /endpoints) y no en Python --
    # restar 'datetime.now()' (naive) de 'agents.last_seen_at'
    # (TIMESTAMPTZ, llega con tz desde psycopg) revienta con "can't
    # subtract offset-naive and offset-aware datetimes" apenas hay un
    # agente ONLINE real con last_seen_at seteado. El mismo bug late
    # en get_incidente_drawer -- corregido ahí también.
    cursor.execute(
        f"""
        SELECT endpoints.hostname, endpoints.ip_address, endpoints.os,
               agents.status,
               (
                   agents.status = 'ONLINE'
                   AND agents.last_seen_at >= CURRENT_TIMESTAMP - INTERVAL '{stale_seconds} seconds'
               ) AS is_online,
               (
                   SELECT COUNT(*) FROM events
                   WHERE events.agent_id = agents.id
                     AND events.detected_at BETWEEN %(start)s AND %(end)s
               ) AS event_count,
               (
                   SELECT COUNT(*) FROM honeyfiles
                   WHERE honeyfiles.agent_id = agents.id
               ) AS honeyfiles_deployed,
               (
                   SELECT COUNT(*) FROM agent_honeyfile_templates
                   WHERE agent_honeyfile_templates.agent_id = agents.id
                     AND agent_honeyfile_templates.status = 'PENDING'
               ) AS honeyfiles_pending
        FROM agents
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        WHERE 1=1 {ep_filter}
        ORDER BY endpoints.hostname ASC;
        """,
        params
    )
    rows = cursor.fetchall()

    endpoints_data = []
    for hostname, ip, os_name, status, is_online, event_count, hf_deployed, hf_pending in rows:
        status_label = "En línea" if is_online else ("Sin señal reciente" if status == "ONLINE" else "Desconectado")
        endpoints_data.append({
            "hostname": hostname,
            "ip_address": str(ip) if ip else "—",
            "os": os_name,
            "is_online": is_online,
            "status_label": status_label,
            "event_count": event_count,
            "honeyfiles_deployed": hf_deployed,
            "honeyfiles_pending": hf_pending
        })

    return {
        "endpoints": endpoints_data,
        "total_endpoints": len(endpoints_data),
        "online_count": sum(1 for e in endpoints_data if e["is_online"]),
        "total_events": sum(e["event_count"] for e in endpoints_data),
    }


def _gather_incidents_report_data(cursor, start, end, endpoint_id):
    params = {"start": start, "end": end}
    ep_filter = ""
    if endpoint_id:
        ep_filter = "AND endpoints.id = %(endpoint_id)s"
        params["endpoint_id"] = endpoint_id

    cursor.execute(
        f"""
        SELECT incidents.id, incidents.title, incidents.status, incidents.classification,
               incidents.opened_at, incidents.closed_at, endpoints.hostname,
               assigned_user.full_name,
               (
                   SELECT COALESCE(MAX(alerts.risk_score), 0) FROM alerts
                   WHERE alerts.incident_id = incidents.id
               ) AS risk_score,
               (
                   SELECT STRING_AGG(DISTINCT heuristic_rules.name, ' + ') FROM alerts
                   LEFT JOIN alert_rule ON alert_rule.alert_id = alerts.id
                   LEFT JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                   WHERE alerts.incident_id = incidents.id
               ) AS rule_names,
               (
                   SELECT COUNT(*) FROM alerts WHERE alerts.incident_id = incidents.id
               ) AS alert_count,
               (
                   SELECT severity_levels.name FROM alerts
                   JOIN severity_levels ON severity_levels.id = alerts.severity_id
                   WHERE alerts.incident_id = incidents.id
                   ORDER BY severity_levels.min_score DESC
                   LIMIT 1
               ) AS severity
        FROM incidents
        JOIN agents ON agents.id = incidents.agent_id
        JOIN endpoints ON endpoints.id = agents.endpoint_id
        LEFT JOIN users AS assigned_user ON assigned_user.id = incidents.assigned_to
        WHERE incidents.opened_at BETWEEN %(start)s AND %(end)s
        {ep_filter}
        ORDER BY incidents.opened_at DESC;
        """,
        params
    )
    rows = cursor.fetchall()

    incidents_data = [
        {
            "id": r[0], "code": f"INC-{r[0]:05d}",
            # Título general por severidad -- ver alert_general_title()
            # (2026-08-18, ver PENDIENTES.md, "Corrección definitiva en
            # la lógica y presentación de ALERTAS").
            "title": alert_general_title(r[11]),
            "status": r[2], "status_label": INCIDENT_STATUS_LABELS_ES.get(r[2], r[2]),
            "classification_label": INCIDENT_CLASSIFICATION_LABELS_ES.get(r[3], "Sin clasificar"),
            "opened_at": r[4], "closed_at": r[5], "hostname": r[6],
            "assigned_to_name": r[7] or "Sin asignar", "risk_score": float(r[8]),
            "rule_label": r[9] if r[9] else "Sin regla vinculada",
            "alert_count": r[10]
        }
        for r in rows
    ]

    return {
        "incidents": incidents_data,
        "total_incidents": len(incidents_data),
        "closed_count": sum(1 for i in incidents_data if i["status"] == "CLOSED"),
        "open_count": sum(1 for i in incidents_data if i["status"] != "CLOSED"),
    }


REPORT_DATA_GATHERERS = {
    "SECURITY": _gather_security_report_data,
    "ENDPOINTS": _gather_endpoints_report_data,
    "INCIDENTS": _gather_incidents_report_data,
}


# --- Generadores de PDF (reportlab) ---

def _build_security_pdf(data, meta):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.8 * cm, bottomMargin=1.8 * cm, leftMargin=1.8 * cm, rightMargin=1.8 * cm)
    styles = _alfa_styles()
    story = [Paragraph(f"ALFA-Sentinel &mdash; {meta['title']}", styles["AlfaTitle"]), Paragraph(meta["subtitle"], styles["AlfaSubtitle"])]

    story.append(Paragraph("1. Resumen de Alertas por Severidad", styles["AlfaSection"]))
    sev_rows = [["Severidad", "Cantidad"]]
    for key in ("CRÍTICO", "ALTO", "MEDIO"):
        sev_rows.append([key, str(data["severity_counts"].get(key, 0))])
    sev_rows.append(["Total", str(data["total_alerts"])])
    story.append(_alfa_table(sev_rows))

    story.append(Paragraph("2. Incidentes por Estado", styles["AlfaSection"]))
    if data["incident_status_counts"]:
        inc_rows = [["Estado", "Cantidad"]] + [[INCIDENT_STATUS_LABELS_ES.get(s, s), str(n)] for s, n in data["incident_status_counts"].items()]
        story.append(_alfa_table(inc_rows))
    else:
        story.append(Paragraph("No se abrieron incidentes en el período evaluado.", styles["Normal"]))

    if data["classification_counts"]:
        story.append(Spacer(1, 8))
        story.append(Paragraph("Clasificación de los incidentes cerrados en el período:", styles["Normal"]))
        cls_rows = [["Clasificación", "Cantidad"]] + [[INCIDENT_CLASSIFICATION_LABELS_ES.get(c, c), str(n)] for c, n in data["classification_counts"].items()]
        story.append(_alfa_table(cls_rows))

    story.append(Paragraph("3. Reglas Heurísticas Más Activas", styles["AlfaSection"]))
    if data["rule_counts"]:
        rule_rows = [["Regla", "Alertas disparadas"]] + [[name or "Sin regla vinculada", str(n)] for name, n in data["rule_counts"]]
        story.append(_alfa_table(rule_rows))
    else:
        story.append(Paragraph("No hubo alertas en el período evaluado.", styles["Normal"]))

    story.append(Paragraph("4. Cobertura de Honeyfiles", styles["AlfaSection"]))
    coverage_pct = round((data["honeyfiles_active"] + data["honeyfiles_triggered"]) / data["honeyfiles_total"] * 100, 1) if data["honeyfiles_total"] else 0
    hf_kv = [
        ["Honeyfiles desplegados", str(data["honeyfiles_total"])],
        ["Activos (intactos)", str(data["honeyfiles_active"])],
        ["Activados (comprometidos)", str(data["honeyfiles_triggered"])],
        ["Cobertura operativa", f"{coverage_pct}%"],
    ]
    story.append(_alfa_kv_table([[Paragraph(k, styles["Normal"]), Paragraph(v, styles["Normal"])] for k, v in hf_kv]))

    story.append(Paragraph("5. Tiempo de Resolución", styles["AlfaSection"]))
    if data["avg_resolution_seconds"] is not None:
        minutes = round(data["avg_resolution_seconds"] / 60, 1)
        story.append(Paragraph(f"Tiempo promedio de resolución de alertas en el período: {minutes} minutos.", styles["Normal"]))
    else:
        story.append(Paragraph("Sin datos aún -- no hay alertas resueltas en el período evaluado.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _build_endpoints_pdf(data, meta):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.8 * cm, bottomMargin=1.8 * cm, leftMargin=1.8 * cm, rightMargin=1.8 * cm)
    styles = _alfa_styles()
    story = [Paragraph(f"ALFA-Sentinel &mdash; {meta['title']}", styles["AlfaTitle"]), Paragraph(meta["subtitle"], styles["AlfaSubtitle"])]

    story.append(Paragraph("1. Resumen", styles["AlfaSection"]))
    resumen = [
        ["Endpoints evaluados", str(data["total_endpoints"])],
        ["En línea al momento de generar", str(data["online_count"])],
        ["Eventos totales en el período", str(data["total_events"])],
    ]
    story.append(_alfa_kv_table([[Paragraph(k, styles["Normal"]), Paragraph(v, styles["Normal"])] for k, v in resumen]))

    story.append(Paragraph("2. Detalle por Endpoint", styles["AlfaSection"]))
    if data["endpoints"]:
        rows = [["Endpoint", "SO", "Conectividad", "Eventos", "Honeyfiles (desplegados / pendientes)"]]
        for ep in data["endpoints"]:
            rows.append([
                Paragraph(f"{ep['hostname']}<br/>{ep['ip_address']}", styles["Normal"]),
                ep["os"] or "—", ep["status_label"], str(ep["event_count"]),
                f"{ep['honeyfiles_deployed']} / {ep['honeyfiles_pending']}"
            ])
        story.append(_alfa_table(rows, col_widths=[4.5 * cm, 2.3 * cm, 3 * cm, 2 * cm, 4.5 * cm]))
    else:
        story.append(Paragraph("No hay endpoints registrados que coincidan con el filtro.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


def _build_incidents_pdf(data, meta):
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter, topMargin=1.8 * cm, bottomMargin=1.8 * cm, leftMargin=1.8 * cm, rightMargin=1.8 * cm)
    styles = _alfa_styles()
    story = [Paragraph(f"ALFA-Sentinel &mdash; {meta['title']}", styles["AlfaTitle"]), Paragraph(meta["subtitle"], styles["AlfaSubtitle"])]

    story.append(Paragraph(
        "Este informe es un resumen de auditoría de los incidentes abiertos en el período -- no "
        "incluye la traza técnica completa de cada uno (proceso, hashes, cadena de evidencia). "
        "Para el detalle forense de un incidente puntual, usar su Reporte PDF individual desde el "
        "Expediente en Incidentes y Alertas.",
        styles["AlfaNote"]
    ))

    story.append(Paragraph("1. Resumen", styles["AlfaSection"]))
    resumen = [
        ["Incidentes en el período", str(data["total_incidents"])],
        ["Cerrados", str(data["closed_count"])],
        ["Abiertos / en curso", str(data["open_count"])],
    ]
    story.append(_alfa_kv_table([[Paragraph(k, styles["Normal"]), Paragraph(v, styles["Normal"])] for k, v in resumen]))

    story.append(Paragraph("2. Detalle de Incidentes", styles["AlfaSection"]))
    if data["incidents"]:
        rows = [["ID", "Título", "Endpoint", "Regla", "Estado", "Responsable", "Abierto"]]
        for inc in data["incidents"]:
            rows.append([
                inc["code"], Paragraph(inc["title"], styles["Normal"]), inc["hostname"],
                inc["rule_label"], inc["status_label"], inc["assigned_to_name"],
                inc["opened_at"].strftime("%d/%m/%Y")
            ])
        story.append(_alfa_table(rows, col_widths=[2 * cm, 4 * cm, 2.7 * cm, 3 * cm, 2.3 * cm, 2.5 * cm, 2 * cm]))
    else:
        story.append(Paragraph("No se abrieron incidentes en el período evaluado.", styles["Normal"]))

    doc.build(story)
    buffer.seek(0)
    return buffer


# --- Generadores de XLSX (openpyxl) ---

def _build_security_xlsx(data, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Seguridad"
    ws.append([meta["title"]])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([meta["subtitle"]])
    ws.append([])

    sev_rows = [[k, data["severity_counts"].get(k, 0)] for k in ("CRÍTICO", "ALTO", "MEDIO")]
    sev_rows.append(["Total", data["total_alerts"]])
    _xlsx_section(ws, "Alertas por severidad", ["Severidad", "Cantidad"], sev_rows)

    inc_rows = [[INCIDENT_STATUS_LABELS_ES.get(s, s), n] for s, n in data["incident_status_counts"].items()]
    _xlsx_section(ws, "Incidentes por estado", ["Estado", "Cantidad"], inc_rows or [["Sin incidentes en el período", ""]])

    if data["classification_counts"]:
        cls_rows = [[INCIDENT_CLASSIFICATION_LABELS_ES.get(c, c), n] for c, n in data["classification_counts"].items()]
        _xlsx_section(ws, "Clasificación de incidentes cerrados", ["Clasificación", "Cantidad"], cls_rows)

    rule_rows = [[name or "Sin regla vinculada", n] for name, n in data["rule_counts"]]
    _xlsx_section(ws, "Reglas más activas", ["Regla", "Alertas disparadas"], rule_rows or [["Sin alertas en el período", ""]])

    coverage_pct = round((data["honeyfiles_active"] + data["honeyfiles_triggered"]) / data["honeyfiles_total"] * 100, 1) if data["honeyfiles_total"] else 0
    hf_rows = [
        ["Honeyfiles desplegados", data["honeyfiles_total"]],
        ["Activos (intactos)", data["honeyfiles_active"]],
        ["Activados (comprometidos)", data["honeyfiles_triggered"]],
        ["Cobertura operativa (%)", coverage_pct],
    ]
    _xlsx_section(ws, "Cobertura de honeyfiles", None, hf_rows)

    minutes = round(data["avg_resolution_seconds"] / 60, 1) if data["avg_resolution_seconds"] is not None else "Sin datos aún"
    _xlsx_section(ws, "Tiempo de resolución", None, [["Promedio (minutos)", minutes]])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 32

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_endpoints_xlsx(data, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Endpoints"
    ws.append([meta["title"]])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([meta["subtitle"]])
    ws.append([])

    resumen_rows = [
        ["Endpoints evaluados", data["total_endpoints"]],
        ["En línea al momento de generar", data["online_count"]],
        ["Eventos totales en el período", data["total_events"]],
    ]
    _xlsx_section(ws, "Resumen", None, resumen_rows)

    ep_rows = [
        [ep["hostname"], ep["ip_address"], ep["os"] or "—", ep["status_label"], ep["event_count"], ep["honeyfiles_deployed"], ep["honeyfiles_pending"]]
        for ep in data["endpoints"]
    ]
    _xlsx_section(
        ws, "Detalle por endpoint",
        ["Endpoint", "IP", "SO", "Conectividad", "Eventos", "Honeyfiles desplegados", "Honeyfiles pendientes"],
        ep_rows or [["Sin endpoints que coincidan con el filtro", "", "", "", "", "", ""]]
    )

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 22

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _build_incidents_xlsx(data, meta):
    wb = Workbook()
    ws = wb.active
    ws.title = "Incidentes"
    ws.append([meta["title"]])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([meta["subtitle"]])
    ws.append(["Resumen de auditoría -- para traza forense completa de un incidente puntual, usar su Reporte PDF individual."])
    ws.append([])

    resumen_rows = [
        ["Incidentes en el período", data["total_incidents"]],
        ["Cerrados", data["closed_count"]],
        ["Abiertos / en curso", data["open_count"]],
    ]
    _xlsx_section(ws, "Resumen", None, resumen_rows)

    inc_rows = [
        [
            inc["code"], inc["title"], inc["hostname"], inc["rule_label"], inc["status_label"],
            inc["classification_label"], inc["assigned_to_name"],
            inc["opened_at"].strftime("%d/%m/%Y %H:%M"),
            inc["closed_at"].strftime("%d/%m/%Y %H:%M") if inc["closed_at"] else "Sigue abierto",
            inc["risk_score"]
        ]
        for inc in data["incidents"]
    ]
    _xlsx_section(
        ws, "Detalle de incidentes",
        ["ID", "Título", "Endpoint", "Regla", "Estado", "Clasificación", "Responsable", "Abierto", "Cerrado", "Riesgo máx."],
        inc_rows or [["Sin incidentes en el período", "", "", "", "", "", "", "", "", ""]]
    )

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


REPORT_BUILDERS = {
    ("SECURITY", "PDF"): _build_security_pdf,
    ("SECURITY", "XLSX"): _build_security_xlsx,
    ("ENDPOINTS", "PDF"): _build_endpoints_pdf,
    ("ENDPOINTS", "XLSX"): _build_endpoints_xlsx,
    ("INCIDENTS", "PDF"): _build_incidents_pdf,
    ("INCIDENTS", "XLSX"): _build_incidents_xlsx,
}


@app.post("/reportes/generar")
def generar_reporte(payload: ReportGenerate, user: dict = Depends(get_current_user)):
    """Genera el informe en el momento (PDF con reportlab o XLSX con
    openpyxl), lo guarda en disco (server/generated_reports/, no en la
    base -- ver comentario en 'reports' en database/schema.sql) e
    inserta la fila de auditoría. 'generated_by' siempre es quien tiene
    la sesión activa; no existe generación automática/por sistema hoy."""

    if payload.report_type not in REPORT_TYPE_LABELS_ES:
        raise HTTPException(status_code=422, detail=f"Tipo de informe desconocido: '{payload.report_type}'")

    if payload.format not in ("PDF", "XLSX"):
        raise HTTPException(status_code=422, detail=f"Formato desconocido: '{payload.format}'")

    start, end, period_label = _resolve_report_period(payload.period)

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            endpoint_hostname = None
            if payload.endpoint_id is not None:
                cursor.execute("SELECT hostname FROM endpoints WHERE id = %s;", (payload.endpoint_id,))
                ep_row = cursor.fetchone()
                if ep_row is None:
                    raise HTTPException(status_code=404, detail="Endpoint no encontrado")
                endpoint_hostname = ep_row[0]

            data = REPORT_DATA_GATHERERS[payload.report_type](cursor, start, end, payload.endpoint_id)

            type_label = REPORT_TYPE_LABELS_ES[payload.report_type]
            title = f"{type_label} - {period_label}"
            if endpoint_hostname:
                title += f" - {endpoint_hostname}"

            subtitle = (
                f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')} por {user.get('full_name', 'usuario')} "
                f"&mdash; Período evaluado: {start.strftime('%d/%m/%Y')} al {end.strftime('%d/%m/%Y')} "
                f"&mdash; {('Endpoint: ' + endpoint_hostname) if endpoint_hostname else 'Todos los endpoints'}"
            )
            meta = {"title": title, "subtitle": subtitle}

            buffer = REPORT_BUILDERS[(payload.report_type, payload.format)](data, meta)

            cursor.execute(
                """
                INSERT INTO reports (
                    title, report_type, format, period_label,
                    start_date, end_date, endpoint_id, generated_by, file_path
                )
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id, created_at;
                """,
                (title, payload.report_type, payload.format, period_label,
                 start, end, payload.endpoint_id, user["id"], "")
            )
            report_id, created_at = cursor.fetchone()

            os.makedirs(REPORTS_DIR, exist_ok=True)
            ext = "pdf" if payload.format == "PDF" else "xlsx"
            code = f"REP-{created_at.year}-{report_id:04d}"
            file_path = os.path.join(REPORTS_DIR, f"{code}.{ext}")

            with open(file_path, "wb") as f:
                f.write(buffer.getvalue())

            cursor.execute("UPDATE reports SET file_path = %s WHERE id = %s;", (file_path, report_id))

            connection.commit()

    finally:
        connection.close()

    return {
        "message": "Informe generado",
        "report": {
            "id": report_id,
            "code": code,
            "title": title,
            "report_type_label": type_label,
            "format": payload.format,
            "period_label": period_label,
            "endpoint": endpoint_hostname or "Todos los endpoints",
            "generated_by": user.get("full_name", "usuario"),
            "created_at": created_at.strftime("%d/%m/%Y %H:%M")
        }
    }


@app.get("/api/reportes")
def api_reportes(page: int = Query(1, ge=1), user: dict = Depends(get_current_user)):
    """Datos de la pantalla Reports en React -- consulta sobre
    'reports'. La generación (POST /reportes/generar) y la descarga
    (GET /reportes/{id}/archivo) son endpoints propios, no se duplican
    acá."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM reports;")
            total_reports = cursor.fetchone()[0]

            cursor.execute(
                """
                SELECT reports.created_at, users.full_name
                FROM reports
                LEFT JOIN users ON users.id = reports.generated_by
                ORDER BY reports.created_at DESC
                LIMIT 1;
                """
            )
            last_row = cursor.fetchone()
            last_generated_at, last_generated_by = last_row if last_row else (None, None)

            cursor.execute("SELECT id, hostname FROM endpoints ORDER BY hostname;")
            endpoint_options = [{"id": r[0], "hostname": r[1]} for r in cursor.fetchall()]

            page_size = 20
            total_pages = max(1, -(-total_reports // page_size))
            current_page = min(page, total_pages)
            offset = (current_page - 1) * page_size

            cursor.execute(
                """
                SELECT reports.id, reports.title, reports.report_type, reports.format,
                       reports.period_label, reports.created_at, endpoints.hostname,
                       users.full_name
                FROM reports
                LEFT JOIN endpoints ON endpoints.id = reports.endpoint_id
                LEFT JOIN users ON users.id = reports.generated_by
                ORDER BY reports.created_at DESC
                LIMIT %s OFFSET %s;
                """,
                (page_size, offset)
            )
            rows = cursor.fetchall()
    finally:
        connection.close()

    history = [
        {
            "id": r[0],
            "code": f"REP-{r[5].year}-{r[0]:04d}",
            "title": r[1],
            "report_type": r[2],
            "report_type_label": REPORT_TYPE_LABELS_ES.get(r[2], r[2]),
            "format": r[3],
            "period_label": r[4],
            "created_at": r[5].strftime("%d/%m/%Y %H:%M"),
            "endpoint": r[6] or "Todos los endpoints",
            "generated_by": r[7] or "Usuario eliminado",
        }
        for r in rows
    ]

    return {
        "total_reports": total_reports,
        "last_generated_at": last_generated_at.strftime("%d/%m/%Y %H:%M") if last_generated_at else None,
        "last_generated_by": last_generated_by,
        "endpoint_options": endpoint_options,
        "report_type_options": [{"value": k, "label": v} for k, v in REPORT_TYPE_LABELS_ES.items()],
        "period_options": [{"value": k, "label": v} for k, v in REPORT_PERIOD_OPTIONS],
        "history": history,
        "page": current_page,
        "page_size": page_size,
        "total_pages": total_pages,
    }


@app.get("/reportes/{report_id}/archivo")
def descargar_reporte(report_id: int, request: Request, disposition: str = Query("attachment")):
    """Sirve el archivo ya generado y guardado en disco -- no regenera
    con datos más nuevos, para que la copia descargada sea siempre la
    misma que quedó auditada en 'reports' en el momento de generarla."""

    user = require_session_user(request)

    if user is None:
        return RedirectResponse(url="/login", status_code=302)

    if disposition not in ("inline", "attachment"):
        disposition = "attachment"

    connection = get_connection()

    try:
        with connection.cursor() as cursor:
            cursor.execute(
                "SELECT format, file_path, created_at FROM reports WHERE id = %s;",
                (report_id,)
            )
            row = cursor.fetchone()

    finally:
        connection.close()

    if row is None:
        raise HTTPException(status_code=404, detail="Informe no encontrado")

    fmt, file_path, created_at = row

    if not file_path or not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="El archivo de este informe ya no está disponible en el servidor.")

    with open(file_path, "rb") as f:
        content = f.read()

    ext = "pdf" if fmt == "PDF" else "xlsx"
    code = f"REP-{created_at.year}-{report_id:04d}"

    return Response(
        content=content,
        media_type=REPORT_FORMAT_MEDIA_TYPES.get(fmt, "application/octet-stream"),
        headers={"Content-Disposition": f'{disposition}; filename="{code}.{ext}"'}
    )


# Etiquetas de 'audit_logs.action' -- strings libres a propósito
# (columna VARCHAR sin CHECK), cada endpoint que llama a log_audit()
# manda el código que quiere; este diccionario solo traduce lo que ya
# se está mandando hoy. Una acción que llegue sin traducción se
# muestra tal cual (nunca se oculta una fila de auditoría por no
# tener label).
AUDIT_ACTION_LABELS_ES = {
    "UPDATE_RULE": "Regla heurística modificada",
    "UPDATE_SETTING": "Parámetro global modificado",
    "CREATE_USER": "Usuario creado",
    "UPDATE_USER": "Usuario modificado",
    "ASSIGN_INCIDENT": "Responsable asignado",
    "UNASSIGN_INCIDENT": "Responsable removido",
    "UPDATE_INCIDENT_STATUS": "Estado de incidente cambiado",
    "UPDATE_ALERT_STATUS": "Estado de alerta cambiado",
}

CONFIGURACION_AUDIT_PAGE_SIZE = 30


@app.get("/api/config/agentes")
def api_config_agentes(user: dict = Depends(get_current_user)):
    """Datos de la subsección Configuración > Agentes de
    Administración en React. Solo
    'agent_stale_seconds' es un parámetro real editable (ver
    PATCH /settings/{key}, whitelist KNOWN_SETTINGS) -- heartbeat e
    intervalo de sincronización de reglas no son parámetros
    configurables porque no existe ningún mecanismo que los consuma
    (el agente es un script de una sola pasada, sin bucle)."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            agent_stale_seconds = get_agent_stale_seconds(cursor)
    finally:
        connection.close()

    return {"agent_stale_seconds": agent_stale_seconds}


@app.get("/api/audit-logs")
def api_audit_logs(page: int = Query(1, ge=1), user: dict = Depends(get_current_user)):
    """Datos de la subsección Configuración > Registro de actividad de
    Administración en React -- consulta sobre 'audit_logs', poblada
    desde 2026-08-12
    por log_audit() en los puntos reales donde se llama (ver
    AUDIT_ACTION_LABELS_ES para la lista completa de acciones que de
    verdad quedan registradas)."""

    connection = get_connection()
    try:
        with connection.cursor() as cursor:
            cursor.execute("SELECT COUNT(*) FROM audit_logs;")
            audit_total = cursor.fetchone()[0]

            total_pages = max(1, -(-audit_total // CONFIGURACION_AUDIT_PAGE_SIZE))
            current_page = min(page, total_pages)
            offset = (current_page - 1) * CONFIGURACION_AUDIT_PAGE_SIZE

            cursor.execute(
                """
                SELECT audit_logs.created_at, users.full_name, audit_logs.action,
                       audit_logs.entity_type, audit_logs.entity_id, audit_logs.description
                FROM audit_logs
                LEFT JOIN users ON users.id = audit_logs.user_id
                ORDER BY audit_logs.created_at DESC
                LIMIT %s OFFSET %s;
                """,
                (CONFIGURACION_AUDIT_PAGE_SIZE, offset)
            )
            rows = cursor.fetchall()
    finally:
        connection.close()

    entries = [
        {
            "created_at": r[0].strftime("%d/%m/%Y %H:%M:%S") if r[0] else None,
            "user_name": r[1] or "Usuario eliminado",
            "action": r[2],
            "action_label": AUDIT_ACTION_LABELS_ES.get(r[2], r[2]),
            "entity_type": r[3],
            "entity_id": r[4],
            "description": r[5],
        }
        for r in rows
    ]

    return {
        "entries": entries,
        "total": audit_total,
        "page": current_page,
        "page_size": CONFIGURACION_AUDIT_PAGE_SIZE,
        "total_pages": total_pages,
    }


@app.post("/agents")
def register_agent(agent: AgentCreate):
    # 'endpoints' (el host físico) y 'agents' (la instalación del
    # agente en ese host) son tablas separadas desde la reestructuración
    # a alfa_sentinel -- hay que crear una fila en cada una.

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO endpoints (hostname, os, os_version, ip_address)
                VALUES (%s, %s, %s, %s)
                RETURNING id;
                """,
                (agent.hostname, agent.os, agent.os_version, agent.ip_address)
            )

            endpoint_id = cursor.fetchone()[0]

            cursor.execute(
                """
                INSERT INTO agents (endpoint_id, agent_version)
                VALUES (%s, %s)
                RETURNING id;
                """,
                (endpoint_id, agent.agent_version or "desconocido")
            )

            agent_id = cursor.fetchone()[0]

            connection.commit()

        return {
            "message": "Agente registrado correctamente",
            "agent_id": agent_id
        }

    finally:
        connection.close()


@app.post("/enrollment-tokens")
def create_enrollment_token(
    user: dict = Depends(require_role("admin"))
):
    # Antes esto recibía "created_by" como un número suelto en la URL
    # -- cualquiera que supiera la ruta podía generar tokens a nombre
    # de cualquier usuario, sin loguearse. Ahora exige sesión con rol
    # 'admin', y el id sale de la sesión, no de lo que mande quien
    # llama.
    created_by = user["id"]

    token = secrets.token_urlsafe(32)

    token_hash = hashlib.sha256(
        token.encode()
    ).hexdigest()

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            cursor.execute(
                """
                INSERT INTO enrollment_tokens (
                    token_hash,
                    created_by,
                    expires_at
                )
                VALUES (
                    %s,
                    %s,
                    CURRENT_TIMESTAMP + INTERVAL '15 minutes'
                )
                RETURNING id, expires_at;
                """,
                (
                    token_hash,
                    created_by
                )
            )

            result = cursor.fetchone()

            connection.commit()

        return {
            "message": "Token de enrollment creado",
            "token": token,
            "token_id": result[0],
            "expires_at": result[1]
        }

    finally:
        connection.close()


@app.post("/enrollment")
def enroll_agent(enrollment: EnrollmentRequest):

    token_hash = hashlib.sha256(
        enrollment.token.encode()
    ).hexdigest()

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            # 1. Buscar el token
            cursor.execute(
                """
                SELECT id
                FROM enrollment_tokens
                WHERE token_hash = %s
                  AND status = 'ACTIVE'
                  AND used_at IS NULL
                  AND expires_at > CURRENT_TIMESTAMP;
                """,
                (token_hash,)
            )

            token_record = cursor.fetchone()

            if token_record is None:
                raise HTTPException(
                    status_code=401,
                    detail="Token de enrollment inválido o expirado"
                )

            token_id = token_record[0]

            # 2. Registrar el endpoint (host físico) y el agente
            # (instalación en ese host) -- son tablas separadas desde
            # la reestructuración a alfa_sentinel.
            cursor.execute(
                """
                INSERT INTO endpoints (hostname, os, os_version, ip_address)
                VALUES (%s, %s, %s, %s)
                RETURNING id;
                """,
                (
                    enrollment.hostname,
                    enrollment.os,
                    enrollment.os_version,
                    enrollment.ip_address
                )
            )

            endpoint_id = cursor.fetchone()[0]

            cursor.execute(
                """
                INSERT INTO agents (endpoint_id, agent_version)
                VALUES (%s, %s)
                RETURNING id;
                """,
                (endpoint_id, enrollment.agent_version or "desconocido")
            )

            agent_id = cursor.fetchone()[0]

            # 3. Generar credencial del agente
            credential = secrets.token_urlsafe(32)

            credential_hash = hashlib.sha256(
                credential.encode()
            ).hexdigest()

            cursor.execute(
                """
                INSERT INTO agent_credentials (
                    agent_id,
                    credential_hash
                )
                VALUES (%s, %s);
                """,
                (
                    agent_id,
                    credential_hash
                )
            )

            # 4. Marcar el token como utilizado
            cursor.execute(
                """
                UPDATE enrollment_tokens
                SET used_at = CURRENT_TIMESTAMP,
                    status = 'USED'
                WHERE id = %s;
                """,
                (token_id,)
            )

            connection.commit()

        return {
            "message": "Agente registrado correctamente",
            "agent_id": agent_id,
            "credential": credential
        }

    finally:
        connection.close()


@app.get("/agent/test")
def test_agent(x_agent_credential: str = Header(...)):

    connection = get_connection()

    try:

        with connection.cursor() as cursor:

            agent_id = resolve_agent_id(cursor, x_agent_credential)

            return {
                "message": "Agente autenticado correctamente",
                "agent_id": agent_id
            }

    finally:

        connection.close()


@app.post("/agent/heartbeat")
def agent_heartbeat(x_agent_credential: str = Header(...)):

    connection = get_connection()

    try:

        with connection.cursor() as cursor:

            agent_id = resolve_agent_id(cursor, x_agent_credential)

            cursor.execute(
                """
                UPDATE agents
                SET last_seen_at = CURRENT_TIMESTAMP,
                    status = 'ONLINE'
                WHERE id = %s;
                """,
                (agent_id,)
            )

            connection.commit()

            return {
                "message": "Heartbeat recibido",
                "agent_id": agent_id
            }

    finally:
        connection.close()


@app.post("/agent/events")
def report_event(
    event: EventCreate,
    x_agent_credential: str = Header(...)
):
    """Recibe eventos crudos del monitor de archivos/procesos del
    agente (tabla 'events')."""

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            agent_id = resolve_agent_id(cursor, x_agent_credential)

            # 'events.event_type' pasó a ser 'event_type_id' (FK a la
            # tabla catálogo 'event_types'). El agente sigue mandando
            # el nombre como texto -- se traduce acá. Si no matchea
            # ningún catálogo, se rechaza en vez de guardar un evento
            # con un tipo inventado.
            cursor.execute(
                "SELECT id FROM event_types WHERE name = %s;",
                (event.event_type,)
            )

            event_type_row = cursor.fetchone()

            if event_type_row is None:
                raise HTTPException(
                    status_code=422,
                    detail=f"Tipo de evento desconocido: '{event.event_type}'"
                )

            event_type_id = event_type_row[0]

            cursor.execute(
                """
                INSERT INTO events (
                    agent_id,
                    event_type_id,
                    process_id,
                    process_name,
                    file_path
                )
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id;
                """,
                (
                    agent_id,
                    event_type_id,
                    event.process_id,
                    event.process_name,
                    (event.metadata or {}).get("file_path")
                )
            )

            event_id = cursor.fetchone()[0]

            connection.commit()

        return {
            "message": "Evento registrado",
            "event_id": event_id
        }

    finally:
        connection.close()


# ------------------------------------------------------------------
# Motor heurístico -- lado servidor (2026-08-16, ver PENDIENTES.md,
# "Motor de reglas heurísticas -- especificación definitiva").
#
# "Mismo episodio": una ráfaga de eventos del mismo agente se agrupa
# en una sola alerta en vez de crear una nueva por cada evento
# (sección 27 -- "no crear cientos de alertas idénticas"). Se
# considera mismo episodio una alerta del mismo agente que sigue
# NEW/ACKNOWLEDGED (no cerrada/descartada) y cuya evidencia más
# reciente (MAX(alert_rule.matched_at), no 'alerts.created_at' --
# corregido 2026-08-17, ver PENDIENTES.md) llegó hace menos de
# EPISODE_WINDOW_SECONDS. Es DESLIZANTE a propósito: mientras sigan
# llegando indicadores compatibles, el episodio se mantiene abierto sin
# importar cuánto dure en total -- solo se considera cerrado (y la
# próxima evidencia abre uno nuevo) cuando pasan EPISODE_WINDOW_SECONDS
# reales sin ninguna coincidencia nueva. 120s es una decisión de
# producto razonable (no hay un valor "correcto" único) -- generoso
# respecto a las ventanas de las reglas individuales (10-20s) para que
# una ráfaga que dispara varias reglas en sucesión siga cayendo en la
# misma alerta.
EPISODE_WINDOW_SECONDS = 120

# Reglas "fuertes" -- peso >= 15, es decir todas menos las señales
# secundarias explícitas de la especificación (HR-06 CPU, HR-08
# archivos temporales) y las contextuales de menor peso (HR-10, y las
# diferidas HR-05/11). Se usan para decidir "evidencia fuerte" en la
# condición de incidente (sección 28) y de aislamiento (sección 30).
STRONG_RULE_NAMES = {
    "Modificacion Masiva Archivos",
    "Renombrado Extension Anomala",
    "Acceso Honeyfile",
    "Escritura Intensiva Archivos",
    "Acceso Recursos Compartidos",
    "Eliminacion Anomala Archivos",
}

# Igual que STRONG_RULE_NAMES pero sin 'Acceso Honeyfile' -- para la
# Condición A de aislamiento (sección 30: "honeyfile + al menos un
# indicador FUERTE DE ACTIVIDAD DE ARCHIVOS", el honeyfile no cuenta
# como su propio segundo indicador).
STRONG_FILE_ACTIVITY_RULES = STRONG_RULE_NAMES - {"Acceso Honeyfile"}


def sort_contributing_rules(rows):
    """Orden de relevancia para las reglas/señales que contribuyeron a
    una alerta o incidente (sección 3 de "Corrección definitiva en la
    lógica y presentación de ALERTAS", 2026-08-18, ver PENDIENTES.md):
    1) las más específicas/críticas primero -- 'Acceso Honeyfile'
       siempre al tope (sección 7: "señal especialmente relevante"),
       después el resto de STRONG_RULE_NAMES (ya existente, ver arriba,
       reutilizado tal cual en vez de inventar una segunda noción de
       "regla fuerte"), después el resto;
    2) dentro de cada nivel, mayor 'weight_applied' primero;
    3) en caso de empate, la coincidencia más reciente ('matched_at')
       primero.
    Reemplaza cualquier 'ORDER BY' que dependiera del orden en que
    Postgres decide devolver filas empatadas (nunca garantizado) o de
    'matched_at ASC' simple (que no distingue relevancia, solo
    cronología). No altera los pesos guardados -- solo el orden en que
    se muestran.

    'rows' es una lista de tuplas (rule_name, weight_applied, matched_at)
    -- 'matched_at' puede ser None (no debería pasar en la práctica,
    pero se maneja sin romper por las dudas)."""

    def _tier(rule_name):
        if rule_name == "Acceso Honeyfile":
            return 0
        if rule_name in STRONG_RULE_NAMES:
            return 1
        return 2

    def _key(row):
        rule_name, weight_applied, matched_at = row
        return (
            _tier(rule_name),
            -float(weight_applied),
            -(matched_at.timestamp() if matched_at else 0.0),
        )

    return sorted(rows, key=_key)

# HR-05/06/11 -- hasta el 2026-08-16 estaban acá porque requerían
# datos que el agente no recopilaba (atribución de proceso a evento
# de archivo, muestreo de CPU por proceso). Esa capacidad ya existe
# (agent/adapters/, agent/process_monitor.py -- ver PENDIENTES.md,
# "Implementación final del motor heurístico y configuración por
# endpoint"), así que el set queda vacío -- se conserva la variable
# (en vez de borrarla y sacar las 2 validaciones que la usan) porque
# el mecanismo de "esta regla no se puede activar todavía" sigue
# siendo válido para el día que se agregue una regla 13 que dependa
# de una capacidad que el agente aún no tenga.
DEFERRED_RULE_NAMES = set()

# Reglas cuyo weight/threshold/window_seconds NO deben editarse desde
# PATCH /rules/{id} porque no funcionan como una regla convencional de
# puntuación (ver sección 13 de la especificación de la pantalla de
# Reglas Heurísticas, 2026-08-16):
# - 'Acceso Honeyfile': su weight=100 es lo que hace que CUALQUIER
#   interacción con un honeyfile llegue a risk_score=100 "gratis" (ver
#   report_alert, MIN(100, suma_pesos + correlación)) -- cambiarlo
#   rompería esa garantía sin que se note hasta la próxima alerta real.
# - 'Correlacion Multiples Indicadores': su weight/threshold en la base son
#   solo documentales -- report_alert calcula la bonificación real por
#   tramos fijos (2/3/4+ reglas -> +5/+10/+15), no lee estas columnas.
#   Editarlas no cambiaría ningún cálculo, así que se bloquea para no
#   sugerir un control que no hace nada.
FIXED_SCORING_RULE_NAMES = {"Acceso Honeyfile", "Correlacion Multiples Indicadores"}


@app.post("/agent/alerts")
def report_alert(
    alert: AlertCreate,
    x_agent_credential: str = Header(...)
):
    """El agente ya NO calcula severidad/score -- solo reporta qué
    reglas detectó activas (alert.matched_rules). Este endpoint es el
    responsable único de: (1) buscar el peso real de cada regla en
    'heuristic_rules', (2) registrar trazabilidad completa en
    'alert_rule' -- una fila por regla que participó, sin duplicar
    evidencia ya registrada (secciones 24 y 26), (3) calcular la
    bonificación de correlación HR-12 (sección 21), (4) sumar y acotar
    a 100 (sección 8; HR-03/honeyfile llega a 100 'gratis' porque su
    weight en la base ya es 100, sin necesidad de un caso especial acá),
    (5) derivar la severidad consultando 'severity_levels' (sección 3),
    (6) decidir si corresponde crear un incidente automáticamente
    (sección 28) y (7) evaluar si se cumple la condición de aislamiento
    (sección 30) y, si corresponde, ORDENARLO -- corregido 2026-08-17
    (ver PENDIENTES.md, "Corrección definitiva del motor heurístico...
    "): ya no se queda en una recomendación sin ejecutar; deja una
    orden real en 'host_isolations' (status='REQUESTED') que el agente
    de ese endpoint recoge y ejecuta de verdad (agent/isolation_sync.py
    + agent/isolation_executor.py, real solo en producción y con
    privilegios; en desarrollo, el flujo completo se ejerce igual pero
    la acción de red queda simulada, nunca toca el firewall real de
    quien está probando el sistema)."""

    if not alert.matched_rules:
        raise HTTPException(status_code=422, detail="matched_rules no puede estar vacío")

    connection = get_connection()

    try:
        with connection.cursor() as cursor:

            agent_id = resolve_agent_id(cursor, x_agent_credential)

            # Solo reglas cuyo is_active EFECTIVO (heuristic_rules +
            # override de agent_rule para ESTE agente, ver
            # _effective_agent_rules_cte()) sea TRUE -- si el agente
            # reportara el nombre de una regla diferida o desconocida,
            # se ignora en vez de inventarle un peso. El peso aplicado
            # ('effective_weight') también respeta el override por
            # endpoint si existe -- el agente nunca manda su propio
            # weight, y aunque lo mandara, se ignoraría (sección 16 de
            # la especificación de configuración por endpoint).
            cursor.execute(
                _effective_agent_rules_cte() + """
                SELECT id, name, effective_weight FROM effective_rules
                WHERE name = ANY(%(names)s) AND effective_is_active = TRUE;
                """,
                {"agent_id": agent_id, "names": alert.matched_rules}
            )
            matched = cursor.fetchall()

            if not matched:
                raise HTTPException(
                    status_code=422,
                    detail="Ninguna de las reglas reportadas es una regla activa conocida"
                )

            is_honeyfile = any(name == "Acceso Honeyfile" for _, name, _ in matched)

            # ¿Actualiza una alerta existente del mismo episodio, o
            # crea una nueva? -- corregido 2026-08-17 (ver PENDIENTES.md,
            # "Corrección definitiva del motor heurístico..."): ANTES
            # comparaba contra 'alerts.created_at', una ventana FIJA
            # desde que se creó la alerta -- un ataque sostenido que
            # sigue mandando evidencia real más allá de esos 120s
            # (ej. cada 90s durante 10 minutos) se fragmentaba en varias
            # alertas nuevas en vez de seguir siendo el mismo episodio,
            # aunque nunca hubo un hueco real de inactividad. Ahora la
            # ventana es DESLIZANTE: se mide desde la última evidencia
            # real vinculada a la alerta (MAX(alert_rule.matched_at)),
            # no desde su creación -- "un episodio permanece abierto
            # mientras sigan llegando indicadores compatibles", y solo
            # se considera cerrado cuando pasan 120s sin ninguna
            # evidencia nueva. Con evidencia continua, un episodio puede
            # durar mucho más de 120s en total y seguir siendo UNO solo.
            cursor.execute(
                """
                SELECT alerts.id
                FROM alerts
                LEFT JOIN alert_rule ON alert_rule.alert_id = alerts.id
                WHERE alerts.agent_id = %s
                  AND alerts.status IN ('NEW', 'ACKNOWLEDGED')
                GROUP BY alerts.id, alerts.created_at
                HAVING GREATEST(alerts.created_at, COALESCE(MAX(alert_rule.matched_at), alerts.created_at))
                       >= NOW() - (%s || ' seconds')::INTERVAL
                ORDER BY GREATEST(alerts.created_at, COALESCE(MAX(alert_rule.matched_at), alerts.created_at)) DESC
                LIMIT 1;
                """,
                (agent_id, EPISODE_WINDOW_SECONDS)
            )
            existing = cursor.fetchone()
            alert_id = existing[0] if existing else None

            if alert_id is None:
                cursor.execute(
                    """
                    INSERT INTO alerts (agent_id, severity_id, title, description, risk_score)
                    VALUES (%s, (SELECT id FROM severity_levels ORDER BY min_score ASC LIMIT 1), %s, %s, 0)
                    RETURNING id;
                    """,
                    (agent_id, alert.title, alert.description)
                )
                alert_id = cursor.fetchone()[0]

            # Reglas ya vinculadas a esta alerta (sin contar la fila
            # sintética de correlación) -- para no duplicar evidencia
            # ya registrada si la misma regla vuelve a matchear en un
            # evento posterior del mismo episodio (sección 24).
            cursor.execute(
                """
                SELECT alert_rule.rule_id FROM alert_rule
                JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                WHERE alert_rule.alert_id = %s
                  AND heuristic_rules.name != 'Correlacion Multiples Indicadores';
                """,
                (alert_id,)
            )
            already_linked = {row[0] for row in cursor.fetchall()}

            for rule_id, _, weight in matched:
                if rule_id not in already_linked:
                    cursor.execute(
                        """
                        INSERT INTO alert_rule (alert_id, rule_id, weight_applied)
                        VALUES (%s, %s, %s);
                        """,
                        (alert_id, rule_id, weight)
                    )

            # Recalcular el score a partir de TODAS las reglas
            # vinculadas hasta ahora (evidencia acumulada del
            # episodio completo, no solo la de este request).
            cursor.execute(
                """
                SELECT heuristic_rules.id, heuristic_rules.name, alert_rule.weight_applied
                FROM alert_rule
                JOIN heuristic_rules ON heuristic_rules.id = alert_rule.rule_id
                WHERE alert_rule.alert_id = %s
                  AND heuristic_rules.name != 'Correlacion Multiples Indicadores';
                """,
                (alert_id,)
            )
            linked_rules = cursor.fetchall()

            base_score = sum(float(w) for _, _, w in linked_rules)
            linked_names = {name for _, name, _ in linked_rules}
            distinct_rule_count = len(linked_rules)

            # HR-12 -- bonificación de correlación (sección 21): NO es
            # una regla de conteo convencional, es una segunda capa.
            if distinct_rule_count >= 4:
                correlation_bonus = 15.0
            elif distinct_rule_count == 3:
                correlation_bonus = 10.0
            elif distinct_rule_count == 2:
                correlation_bonus = 5.0
            else:
                correlation_bonus = 0.0

            # Respeta el mismo interruptor is_active EFECTIVO que
            # cualquier otra regla (editable desde /configuracion a
            # nivel global, o por endpoint vía agent_rule -- sección 19
            # de la especificación: HR-12 solo admite override de
            # is_active, nunca de weight/threshold/window) -- si un
            # analista desactiva la correlación (global o para este
            # endpoint puntual), deja de sumar bonificación.
            cursor.execute(
                _effective_agent_rules_cte() + """
                SELECT id FROM effective_rules
                WHERE name = 'Correlacion Multiples Indicadores' AND effective_is_active = TRUE;
                """,
                {"agent_id": agent_id}
            )
            correlation_row = cursor.fetchone()
            correlation_rule_id = correlation_row[0] if correlation_row else None

            if correlation_rule_id is None:
                correlation_bonus = 0.0
                existing_correlation_row = None
            else:
                cursor.execute(
                    "SELECT id FROM alert_rule WHERE alert_id = %s AND rule_id = %s;",
                    (alert_id, correlation_rule_id)
                )
                existing_correlation_row = cursor.fetchone()

            if correlation_bonus > 0:
                if existing_correlation_row:
                    cursor.execute(
                        "UPDATE alert_rule SET weight_applied = %s, matched_at = CURRENT_TIMESTAMP WHERE id = %s;",
                        (correlation_bonus, existing_correlation_row[0])
                    )
                else:
                    cursor.execute(
                        """
                        INSERT INTO alert_rule (alert_id, rule_id, weight_applied)
                        VALUES (%s, %s, %s);
                        """,
                        (alert_id, correlation_rule_id, correlation_bonus)
                    )

            # Sección 8: nunca almacenar un score superior a 100.
            # HR-03 (honeyfile) llega a 100 sin caso especial: su
            # weight en 'heuristic_rules' ya es 100.
            final_score = min(100.0, base_score + correlation_bonus)

            cursor.execute(
                "SELECT id, name FROM severity_levels WHERE %s BETWEEN min_score AND max_score;",
                (final_score,)
            )
            severity_row = cursor.fetchone()
            severity_id, severity_name = severity_row if severity_row else (None, "BAJO")

            cursor.execute(
                """
                UPDATE alerts SET severity_id = %s, risk_score = %s, description = %s
                WHERE id = %s RETURNING incident_id;
                """,
                (severity_id, final_score, alert.description, alert_id)
            )
            incident_id = cursor.fetchone()[0]

            # Sección 28: CRÍTICO por score solo no alcanza -- hace
            # falta evidencia fuerte además (honeyfile, correlación de
            # al menos 3 reglas distintas, o al menos 2 reglas
            # "fuertes" distintas coincidiendo).
            strong_count = len(linked_names & STRONG_RULE_NAMES)
            meets_incident_condition = (
                final_score >= 75
                and (is_honeyfile or distinct_rule_count >= 3 or strong_count >= 2)
            )

            incident_created = False

            if incident_id is None and meets_incident_condition:
                cursor.execute(
                    """
                    INSERT INTO incidents (agent_id, title, description)
                    VALUES (%s, %s, %s)
                    RETURNING id;
                    """,
                    (
                        agent_id,
                        alert.title,
                        (
                            f"Creado automáticamente por el motor heurístico -- "
                            f"risk_score {final_score:.2f}, reglas: {', '.join(sorted(linked_names))}."
                        )
                    )
                )
                incident_id = cursor.fetchone()[0]

                cursor.execute(
                    "UPDATE alerts SET incident_id = %s WHERE id = %s;",
                    (incident_id, alert_id)
                )

                incident_created = True

            # Sección 30: la condición de aislamiento es DISTINTA de la
            # de incidente, y solo se evalúa una vez que existe un
            # incidente (la detección/contención del diagrama de la
            # especificación pasa primero por "¿corresponde incidente?").
            # Corregido 2026-08-17 (ver PENDIENTES.md): ANTES esto se
            # quedaba en una fila 'RECOMMENDED' sin ejecutar nada. Ahora
            # SÍ se ordena: se inserta como 'REQUESTED' y el agente de
            # ESE endpoint (agent/isolation_sync.py, polling periódico
            # vía GET /agent/isolation-status, mismo patrón que
            # honeyfile_sync.py) la recoge, la ejecuta de verdad
            # (agent/isolation_executor.py) y confirma el resultado real
            # vía POST /agent/isolation-status/report -- recién ahí la
            # fila pasa a 'EXECUTED' o 'ISOLATION_FAILED'. El servidor
            # nunca ejecuta nada él mismo (no tiene acceso a la red del
            # endpoint) -- solo ordena y registra el resultado.
            isolation_requested = False

            if incident_id is not None:
                strong_file_matched = linked_names & STRONG_FILE_ACTIVITY_RULES
                condition_a = is_honeyfile and len(strong_file_matched) >= 1
                condition_b = final_score >= 75 and len(strong_file_matched) >= 2

                if condition_a or condition_b:
                    # No duplicar una orden ya en curso o ya cumplida
                    # para este mismo ENDPOINT (mismo criterio de "no
                    # duplicar evidencia/órdenes ya registradas" que
                    # already_linked más arriba).
                    #
                    # BUG REAL corregido 2026-08-18 (problema H, ver
                    # PENDIENTES.md): filtraba por 'incident_id = %s' --
                    # el incidente de ESTE episodio puntual. Un agente con
                    # más de un incidente (ej. ya aislado por un episodio
                    # anterior) podía disparar esta condición de nuevo
                    # desde un incidente DISTINTO -- 'incident_id' nunca
                    # coincidía con la orden ya existente (que tiene el
                    # incident_id del episodio anterior), así que el
                    # servidor insertaba una SEGUNDA orden 'REQUESTED'
                    # automática para un endpoint que ya estaba aislado.
                    # Mismo criterio único que _agent_is_isolated_sql().
                    cursor.execute(
                        f"SELECT 1 WHERE {_agent_is_isolated_sql('%s')};",
                        (agent_id,)
                    )
                    if cursor.fetchone() is None:
                        reason = (
                            "Condición A: honeyfile activado + al menos un indicador fuerte de actividad de archivos "
                            f"({', '.join(sorted(strong_file_matched))})."
                            if condition_a else
                            "Condición B: severidad CRÍTICA + al menos dos indicadores fuertes de actividad maliciosa "
                            f"({', '.join(sorted(strong_file_matched))})."
                        )
                        cursor.execute(
                            """
                            INSERT INTO host_isolations (agent_id, incident_id, isolation_type, status, reason)
                            VALUES (%s, %s, 'NETWORK', 'REQUESTED', %s);
                            """,
                            (agent_id, incident_id, reason)
                        )
                        isolation_requested = True

            connection.commit()

        return {
            "message": "Alerta registrada",
            "alert_id": alert_id,
            "risk_score": final_score,
            "severity": severity_name,
            "incident_id": incident_id,
            "incident_created": incident_created,
            "isolation_requested": isolation_requested,
        }

    finally:
        connection.close()
